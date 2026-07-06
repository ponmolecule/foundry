"""Excel configuration I/O.

Bankers speak Excel, not JSON. This module renders an engagement
configuration as a structured workbook (writer) and converts an uploaded
workbook back into the canonical config dict (parser). The parser output
feeds the SAME validate_config -> register -> run pipeline as JSON uploads;
Excel is a front-end format, never a second code path.

Workbook contract (sheet names are fixed; the parser fails closed on
missing required sheets):
  Engagement       field | value            identity, narrative, license
  Modules          module | loaded (yes/no)
  Constraints      key | value | text | source
  Target State     field | value
  Peer Query       field | value            (+ prior_metrics rows)
  Assumptions      key | value | confidence | prior_metric
  Marketing Budget month | amount           (funnel-channel banks)
  Loan Segments    one row per segment      (commercial banks)
  Entity Map       field/flag rows          (step 0A)
"""
import io
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from .configio import ConfigError
from .modules import REGISTRY

NAVY = "0D1626"; GOLD = "DFA85A"; PANEL = "182640"
HDR = Font(name="Arial", bold=True, color=GOLD, size=10)
KEY = Font(name="Arial", size=10)
INPUT = Font(name="Arial", size=10, color="0000FF")   # blue = user-editable input
FILL = PatternFill("solid", start_color=NAVY)
NOTE = Font(name="Arial", size=8, italic=True, color="808080")


def _sheet(wb, name, headers, widths):
    ws = wb.create_sheet(name)
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=j, value=h)
        c.font = HDR; c.fill = FILL
        ws.column_dimensions[chr(64 + j)].width = widths[j - 1]
    ws.freeze_panes = "A2"
    return ws


def _kv(ws, rows, start=2):
    r = start
    for k, v in rows:
        ws.cell(row=r, column=1, value=k).font = KEY
        c = ws.cell(row=r, column=2, value=v); c.font = INPUT
        c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    return r


def workbook_from_config(cfg):
    wb = Workbook(); wb.remove(wb.active)

    ws = _sheet(wb, "Engagement", ["Field", "Value"], [28, 90])
    s1 = cfg.get("step_1", {}); s0 = cfg["step_0"]; s0a = cfg.get("step_0a", {})
    _kv(ws, [
        ("engagement_id", cfg["engagement_id"]),
        ("client_legal_name", cfg["client_legal_name"]),
        ("proposed_bank", cfg["proposed_bank"]),
        ("hq", cfg.get("hq", "")),
        ("prepared_by", cfg.get("prepared_by", "")),
        ("config_version", cfg["config_version"]),
        ("config_frozen", str(cfg["config_frozen"])),
        ("archetype", cfg.get("archetype", "digital_consumer")),
        ("one_sentence", s0["one_sentence"]),
        ("earning_engine", s0.get("earning_engine", "")),
        ("charter", s1.get("charter", "")),
        ("regulators", ", ".join(s1.get("regulators", []))),
        ("license_rationale", s1.get("rationale", "")),
        ("step_minus_1_decision", cfg.get("step_minus_1", {}).get("decision", "proceed_de_novo")),
        ("parent_entity", s0a.get("entities", {}).get("parent", "")),
        ("bank_entity", s0a.get("entities", {}).get("bank", "")),
        ("map_approved", str(s0a.get("map_approved", ""))),
    ])

    ws = _sheet(wb, "Modules", ["Module", "Loaded (yes/no)"], [30, 18])
    r = 2
    for m in sorted(REGISTRY):
        ws.cell(row=r, column=1, value=m).font = KEY
        c = ws.cell(row=r, column=2, value="yes" if m in s0["modules"] else "no"); c.font = INPUT
        r += 1

    ws = _sheet(wb, "Constraints", ["key", "value", "text", "source"], [26, 12, 60, 42])
    for i, c0 in enumerate(cfg["constraints"], 2):
        for j, k in enumerate(("key", "value", "text", "source"), 1):
            cc = ws.cell(row=i, column=j, value=c0[k]); cc.font = INPUT
            cc.alignment = Alignment(wrap_text=True, vertical="top")

    ws = _sheet(wb, "Target State", ["Field", "Value"], [26, 20])
    _kv(ws, list(cfg["target_state"].items()))

    ws = _sheet(wb, "Peer Query", ["Field", "Value"], [26, 22])
    r = _kv(ws, [(k, v) for k, v in cfg["peer_query"].items() if k != "log_assets_yr3"])
    ws.cell(row=r, column=1, value="prior_metrics").font = HDR
    r += 1
    for m in cfg.get("prior_metrics", []):
        ws.cell(row=r, column=1, value="metric").font = KEY
        ws.cell(row=r, column=2, value=m).font = INPUT
        r += 1

    ws = _sheet(wb, "Assumptions", ["key", "value", "confidence", "prior_metric"], [32, 16, 24, 26])
    tags = cfg.get("assumption_tags", {})
    r = 2
    for k, v in cfg["assumptions"].items():
        if k in ("marketing_budget_m", "loan_segments"):
            continue
        ws.cell(row=r, column=1, value=k).font = KEY
        ws.cell(row=r, column=2, value=v).font = INPUT
        t = tags.get(k)
        if t:
            ws.cell(row=r, column=3, value=t[0]).font = INPUT
            if t[1]:
                ws.cell(row=r, column=4, value=t[1]).font = INPUT
        r += 1
    dg = tags.get("deposit_growth_yr1")
    if dg:
        ws.cell(row=r, column=1, value="deposit_growth_yr1").font = KEY
        ws.cell(row=r, column=2, value="(derived by engine)").font = NOTE
        ws.cell(row=r, column=3, value=dg[0]).font = INPUT
        ws.cell(row=r, column=4, value=dg[1] or "").font = INPUT

    ws = _sheet(wb, "Marketing Budget", ["month", "amount ($)"], [10, 16])
    for i, v in enumerate(cfg["assumptions"].get("marketing_budget_m", []), 1):
        ws.cell(row=i + 1, column=1, value=i).font = KEY
        ws.cell(row=i + 1, column=2, value=v).font = INPUT

    seg_cols = ["name", "orig_per_lender_m", "ramp_m", "amort_annual", "yield",
                "nco_mature", "nco_ramp_m", "allowance_coverage", "avg_loan_size"]
    ws = _sheet(wb, "Loan Segments", seg_cols, [12] + [16] * 8)
    for i, seg in enumerate(cfg["assumptions"].get("loan_segments", []), 2):
        for j, k in enumerate(seg_cols, 1):
            ws.cell(row=i, column=j, value=seg[k]).font = INPUT

    ws = _sheet(wb, "Entity Map", ["id", "class", "text"], [10, 34, 80])
    for i, f in enumerate(cfg.get("step_0a", {}).get("flags_from_map", []), 2):
        for j, k in enumerate(("id", "class", "text"), 1):
            c = ws.cell(row=i, column=j, value=f[k]); c.font = INPUT
            c.alignment = Alignment(wrap_text=True, vertical="top")

    _dictionary_sheet(wb)
    return wb


# ---------------------------------------------------------------- parser

def _num(v):
    if isinstance(v, float) and v.is_integer():
        return int(v)
    return v


def _kv_read(ws):
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        out[str(row[0]).strip()] = row[1]
    return out


def parse_workbook(data: bytes):
    try:
        wb = load_workbook(io.BytesIO(data), data_only=True)
    except Exception as e:
        raise ConfigError(f"could not read workbook: {e}")
    need = ["Engagement", "Modules", "Constraints", "Target State",
            "Peer Query", "Assumptions"]
    missing = [n for n in need if n not in wb.sheetnames]
    if missing:
        raise ConfigError(f"workbook missing required sheet(s): {missing}")

    e = _kv_read(wb["Engagement"])
    def E(k, default=""):
        v = e.get(k, default)
        return "" if v is None else v

    modules = []
    for row in wb["Modules"].iter_rows(min_row=2, values_only=True):
        if row[0] and str(row[1] or "").strip().lower() in ("yes", "y", "true", "1"):
            modules.append(str(row[0]).strip())

    constraints = []
    for row in wb["Constraints"].iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        constraints.append({"key": str(row[0]).strip(), "value": _num(row[1]),
                            "text": row[2] or "", "source": row[3] or ""})

    target = {k: _num(v) for k, v in _kv_read(wb["Target State"]).items()}

    peer, prior_metrics = {}, []
    for row in wb["Peer Query"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        k = str(row[0]).strip()
        if k == "prior_metrics":
            continue
        if k == "metric":
            if row[1]:
                prior_metrics.append(str(row[1]).strip())
        else:
            peer[k] = _num(row[1])
    peer.setdefault("log_assets_yr3", None)

    assumptions, tags = {}, {}
    for row in wb["Assumptions"].iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        k = str(row[0]).strip()
        v = row[1]
        derived = isinstance(v, str) and "derived" in v
        if not derived:
            assumptions[k] = _num(v)
        if len(row) > 2 and row[2]:
            tags[k] = (str(row[2]).strip(), str(row[3]).strip() if len(row) > 3 and row[3] else None)

    if "Marketing Budget" in wb.sheetnames:
        mb = [(row[0], _num(row[1])) for row in wb["Marketing Budget"].iter_rows(min_row=2, values_only=True)
              if row[0] is not None and row[1] is not None]
        if mb:
            assumptions["marketing_budget_m"] = [v for _, v in sorted(mb)]

    if "Loan Segments" in wb.sheetnames:
        seg_cols = [c.value for c in wb["Loan Segments"][1]]
        segs = []
        for row in wb["Loan Segments"].iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            segs.append({k: (_num(v) if k != "name" else str(v).strip())
                         for k, v in zip(seg_cols, row)})
        if segs:
            assumptions["loan_segments"] = segs

    flags = []
    if "Entity Map" in wb.sheetnames:
        for row in wb["Entity Map"].iter_rows(min_row=2, values_only=True):
            if row[0]:
                flags.append({"id": str(row[0]).strip(), "class": str(row[1] or "advisory").strip(),
                              "text": row[2] or ""})

    cfg = {
        "engagement_id": E("engagement_id"),
        "client_legal_name": E("client_legal_name"),
        "proposed_bank": E("proposed_bank"),
        "hq": E("hq"),
        "prepared_by": E("prepared_by"),
        "config_version": str(E("config_version")),
        "config_frozen": str(E("config_frozen"))[:10],
        "archetype": E("archetype", "digital_consumer") or "digital_consumer",
        "step_minus_1": {"decision": E("step_minus_1_decision", "proceed_de_novo"),
                         "alternatives_priced": {}},
        "step_0": {"one_sentence": E("one_sentence"),
                   "earning_engine": E("earning_engine"),
                   "modules": modules},
        "step_0a": {"entities": {"parent": E("parent_entity"), "bank": E("bank_entity")},
                    "flags_from_map": flags,
                    "map_approved": str(E("map_approved"))[:10]},
        "step_1": {"charter": E("charter"),
                   "regulators": [x.strip() for x in str(E("regulators")).split(",") if x.strip()],
                   "rationale": E("license_rationale")},
        "constraints": constraints,
        "target_state": target,
        "peer_query": peer,
        "prior_metrics": prior_metrics,
        "assumptions": assumptions,
        "assumption_tags": tags,
    }
    return cfg


# ---------------------------------------------------------------- dictionary

# (sheet, field, type, units, required_for, range_allowed, description, example)
DICTIONARY = [
 ("Engagement","engagement_id","text","—","always","free text","Unique engagement identifier assigned by the advisory team; appears on every output and in the audit trail.","ENG-2026-0004"),
 ("Engagement","client_legal_name","text","—","always","free text","Legal name of the organizing entity (the applicant company, not the bank).","Prairie Digital Financial, Inc."),
 ("Engagement","proposed_bank","text","—","always","free text","Name of the bank in organization. Also used to generate the engagement's URL slug.","Prairie Digital Bank (in organization)"),
 ("Engagement","hq","text","—","optional","City, ST","Headquarters city of the proposed bank.","Des Moines, IA"),
 ("Engagement","prepared_by","text","—","optional","free text","Team responsible for this configuration.","Klaros Group / Foundry engagement team"),
 ("Engagement","config_version","text","—","always","e.g. 1.0","Version of this configuration. Increment on any change after freeze; version history is part of the exam record.","1.0"),
 ("Engagement","config_frozen","date","YYYY-MM-DD","always","valid date","Date this configuration was frozen. Doubles as rules_as_of in the run manifest.","2026-07-05"),
 ("Engagement","archetype","text","—","always","digital_consumer / community_commercial","Business-model archetype; selects the examiner-book template family.","digital_consumer"),
 ("Engagement","one_sentence","text","—","always","one sentence","Step 0: the bank's economics in one sentence. If management cannot produce this, the model is not the blocker.","We gather consumer deposits through the app..."),
 ("Engagement","earning_engine","text","—","optional","spread / fee / hybrid...","Primary earning engine classification.","hybrid_deposit_led"),
 ("Engagement","charter","text","—","optional","free text","Step 1 license decision (derived from the business model, not assumed).","Iowa state nonmember bank"),
 ("Engagement","regulators","text list","comma-separated","optional","free text","Chartering and insuring agencies.","Iowa Division of Banking, FDIC"),
 ("Engagement","license_rationale","text","—","optional","free text","Why this charter and not the alternatives.","Deposit-led consumer model; state charter matches footprint"),
 ("Engagement","step_minus_1_decision","text","—","optional","proceed_de_novo / remain_sponsor / acquire","Step -1: should this company become a bank at all.","proceed_de_novo"),
 ("Engagement","parent_entity","text","—","optional","free text","Step 0A: what the parent/holding company owns and does.","Parent - app, brand, marketing, 410k accounts"),
 ("Engagement","bank_entity","text","—","optional","free text","Step 0A: what the bank owns and does.","Bank - deposits, card program, securities"),
 ("Engagement","map_approved","date","YYYY-MM-DD","optional","valid date","Date the entity/funds-flow map was approved (a gate before calibration).","2026-06-25"),

 ("Modules","core_deposits","yes/no","—","deposit banks","yes or no","App/marketing-funnel consumer deposit engine: accounts from migration + acquisition spend; balances ramp over balance_ramp_months.","yes"),
 ("Modules","relationship_deposits","yes/no","—","community banks","yes or no","Banker-driven commercial deposit engine: relationships per banker per month, hiring ramp capped at bankers_max.","no"),
 ("Modules","revolving_credit","yes/no","—","card banks","yes or no","Consumer card book: penetration of active accounts, average balance, vintage-ramped losses, allowance coverage.","yes"),
 ("Modules","commercial_lending","yes/no","—","commercial banks","yes or no","CRE/C&I lending: originations per lender by segment (Loan Segments sheet), amortization, segment loss curves.","no"),
 ("Modules","payments_fees","yes/no","—","digital banks","yes or no","Debit interchange plus per-account fees. Triggers the Durbin exemption rule check.","yes"),
 ("Modules","relationship_fees","yes/no","—","community banks","yes or no","Commercial service charges and treasury-management fees on relationship deposits.","no"),
 ("Modules","capacity_expenses","yes/no","—","digital banks","yes or no","Digital ops staffing from demand: KYC reviews, fraud alerts, service contacts converted to FTE.","yes"),
 ("Modules","branch_capacity_expenses","yes/no","—","community banks","yes or no","Branch/credit-admin staffing from demand: onboarding, relationship service, credit administration per loan.","no"),
 ("Modules","investment_portfolio","yes/no","—","always","yes","Securities book as the funding-waterfall residual (chassis-level). Load for every bank.","yes"),

 ("Constraints","key","text","—","always","leverage_min / card_receivables_max_share / cre_max_pct_capital / brokered_max_share / marketing_linkage","Machine-readable constraint identifier. leverage_min is REQUIRED for every engagement. Keys with evaluators are tested in every scenario; structural keys are attested via flags.","leverage_min"),
 ("Constraints","value","number","ratio","evaluated keys","0-1 for ratios; multiples for /capital caps","The binding limit. leverage_min: minimum tier 1 leverage. card_receivables_max_share: max card/assets. cre_max_pct_capital: max CRE/tier 1 (e.g. 3.5 = 350%).","0.095"),
 ("Constraints","text","text","—","always","free text","Human-readable statement of the constraint as the regulator expressed it.","Tier 1 leverage >= 9.5% through year 3"),
 ("Constraints","source","text","—","always","document reference","Provenance pointer to the pre-filing document. Every constraint must trace to a named source.","FDIC pre-filing memo 2026-06-28"),

 ("Target State","assets_yr3","number","$","always","> 0","Stated intent: total assets at end of year 3. Used in peer selection (log-scaled); the model's own trajectory is compared back to it.","430000000"),
 ("Target State","deposits_yr3","number","$","optional","> 0","Stated intent: total deposits at end of year 3.","360000000"),
 ("Target State","card_receivables_yr3","number","$","card banks","> 0","Stated intent: card receivables at year 3 (digital/card banks).","95000000"),
 ("Target State","loans_yr3","number","$","commercial banks","> 0","Stated intent: total loans at year 3 (commercial banks).","225000000"),
 ("Target State","securities_yr3","number","$","optional","> 0","Stated intent: securities portfolio at year 3.","240000000"),
 ("Target State","headcount_yr3","number","FTE","optional","> 0","Stated intent: total headcount at year 3.","64"),
 ("Target State","footprint","text","—","optional","free text","Market footprint description.","national digital, Midwest-weighted"),
 ("Target State","initial_capital","number","$","always","> 0","Paid-in capital at organization. Drives the equity base; org costs are expensed against it at month 0.","70000000"),

 ("Peer Query","consumer_loan_share","number","share 0-1","always","0-1","Intended consumer loans / total assets at maturity. Peer-selection feature (stated intent, never model output).","0.22"),
 ("Peer Query","fee_income_share","number","share 0-1","always","0-1","Intended fee income / total revenue at maturity. Peer-selection feature.","0.25"),
 ("Peer Query","core_funding_share","number","share 0-1","always","0-1","Intended core deposits / total funding. Peer-selection feature.","0.95"),
 ("Peer Query","digital_channel","number","0 or 1","always","0 or 1","1 if primary channel is digital/app, 0 if branch/relationship. Peer-selection feature.","1.0"),
 ("Peer Query","metric (prior_metrics rows)","text","—","optional","deposit_growth_yr1 / cost_of_deposits_spread / card_nco_mature / cac_per_funded_account / opex_per_active_acct / efficiency_q12","Metrics to benchmark against the frozen cohort. List only metrics meaningful for this business model (e.g. omit card_nco_mature for a bank with no card book).","deposit_growth_yr1"),

 ("Assumptions","(confidence column)","text","—","recommended","observed / contractual / externally_benchmarked / expert_judgment / management_estimate / derived / unsubstantiated","Evidence-quality tag per assumption. Rolls up into the assumption-quality map; 'unsubstantiated' on a material driver blocks a green readiness status.","externally_benchmarked"),
 ("Assumptions","(prior_metric column)","text","—","optional","a prior_metrics name","Which cohort prior this assumption maps to for percentile placement in the assumption book.","cost_of_deposits_spread"),

 ("Assumptions","fed_funds","number","annual rate","chassis","0-0.30","Reference short rate for the projection horizon. Anchors funding spreads and the borrowing rate.","0.045"),
 ("Assumptions","savings_rate","number","annual rate","chassis","0-0.30","Digital banks: savings/HYSA rate. Community banks: blended interest-bearing deposit rate. Shocked by deposit_beta_up in rate scenarios.","0.0385"),
 ("Assumptions","deposit_beta_up","number","beta","chassis","0-1.5","Share of a rate shock passed through to deposit pricing (rises to 0.75 in the +300bp scenario for digital banks).","0.6"),
 ("Assumptions","monthly_attrition","number","monthly rate","chassis","0-0.15","Monthly account/relationship attrition. Negative values are rejected: negative attrition mints customers.","0.015"),
 ("Assumptions","balance_ramp_months","integer","months","chassis","1-36","Months for a new account/relationship to reach average balance. Drives the immature-balance discount.","9"),
 ("Assumptions","cash_target_pct_deposits","number","share","chassis","0-0.5","Operating cash held as a share of deposits; securities absorb the residual.","0.08"),
 ("Assumptions","cash_yield","number","annual rate","chassis","0-0.30","Yield on operating cash.","0.044"),
 ("Assumptions","securities_yield","number","annual rate","chassis","0-0.30","Yield on the securities portfolio (the funding-waterfall residual).","0.0445"),
 ("Assumptions","borrow_spread","number","annual spread","optional","0-0.05","Spread over fed_funds on wholesale (FHLB-style) borrowings, drawn automatically when loans outrun deposits+equity. Default 0.0045.","0.0045"),
 ("Assumptions","productive_hours_m","number","hours/month","chassis","> 0","Productive hours per ops FTE per month; converts capacity minutes to headcount.","140"),
 ("Assumptions","loaded_cost_ops_fte_m","number","$/month","chassis","> 0","Fully loaded monthly cost per operations FTE.","7800"),
 ("Assumptions","fixed_exec_team_m","number","$/month","chassis","> 0","Fixed monthly cost of executives, risk, compliance leadership, finance.","420000"),
 ("Assumptions","tech_core_base_m","number","$/month","chassis","> 0","Fixed monthly core-processor / technology base cost.","160000"),
 ("Assumptions","tech_per_acct_m","number","$/account/month","chassis",">= 0","Variable technology cost per account per month.","0.42"),
 ("Assumptions","occupancy_other_m","number","$/month","chassis",">= 0","Occupancy and other fixed monthly expenses.","110000"),
 ("Assumptions","org_costs_pre_open","number","$","chassis",">= 0","Organization costs expensed at month 0 against opening capital.","5000000"),
 ("Assumptions","tax_rate","number","rate","chassis","0-0.6","Effective tax rate, applied once cumulative pre-tax income turns positive (NOL carryforward simplification).","0.24"),

 ("Assumptions","migration_accounts_m1","number","accounts","core_deposits",">= 0","Funded accounts migrating from the parent's existing base in month 1; decays geometrically thereafter.","6000"),
 ("Assumptions","migration_decay","number","monthly factor","core_deposits","0-1","Geometric monthly decay of the migration flow (0.93 = each month brings 93% of the prior month's migrants).","0.93"),
 ("Assumptions","new_accts_per_marketing_dollar","number","accounts/$","core_deposits","> 0","Acquisition efficiency: 1/CAC. 0.005714 = $175 CAC. Benchmarked against the cohort's cac_per_funded_account prior.","0.005714"),
 ("Assumptions","avg_balance_savings","number","$","core_deposits","> 0","Average mature savings balance per account.","2400"),
 ("Assumptions","avg_balance_checking","number","$","core_deposits","> 0","Average mature checking balance per account.","1200"),
 ("Assumptions","savings_share_accounts","number","share 0-1","core_deposits","0-1","Share of accounts that are savings (vs checking). Drives the blended deposit cost.","0.55"),
 ("Assumptions","checking_rate","number","annual rate","core_deposits","0-0.30","Rate paid on checking balances.","0.001"),

 ("Assumptions","bankers_start","number","FTE","relationship_deposits","> 0","Relationship bankers at open.","5"),
 ("Assumptions","bankers_add_per_m","number","FTE/month","relationship_deposits",">= 0","Banker hiring pace per month, capped at bankers_max.","0.25"),
 ("Assumptions","bankers_max","number","FTE","relationship_deposits","> 0","Maximum banker headcount.","13"),
 ("Assumptions","new_relationships_per_banker_m","number","relationships","relationship_deposits","> 0","New commercial relationships each banker sources per month. The growth-linkage flag ties deposit growth to this and the hiring ramp.","4.8"),
 ("Assumptions","avg_deposit_per_relationship","number","$","relationship_deposits","> 0","Average mature deposit balance per commercial relationship.","340000"),
 ("Assumptions","interest_bearing_share","number","share 0-1","relationship_deposits","0-1","Share of relationship deposits that are interest-bearing.","0.70"),

 ("Assumptions","card_penetration","number","share 0-1","revolving_credit","0-1","Share of active accounts holding the card at maturity; ramps over card_penetration_ramp_m.","0.22"),
 ("Assumptions","card_penetration_ramp_m","integer","months","revolving_credit","1-36","Months for card penetration to reach its mature level.","18"),
 ("Assumptions","card_avg_balance","number","$","revolving_credit","> 0","Average revolving balance per card account.","1900"),
 ("Assumptions","card_yield","number","annual rate","revolving_credit","0-0.40","Effective yield on card receivables.","0.209"),
 ("Assumptions","card_nco_mature","number","annual rate","revolving_credit","0-0.40","Mature net charge-off rate; vintage-ramped over card_nco_ramp_m. Benchmarked against the cohort prior; multiplied 1.75x in credit stress.","0.058"),
 ("Assumptions","card_nco_ramp_m","integer","months","revolving_credit","1-36","Months for losses to season to the mature rate.","12"),
 ("Assumptions","allowance_coverage","number","share of receivables","revolving_credit","0-0.2","Allowance held as a share of receivables (CECL-lite coverage).","0.065"),

 ("Assumptions","lenders_start","number","FTE","commercial_lending","> 0","Commercial lenders at open.","4"),
 ("Assumptions","lenders_add_per_m","number","FTE/month","commercial_lending",">= 0","Lender hiring pace, capped at lenders_max.","0.20"),
 ("Assumptions","lenders_max","number","FTE","commercial_lending","> 0","Maximum lender headcount.","9"),

 ("Assumptions","monthly_debit_spend_per_acct","number","$/account/month","payments_fees",">= 0","Average monthly debit spend per account; interchange = spend x interchange_rate.","540"),
 ("Assumptions","interchange_rate","number","rate","payments_fees","0-0.03","Interchange rate on debit spend. Assumes the Durbin small-issuer exemption (<$10B assets); the rule engine raises a counsel item automatically.","0.0105"),
 ("Assumptions","fee_per_acct_m","number","$/account/month","payments_fees",">= 0","Other monthly fee income per account.","0.50"),

 ("Assumptions","service_charge_per_rel_m","number","$/relationship/month","relationship_fees",">= 0","Monthly service charges per commercial relationship.","95"),
 ("Assumptions","tm_fee_rate_ann","number","annual rate on deposits","relationship_fees","0-0.01","Treasury-management fee income as an annual rate on relationship deposits.","0.0009"),

 ("Assumptions","kyc_reviews_per_new_acct","number","reviews","capacity_expenses",">= 0","KYC reviews per new account.","1.0"),
 ("Assumptions","kyc_min_per_review","number","minutes","capacity_expenses","> 0","Minutes per KYC review.","6"),
 ("Assumptions","fraud_alerts_per_1k_accts_m","number","alerts/1k accts/month","capacity_expenses",">= 0","Monthly fraud alerts per thousand accounts. Alerts x minutes / productive hours = investigators: the capacity engine that makes asset-light stories prove their staffing.","9"),
 ("Assumptions","min_per_alert","number","minutes","capacity_expenses","> 0","Minutes to work one fraud alert.","22"),
 ("Assumptions","service_contacts_per_acct_m","number","contacts/account/month","capacity_expenses",">= 0","Monthly service contacts per account.","0.09"),
 ("Assumptions","min_per_contact","number","minutes","capacity_expenses","> 0","Minutes per service contact.","7.5"),

 ("Assumptions","onboarding_min_per_rel","number","minutes","branch_capacity_expenses","> 0","Minutes to onboard one new commercial relationship (KYC, docs, account setup).","240"),
 ("Assumptions","service_min_per_rel_m","number","minutes/month","branch_capacity_expenses",">= 0","Monthly servicing minutes per relationship.","35"),
 ("Assumptions","credit_admin_min_per_loan_m","number","minutes/month","branch_capacity_expenses",">= 0","Monthly credit-administration minutes per loan (count derived from balances / avg_loan_size).","90"),

 ("Marketing Budget","month","integer","1-36","core_deposits","1-36","Projection month. One row per month; missing tail months repeat the last value.","1"),
 ("Marketing Budget","amount ($)","number","$/month","core_deposits",">= 0","Acquisition spend for that month. Accounts acquired = amount x new_accts_per_marketing_dollar; the pre-filing marketing-linkage requirement is satisfied by construction.","700000"),

 ("Loan Segments","name","text","—","commercial_lending","short id","Segment identifier (e.g. cre, ci). Balance appears as loans_<name> in outputs; 'cre' feeds the CRE/capital concentration test.","cre"),
 ("Loan Segments","orig_per_lender_m","number","$/lender/month","commercial_lending","> 0","Monthly originations per lender at full ramp.","680000"),
 ("Loan Segments","ramp_m","integer","months","commercial_lending","1-36","Months for origination pace to reach full run-rate.","9"),
 ("Loan Segments","amort_annual","number","annual rate","commercial_lending","0-1","Annual amortization/paydown rate of the segment balance.","0.08"),
 ("Loan Segments","yield","number","annual rate","commercial_lending","0-0.30","Segment loan yield.","0.0715"),
 ("Loan Segments","nco_mature","number","annual rate","commercial_lending","0-0.40","Mature net charge-off rate for the segment; vintage-ramped over nco_ramp_m.","0.0035"),
 ("Loan Segments","nco_ramp_m","integer","months","commercial_lending","1-60","Months for segment losses to season.","24"),
 ("Loan Segments","allowance_coverage","number","share of balance","commercial_lending","0-0.2","Allowance held against the segment balance.","0.013"),
 ("Loan Segments","avg_loan_size","number","$","commercial_lending","> 0","Average loan size; converts balances to loan counts for credit-admin capacity.","1600000"),

 ("Entity Map","id","text","—","optional","MAP-01, MAP-02...","Flag identifier from the step 0A entity/funds-flow review.","MAP-01"),
 ("Entity Map","class","text","—","optional","advisory / likely_regulatory_objection / satisfied / commercial_assumption_requiring_support / counsel_determination_required","Flag severity class; drives the open-items count in readiness.","advisory"),
 ("Entity Map","text","text","—","optional","free text","The structural observation (affiliate dependence, shared staffing, funds-flow concerns).","Marketing sits in the parent under a services agreement..."),
]


def _dictionary_sheet(wb):
    ws = _sheet(wb, "Data Dictionary",
                ["Sheet", "Field", "Type", "Units", "Required for", "Range / allowed values",
                 "Description", "Example"],
                [16, 30, 10, 18, 22, 34, 78, 30])
    r = 2
    last = None
    for sheet, field, typ, units, req, rng, desc, ex in DICTIONARY:
        if sheet != last:
            c = ws.cell(row=r, column=1, value=sheet)
            c.font = HDR; last = sheet
        else:
            ws.cell(row=r, column=1, value="")
        for j, v in enumerate((field, typ, units, req, rng, desc, ex), 2):
            c = ws.cell(row=r, column=j, value=v); c.font = KEY
            c.alignment = Alignment(wrap_text=True, vertical="top")
        r += 1
    note = ws.cell(row=r + 1, column=1,
                   value="Blue cells throughout the workbook are inputs. Load a module on the Modules "
                         "sheet and its required assumptions (per this dictionary) become mandatory; "
                         "the upload validator rejects incomplete or out-of-range configurations with "
                         "a precise reason rather than producing partial financials.")
    note.font = NOTE
    ws.merge_cells(start_row=r + 1, start_column=1, end_row=r + 1, end_column=8)
    return ws
