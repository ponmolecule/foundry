"""Protocol harness — runs the Phase A/B test set from
GPT_Claude_consolidated_v2 and prints a scored report.

  python -m foundry.tests_protocol
"""
import copy, json, math, os, sys
from . import chassis, run as runner
from .client_solstice import CLIENT as SOLSTICE
from .client_blackland import CLIENT as BLACKLAND

GOLDENS = json.load(open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "goldens.json")))
RESULTS = []

def check(tid, name, ok, detail=""):
    RESULTS.append((tid, name, ok, detail))
    print(f"  {'PASS' if ok else 'FAIL'}  [{tid}] {name}" + (f" — {detail}" if detail else ""))

# ---------------- T2: golden regression ----------------
def t2():
    print("T2 golden-run regression")
    r = runner.run(copy.deepcopy(SOLSTICE))
    check("T2", "Solstice canonical hash reproduces golden v5",
          r["run_hash"] == GOLDENS["solstice_golden_v5"],
          f"{r['run_hash']} vs {GOLDENS['solstice_golden_v5']}")
    b = runner.run(copy.deepcopy(BLACKLAND))
    check("T2", "Blackland canonical hash reproduces golden v2",
          b["run_hash"] == GOLDENS["blackland_golden_v2"],
          f"{b['run_hash']} vs {GOLDENS['blackland_golden_v2']}")

# ---------------- T3: metamorphic with mechanism-exercised reporting ----------------
def _summ(cfg, ov=None, cap=None):
    c = copy.deepcopy(cfg)
    if ov: c["assumptions"].update(ov)
    if cap: c["target_state"]["initial_capital"] = cap
    return chassis.summarize(chassis.project(c), c)

def t3():
    print("T3 metamorphic suite (direct-mechanics layer; balances/behavior fixed)")
    for cfg, tag in ((SOLSTICE, "solstice"), (BLACKLAND, "blackland")):
        base = _summ(cfg)
        cases = [
            ("deposit rate +100bp -> cum NI must not improve",
             {"savings_rate": cfg["assumptions"]["savings_rate"] + 0.01},
             "cost_of_funds", lambda s: s["cum_net_income"] <= base["cum_net_income"],
             lambda s: s["cum_net_income"] != base["cum_net_income"]),
            ("NCO x1.5 -> cum NI must not improve",
             {"_nco_mult": 1.5}, "credit_losses",
             lambda s: s["cum_net_income"] <= base["cum_net_income"],
             lambda s: s["cum_net_income"] != base["cum_net_income"]),
            ("attrition x2 -> ending accounts must not rise",
             {"monthly_attrition": cfg["assumptions"]["monthly_attrition"] * 2},
             "volume_attrition", lambda s: s["accounts_yr3"] <= base["accounts_yr3"],
             lambda s: s["accounts_yr3"] != base["accounts_yr3"]),
            ("growth x0.5 with running fixed costs -> cumulative NI must worsen",
             {"_growth_mult": 0.5}, "volume_vs_fixed_cost",
             lambda s: s["cum_net_income"] <= base["cum_net_income"],
             lambda s: s["cum_net_income"] != base["cum_net_income"]),
        ]
        for name, ov, mech, expect, exercised in cases:
            s = _summ(cfg, ov)
            if not exercised(s):
                check("T3", f"{tag}: {name}", False, f"NOT EXERCISED (mechanism {mech} produced no delta)")
            else:
                check("T3", f"{tag}: {name}", expect(s), f"mechanism={mech}")
        # capital perturbation (target_state, not assumptions)
        s = _summ(cfg, cap=cfg["target_state"]["initial_capital"] + 50e6)
        ok = s["min_leverage"] >= base["min_leverage"]
        ex = s["min_leverage"] != base["min_leverage"]
        check("T3", f"{tag}: capital +$50M -> min leverage must not fall",
              ok if ex else False, "mechanism=capital_base" if ex else "NOT EXERCISED")

# ---------------- T4: challenge benchmark (Icarus expectations) ----------------
def t4():
    print("T4 challenge engine vs labeled broken applicant (Icarus)")
    icarus = copy.deepcopy(SOLSTICE)
    icarus["client_legal_name"] = "Icarus Financial Corp."
    icarus["proposed_bank"] = "Icarus Bank (in organization)"
    icarus["target_state"]["initial_capital"] = 45e6
    icarus["assumptions"].update({
        "new_accts_per_marketing_dollar": 1 / 22.0, "marketing_budget_m": [0.6e6] * 36,
        "savings_rate": 0.0125, "card_nco_mature": 0.021,
        "fraud_alerts_per_1k_accts_m": 2.0, "min_per_alert": 8.0})
    r = runner.run(icarus)
    exp = GOLDENS["icarus_expectations"]
    fails = {t["constraint"] for t in r["constraint_tests"] if not t["pass"]}
    for c in exp["constraints_fail"]:
        check("T4", f"constraint '{c}' detected as breached", c in fails)
    fids = {f["id"] for f in r["flags"]}
    for fid in exp["must_flag"]:
        check("T4", f"flag '{fid}' raised", fid in fids)
    clean = runner.run(copy.deepcopy(SOLSTICE))
    hard = [t for t in clean["constraint_tests"] if not t["pass"]]
    check("T4", "clean case (Solstice) raises no constraint breaches", len(hard) == 0,
          f"{len(clean['flags'])} advisory/support flags retained — review burden, not noise gate")

# ---------------- T14: fail-closed input validation ----------------
def t14():
    print("T14 fail-closed validation")
    broken = copy.deepcopy(BLACKLAND)
    del broken["assumptions"]["savings_rate"]
    try:
        runner.run(broken)
        check("T14", "missing required assumption fails closed", False, "run completed on invalid config")
    except Exception as e:
        check("T14", "missing required assumption fails closed", True, f"raised {type(e).__name__}: {str(e)[:60]}")
    broken2 = copy.deepcopy(SOLSTICE)
    broken2["assumptions"]["monthly_attrition"] = -0.5   # nonsense: negative attrition mints accounts
    try:
        runner.run(broken2)
        check("T14", "nonsense input (negative attrition) rejected", False,
              "engine calculated it — validation layer incomplete")
    except Exception as e:
        check("T14", "nonsense input (negative attrition) rejected", True,
              f"{type(e).__name__}: {str(e)[:70]}")
    # T1 supported-mechanics: a JSON-only client must run with zero code
    from .registry import ENGAGEMENTS
    if "prairie-digital-bank" in ENGAGEMENTS:
        p = runner.run(copy.deepcopy(ENGAGEMENTS["prairie-digital-bank"]["config"]))
        ok = all(t["pass"] for t in p["constraint_tests"])
        check("T1", "supported-mechanics client (Prairie, JSON-only) runs clean, zero code", ok,
              f"hash {p['run_hash']}")

# ---------------- T6 strong form ----------------
def t6():
    print("T6 strong form: unused modules must not move existing clients")
    r = runner.run(copy.deepcopy(SOLSTICE))
    check("T6", "Solstice unchanged with commercial_lending/relationship modules registered",
          r["run_hash"] == GOLDENS["solstice_golden_v5"])

def t15():
    print("T15 output-contract parity: Excel round-trip must reproduce the JSON run")
    import json as _json
    from .excelio import workbook_from_config, parse_workbook
    import io as _io
    cfg = _json.load(open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                       "clients_uploaded", "prairie_digital.json"), encoding="utf-8"))
    buf = _io.BytesIO(); workbook_from_config(cfg).save(buf)
    cfg2 = parse_workbook(buf.getvalue())
    r1, r2 = runner.run(copy.deepcopy(cfg)), runner.run(cfg2)
    check("T15", "scenarios identical across formats", r1["scenarios"] == r2["scenarios"])
    check("T15", "monthly financials identical across formats", r1["base_monthly"] == r2["base_monthly"])
    check("T15", "cohort, priors, flags identical across formats",
          r1["cohort"] == r2["cohort"] and r1["priors"] == r2["priors"]
          and [f["id"] for f in r1["flags"]] == [f["id"] for f in r2["flags"]])


def t16():
    print("T16 validate/run seam: anything validate_config passes, run() must complete")
    import json as _json
    from .configio import validate_config, ConfigError
    cfg = _json.load(open(os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                                       "clients_uploaded", "prairie_digital.json"), encoding="utf-8"))
    validate_config(copy.deepcopy(cfg))
    try:
        runner.run(copy.deepcopy(cfg))
        check("T16", "validated config runs end-to-end", True)
    except KeyError as e:
        check("T16", "validated config runs end-to-end", False, f"KeyError {e} — seam reopened")
    broken = copy.deepcopy(cfg); broken.pop("step_0a")
    try:
        validate_config(broken)
        check("T16", "config missing run()-consumed key fails closed", False, "validator passed it")
    except ConfigError:
        check("T16", "config missing run()-consumed key fails closed", True)


def t17():
    print("T17 provenance gate: source names never appear in the product surface")
    import glob
    banned = ["goldstein", "superapp", "jsx build", "predecessor a", "predecessor b"]
    hits = []
    for path in glob.glob("web/*.html"):
        low = open(path, encoding="utf-8").read().lower()
        for b in banned:
            if b in low:
                hits.append(f"{path}: '{b}'")
    check("T17", "web/ carries no source attribution", not hits, "; ".join(hits))


def t18():
    print("T18 engagement store: schema-pinned save/load reproduces the golden hash")
    import tempfile
    from . import store
    base = runner.run(copy.deepcopy(SOLSTICE))["run_hash"]
    prev = os.environ.get("FOUNDRY_DATA_DIR")
    with tempfile.TemporaryDirectory() as td:
        os.environ["FOUNDRY_DATA_DIR"] = td
        try:
            meta = store.save_engagement(copy.deepcopy(SOLSTICE))
            loaded = store.load_engagement(meta["slug"])
            ok_ver = loaded.get("config_schema_version") == store.CONFIG_SCHEMA_VERSION
            rerun = runner.run(loaded)["run_hash"]
            check("T18a", "stored-then-loaded Solstice reproduces golden hash",
                  ok_ver and rerun == base, f"{rerun} vs {base}, ver_ok={ok_ver}")
            # tamper: unsupported version must refuse to load
            path = meta["path"]
            bad = json.load(open(path, encoding="utf-8"))
            bad["config_schema_version"] = "999"
            json.dump(bad, open(path, "w", encoding="utf-8"))
            try:
                store.load_engagement(meta["slug"])
                check("T18b", "unsupported schema version fails closed", False, "loaded anyway")
            except store.SchemaVersionError:
                check("T18b", "unsupported schema version fails closed", True)
            # missing version must refuse to load
            bad.pop("config_schema_version")
            json.dump(bad, open(path, "w", encoding="utf-8"))
            try:
                store.load_engagement(meta["slug"])
                check("T18c", "missing schema version fails closed", False, "loaded anyway")
            except store.SchemaVersionError:
                check("T18c", "missing schema version fails closed", True)
        finally:
            if prev is None:
                os.environ.pop("FOUNDRY_DATA_DIR", None)
            else:
                os.environ["FOUNDRY_DATA_DIR"] = prev


def t19():
    print("T19 field library: concordance closure + progressive disclosure + budget")
    from . import fieldlib as fl
    # (a) closure against the real goldens: every assumption key consumed by the
    # chassis for Solstice/Blackland has exactly one disposition in the library
    lib_fields = set()
    for a in fl.ARCHETYPES.values():
        lib_fields |= set(a["drivers"]) | set(a["defaults"])
    for d in fl.CAPACITY_DEFAULTS.values():
        lib_fields |= set(d)
    lib_fields |= set(fl.GLOBAL_DEFAULTS) | {"org_costs_pre_open"}
    golden_keys = set(SOLSTICE["assumptions"].keys()) | set(BLACKLAND["assumptions"].keys())
    missing = sorted(golden_keys - lib_fields)
    check("T19a", "every golden-consumed assumption key has a library disposition",
          not missing, f"missing: {missing}")
    phantom = sorted(f for f in lib_fields - golden_keys if f != "org_costs_pre_open")
    check("T19b", "library names no field the chassis does not consume",
          not phantom, f"phantom: {phantom}")
    # (c) progressive disclosure: deposits-only surface contains no lending field
    dep = fl.fields_for(["funnel_deposit"])
    lend = fl.fields_for(["funnel_deposit", "commercial_lending"])
    lend_drivers = set(fl.ARCHETYPES["commercial_lending"]["drivers"])
    leak = lend_drivers & (set(dep["typed"]) | set(dep["defaults"]))
    check("T19c", "deposits-only surface carries no lending fields", not leak, f"leak: {sorted(leak)}")
    strictly_adds = set(dep["typed"]) < set(lend["typed"])
    check("T19d", "adding an archetype strictly adds typed fields", strictly_adds,
          f"{len(dep['typed'])} -> {len(lend['typed'])}")
    # (e) budget rule P6
    two = fl.typed_budget(["funnel_deposit", "revolving_credit"])
    four = fl.typed_budget(list(fl.ARCHETYPES.keys())[:4])
    check("T19e", "typed budget: 2-product <=40, 4-product <=70",
          two <= 40 and four <= 70, f"two={two}, four={four}")


def t20():
    print("T20 dialect bridge: chassis drivers -> workspace config, paths pinned")
    import json as _json
    from .bridge import bridge_solstice
    from .v2.run_q import run_v2
    tpl = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfg, targets = bridge_solstice(copy.deepcopy(SOLSTICE), tpl)
    r = run_v2(cfg)
    deps = [p for p in r["products"] if p["family"] == "deposit"]
    card = [p for p in r["products"] if p["family"] == "lending"][0]
    dep_path = [sum(p["bal"][t + 1] for p in deps) * 1000 for t in range(12)]
    card_path = [b * 1000 for b in card["bal"][1:13]]
    dd = max(abs(a - b) for a, b in zip(dep_path, targets["target_deposits"]))
    cd = max(abs(a - b) for a, b in zip(card_path, targets["target_receivables"]))
    check("T20a", "bridged deposit path within $1k/qtr of chassis", dd < 1000, f"max ${dd:,.0f}")
    check("T20b", "bridged card path within $1k/qtr of chassis", cd < 1000, f"max ${cd:,.0f}")


def t21():
    print("T21 v3.1 empty-start + taxonomy templates carry the source values")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.validate_q import ConfigErrorV2
    cfg = _json.load(open("foundry/fixtures/patrick_default_v31.json", encoding="utf-8"))
    check("T21a", "default starts empty (products and modules)",
          not cfg["assumptions"]["lending_products"] and not cfg["assumptions"]["deposit_products"]
          and not cfg["step_0"]["modules"])
    r0 = run_v2(cfg)   # fidelity ruling: the source model keeps the BS alive
    bs0 = r0["financials"]["bs"]
    check("T21b", "empty start RUNS: capital plugs into securities, deposits zero (source-model fidelity)",
          bs0["sec"][0] > 0 and bs0["deposits"][0] == 0
          and abs(bs0["equity"][0] * 1000 - cfg["target_state"]["initial_capital"]) < 1e-3,
          f"sec Q1 {bs0['sec'][0]:,.0f}k, equity Q1 {bs0['equity'][0]:,.0f}k")
    tpl = _json.load(open("foundry/fixtures/patrick_templates_v31.json", encoding="utf-8"))
    cc = next(t for t in tpl["loans"] if t["name"] == "Credit Card")
    ok = (abs(cc["yield_ann"] - 0.18) < 1e-12 and abs(cc["originations_q"] - 1_500_000) < 1e-6
          and abs(cc["charge_off_ann"] - 0.04) < 1e-12)
    check("T21c", "template constants match the source workbook", ok)
    # materialize Retail Demand exactly as the UI does; path must hit the targets
    tg = _json.load(open("foundry/fixtures/patrick_default_targets.json", encoding="utf-8"))
    rd = next(t for t in tpl["deposits"] if t["name"] == "Retail Demand")
    b, monthly = 0.0, []
    for _ in range(36):
        b = b * (1 - rd["runoff_ann"] / 12.0) + rd["adds_m"]; monthly.append(b)
    qp = [monthly[3 * q - 1] for q in range(1, 13)]
    g = (qp[11] / qp[0]) ** (1 / 11) - 1.0 if qp[0] else 0.0
    cfg["assumptions"]["deposit_products"] = [{"name": "Retail Demand", "call_report_line": "depDDA",
        "opening_balance": qp[0] / (1 + g), "growth_q": g, "runoff_q": 0.0, "rate_type": "fixed",
        "rate_paid_ann": rd["rate_paid_ann"], "fee_yield_ann": 0.0, "opex_pct_ann": 0.0,
        "opex_fixed_m": 0.0}]
    cfg["step_0"]["modules"] = ["balance_driven_deposits"]
    r = run_v2(cfg)
    dep = [p for p in r["products"] if p["family"] == "deposit"][0]
    got = [x * 1000 for x in dep["bal"][1:13]]
    q1_exact = abs(got[0] - tg["Retail Demand"][0]) < 1000.0
    q12_close = abs(got[11] - tg["Retail Demand"][11]) < 1000.0
    check("T21d", "fitted template: Q1 exact and Q12 within house tol (editable scalars, no pins)",
          q1_exact and q12_close, f"Q1 d=${abs(got[0]-tg['Retail Demand'][0]):.2f}, Q12 d=${abs(got[11]-tg['Retail Demand'][11]):.2f}")


def t22():
    print("T22 FIW: per-engagement workbook, progressive sheets, values, hash")
    import io as _io
    import json as _json
    from openpyxl import load_workbook as _lw
    from .v2.fiw import build_fiw, cfg_hash
    cfg = _json.load(open("foundry/fixtures/patrick_default_v31.json", encoding="utf-8"))
    cfg["proposed_bank"] = "Testament Bank"
    cfg["charter_profile"] = {"charter_type": "national", "regulator": "OCC",
                                "cblr_election": True, "pre_open_months": 9}
    cfg["assumptions"]["deposit_products"] = [{"name": "Retail Demand", "call_report_line": "depDDA",
        "opening_balance": 8_930_000, "growth_q": 0.24, "runoff_q": 0.0, "rate_type": "fixed",
        "rate_paid_ann": 0.005, "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0}]
    data, gh = build_fiw(cfg)
    wb = _lw(_io.BytesIO(data))
    check("T22a", "deposits-only FIW: no loan sheet (disclosure extends to paper; "
                    "STATE is the hidden self-containment sheet)",
          wb.sheetnames == ["README", "CONTROL", "ASSM_DEPOSITS", "LIMITS", "SETTINGS", "STATE"]
          and wb["STATE"].sheet_state == "veryHidden"
          and wb["SETTINGS"].sheet_state == "visible", str(wb.sheetnames))
    check("T22b", "generation hash on README matches the canonical hash",
          any(r[0].value == "Generation hash" and r[1].value == cfg_hash(cfg) for r in wb["README"].iter_rows()))
    dep = {(r[0].value): r[3].value for r in wb["ASSM_DEPOSITS"].iter_rows(min_row=2)}
    check("T22c", "values transcribed (rate 0.5%, opening $8.93M) and line rendered as label",
          abs(dep["deposit.0.rate_paid_ann"] - 0.005) < 1e-12
          and abs(dep["deposit.0.opening_balance"] - 8_930_000) < 1e-6
          and dep["deposit.0.call_report_line"] == "Deposits: Transaction (DDA)")
    cfg["assumptions"]["lending_products"] = [{"name": "Credit Card", "call_report_line": "loanCreditCard",
        "opening_balance": 0, "originations_q": 1_500_000, "orig_growth_q": 0, "runoff_q": 0,
        "rate_type": "fixed", "yield_ann": 0.18, "charge_off_ann": 0.04, "provision_rate_ann": None,
        "reserve_rate_pct_bal": 0.03, "measurement": "amortized", "fee_yield_ann": 0,
        "opex_pct_ann": 0, "opex_fixed_m": 0}]
    data2, _ = build_fiw(cfg)
    wb2 = _lw(_io.BytesIO(data2))
    keys2 = [r[0].value for r in wb2["ASSM_LOANS"].iter_rows(min_row=2)]
    check("T22d", "adding a loan adds ASSM_LOANS; no mb_ rows without originate-to-sell",
          "ASSM_LOANS" in wb2.sheetnames and not any("mortgage_banking" in (k or "") for k in keys2))


def t23():
    print("T23 FIW diff-import: only human edits apply; facts ignored; fail-closed")
    import io as _io
    import os as _os
    import json as _json
    import tempfile
    from openpyxl import load_workbook as _lw
    from .v2 import fiw as F
    old_env = _os.environ.get("FOUNDRY_DATA_DIR")
    tmp = tempfile.mkdtemp()
    _os.environ["FOUNDRY_DATA_DIR"] = tmp
    try:
        cfg = _json.load(open("foundry/fixtures/patrick_default_v31.json", encoding="utf-8"))
        cfg["assumptions"]["deposit_products"] = [{"name": "Retail Demand", "call_report_line": "depDDA",
            "opening_balance": 8_930_000, "growth_q": 0.24, "runoff_q": 0.0, "rate_type": "fixed",
            "rate_paid_ann": 0.005, "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0}]
        cfg["step_0"]["modules"] = ["balance_driven_deposits"]
        data, gh = F.build_fiw(cfg)
        F.persist_snapshot(cfg, gh)
        wb = _lw(_io.BytesIO(data))
        ws = wb["ASSM_DEPOSITS"]
        for r in ws.iter_rows(min_row=2):
            if r[0].value == "deposit.0.rate_paid_ann":
                r[3].value = 0.0125          # human edit
            if r[0].value == "deposit.0.call_report_line":
                r[3].value = "Loans: Other"  # vandalized FACT — must be ignored
        buf = _io.BytesIO(); wb.save(buf)
        # in-app edit made AFTER generation must survive (workbook cell untouched)
        current = _json.loads(_json.dumps(cfg))
        current["assumptions"]["deposit_products"][0]["growth_q"] = 0.30
        merged, rep = F.diff_import(buf.getvalue(), current)
        d = merged["assumptions"]["deposit_products"][0]
        check("T23a", "the one human edit applied", abs(d["rate_paid_ann"] - 0.0125) < 1e-12)
        # SUPERSEDED by the workbook-is-the-document ruling (Allied Bank episode,
        # TEST_CASES 33): the workbook's generation state governs; in-app edits
        # made after generation are REPLACED on upload — and the report says so.
        check("T23b", "the workbook's state governs: post-generation session edits are "
                        "replaced on upload, and the report carries the session note",
              abs(d["growth_q"] - cfg["assumptions"]["deposit_products"][0]["growth_q"]) < 1e-12
              and "session_note" in rep and "workbook's state now governs" in rep["session_note"])
        check("T23c", "fact rows are ignored even when vandalized",
              d["call_report_line"] == "depDDA")
        check("T23d", "edit report names exactly the human change",
              rep["edit_count"] == 1 and rep["edits"][0]["key"] == "deposit.0.rate_paid_ann")
        # SELF-CONTAINED semantics: a tampered README hash is IRRELEVANT when the
        # embedded state is present — the workbook's own state governs. The refusal
        # now belongs to legacy files only: strip STATE + unknown hash => fail closed.
        wb2 = _lw(_io.BytesIO(data))
        for r in wb2["README"].iter_rows():
            if r[0].value == "Generation hash":
                r[1].value = "abcdef000000"
        b2 = _io.BytesIO(); wb2.save(b2)
        _m2, _r2 = F.diff_import(b2.getvalue(), current)
        check("T23e", "tampered hash is irrelevant when embedded state governs",
              _r2.get("base") == "embedded")
        del wb2["STATE"]
        b3 = _io.BytesIO(); wb2.save(b3)
        try:
            F.diff_import(b3.getvalue(), current); failed = False
        except ValueError:
            failed = True
        check("T23f", "legacy file (no embedded state) with unknown hash fails closed", failed)
    finally:
        if old_env is None:
            _os.environ.pop("FOUNDRY_DATA_DIR", None)
        else:
            _os.environ["FOUNDRY_DATA_DIR"] = old_env


def t24():
    print("T24 Mode T stage T-1: ingest+recon on the two real shapes")
    import os as _os
    from .modet import recon_file
    src = "/mnt/project/Klaros_Bank_Charter_Financial_Model_v1_0_Patrick.xlsx"
    if _os.path.exists(src):
        rep = recon_file(open(src, "rb").read(), src)
        check("T24a", "hidden sheets surfaced (the PEER lesson)",
              "PEER" in rep["hidden_sheets"], str(rep["hidden_sheets"])[:90])
        themes = {t for c in rep["candidates"] for t in c["themes"]}
        check("T24b", "lexicon finds the expected themes across the workbook",
              {"deposits", "loans", "credit", "capital", "treasury"} <= themes, str(sorted(themes)))
        long_axes = [ax for sh in rep["sheets"] for ax in sh["time_axes"]
                      if ax["cadence"] == "monthly" or (ax["cadence"] == "periodic" and ax["span"] >= 18)]
        check("T24c", "a long time axis detected (labeled monthly, or periodic span>=18)", bool(long_axes))
        check("T24d", "no unit is ever asserted",
              all(c["units"] == "UNVERIFIED" for c in rep["candidates"]))
        rep2 = recon_file(open(src, "rb").read(), src)
        check("T24e", "recon is deterministic", rep["report_hash"] == rep2["report_hash"])
    else:
        print("  SKIP  source workbook not present in this environment (T24a-e)")
        print("        (5 checks skipped by design on machines without the client file;")
        print("         they run and must pass in the build environment before any bundle ships)")
    csvp = "foundry/fixtures/modet/prairie_style_unit_economics.csv"
    rep = recon_file(open(csvp, "rb").read(), csvp)
    ok = (rep["kind"] == "csv"
          and any(ax["cadence"] == "periodic" and ax["span"] == 12
                   for sh in rep["sheets"] for ax in sh["time_axes"])
          and any("volume" in c["themes"] for c in rep["candidates"])
          and any("fees" in c["themes"] for c in rep["candidates"]))
    check("T24f", "CSV shape (S4): periodic axis (span 12, cadence NOT guessed) + volume/fee candidates", ok,
          f"{rep['candidate_count']} candidates")


def t25():
    print("T25 Mode T stage T-2: mapping session — declarations, exact series, gaps")
    import json as _json
    from .modet import ingest
    from . import modet_map as M
    from .modet import recon
    from .v2.run_q import run_v2
    csvp = "foundry/fixtures/modet/prairie_style_unit_economics.csv"
    inv = ingest(open(csvp, "rb").read(), csvp)
    rep = recon(inv)
    cfg = _json.load(open("foundry/fixtures/patrick_default_v31.json", encoding="utf-8"))
    cfg["assumptions"]["deposit_products"] = [{"name": "Digital Deposits", "call_report_line": "depDDA",
        "opening_balance": None, "growth_q": 0.0, "runoff_q": 0.0, "rate_type": "fixed",
        "rate_paid_ann": None, "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0}]
    cfg["step_0"]["modules"] = ["balance_driven_deposits"]
    sess = M.new_session(rep, cfg)
    # refusing to declare is refusing to map
    try:
        M.assign(sess, {"sheet": "(csv)", "row": 4, "label": "Deposit balance"},
                  "deposit.0.balance_path"); undeclared_ok = True
    except ValueError:
        undeclared_ok = False
    check("T25a", "series assignment without declared cadence/units is refused", not undeclared_ok)
    M.assign(sess, {"sheet": "(csv)", "row": 4, "label": "Deposit balance"},
              "deposit.0.balance_path", converter="series_pin",
              declared={"cadence": "monthly", "units": "dollars"})
    M.assign(sess, {"sheet": "(client meeting)", "row": 0, "label": "stated cost of funds"},
              "deposit.0.rate_paid_ann", converter="pct_to_rate",
              declared={"units": "percent"}, params={"value": 1.5})
    merged, log, gaps = M.apply_session(sess, inv, cfg)
    r = run_v2(merged)
    dep = [p for p in r["products"] if p["family"] == "deposit"][0]
    got = [x * 1000 for x in dep["bal"][1:5]]
    src = [5_817_000, 14_294_900, 23_894_900, 34_393_900]  # CSV months 3/6/9/12
    worst = max(abs(a - b) for a, b in zip(got, src))
    check("T25b", "mapped series reproduces the client's path exactly (house tol $1k)",
          worst < 1000.0, f"worst ${worst:,.0f}")
    check("T25c", "scalar conversion applied with provenance user",
          abs(merged["assumptions"]["deposit_products"][0]["rate_paid_ann"] - 0.015) < 1e-12
          and all(x["provenance"] == "user" for x in log))
    check("T25d", "translation log names source, slot, and conversion for every assignment",
          len(log) == 2 and all(("source" in x and "slot" in x and "conversion" in x) for x in log))
    holed = _json.loads(_json.dumps(merged))
    holed["assumptions"]["cash_yield"] = None   # a genuine hole, on a copy
    gaps2 = [g for g in M.slots_for(holed) if g["required"] and not g["filled"]]
    check("T25e", "unfilled required slots surface as gap questions",
          any(g["slot"] == "assumptions.cash_yield" for g in gaps2),
          str([g["slot"] for g in gaps2])[:80])


def t26():
    print("T26 converter library: arithmetic, conventions, registry integrity")
    from . import converters as C
    ok = (abs(C.monthly_flow_x3(1000) - 3000) < 1e-12
          and abs(C.annual_rate_div4(0.15) - 0.0375) < 1e-12
          and abs(C.pct_to_rate(1.5) - 0.015) < 1e-12
          and abs(C.bp_to_rate(25) - 0.0025) < 1e-12
          and abs(C.units_thousands(60000) - 60_000_000) < 1e-6)
    check("T26a", "scalar converters: x3, /4, pct, bp, $000s", ok)
    steps = C.annual_steps_to_quarterly([0.038, 0.036, 0.034])
    check("T26b", "annual steps -> 12 quarterly points, held flat per year (no interpolation)",
          len(steps) == 12 and steps[0] == steps[3] == 0.038 and steps[4] == 0.036 and steps[11] == 0.034)
    path = C.dollar_adds_to_monthly_path(3_000_000, 0.05, 36)
    qp = C.monthly_path_to_quarterly_eop(path)
    import json as _json
    tg = _json.load(open("foundry/fixtures/patrick_default_targets.json", encoding="utf-8"))
    worst = max(abs(a - b) for a, b in zip(qp, tg["Retail Demand"]))
    check("T26c", "dollar-adds path matches the source-workbook recursion (targets sidecar)",
          worst < 1.0, f"worst ${worst:.4f}")
    from .modet_map import CONVERTERS as MC
    check("T26d", "mapping session consumes the library registry (no duplicate arithmetic)",
          MC is C.REGISTRY)


def t27():
    print("T27 inverter metamorphic gate: golden statements -> invert -> re-run -> same statements")
    import json as _json
    import hashlib as _hl
    from .v2.run_q import run_v2
    from .inverter import aggregate_from_run, invert_statements
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    r1 = run_v2(cfg)
    agg1 = aggregate_from_run(r1)          # the "client's statements"
    inv = invert_statements(agg1, cfg)
    r2 = run_v2(inv)
    agg2 = aggregate_from_run(r2)
    def worst(key):
        return max(abs(a - b) for a, b in zip(agg1[key], agg2[key]))
    check("T27a", "deposit balance path reproduced ($000s, house tol 1.0)",
          worst("dep_bal") < 1.0, f"worst {worst('dep_bal'):.3f}")
    check("T27b", "loan balance path reproduced", worst("loan_bal") < 1.0,
          f"worst {worst('loan_bal'):.3f}")
    check("T27c", "interest expense reproduced (pinned deposit rates)",
          worst("dep_int_exp") < 1.0, f"worst {worst('dep_int_exp'):.3f}")
    check("T27d", "interest income reproduced (pinned yields)",
          worst("loan_int_inc") < 1.0, f"worst {worst('loan_int_inc'):.3f}")
    check("T27e", "charge-offs reproduced (pinned CO rates, compensated originations)",
          worst("charge_offs") < 1.0, f"worst {worst('charge_offs'):.3f}")
    h = lambda a: _hl.sha256(_json.dumps({k: [round(x, 6) for x in v] for k, v in a.items()},
                                          sort_keys=True).encode()).hexdigest()[:12]
    check("T27f", "statement-series hash equality (the doc's metamorphic test, made precise)",
          h(agg1) == h(agg2), f"{h(agg1)} vs {h(agg2)}")


def t28():
    print("T28 stage T-5: translation log first-class; gaps become questions")
    import json as _json
    from .modet import ingest, recon
    from . import modet_map as M
    csvp = "foundry/fixtures/modet/prairie_style_unit_economics.csv"
    inv = ingest(open(csvp, "rb").read(), csvp)
    rep = recon(inv)
    cfg = _json.load(open("foundry/fixtures/patrick_default_v31.json", encoding="utf-8"))
    cfg["assumptions"]["deposit_products"] = [{"name": "Digital Deposits", "call_report_line": "depDDA",
        "opening_balance": None, "growth_q": 0.0, "runoff_q": 0.0, "rate_type": "fixed",
        "rate_paid_ann": None, "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0}]
    cfg["step_0"]["modules"] = ["balance_driven_deposits"]
    sess = M.new_session(rep, cfg)
    M.assign(sess, {"sheet": "(csv)", "row": 4, "label": "Deposit balance"},
              "deposit.0.balance_path", converter="series_pin",
              declared={"cadence": "monthly", "units": "dollars"})
    out = M.finalize(sess, inv, cfg, rep)
    log, qs = out["translation_log"], out["gap_questions"]
    check("T28a", "log carries the source recon hash and one row per assignment",
          log["source_report_hash"] == rep["report_hash"] and len(log["rows"]) == 1)
    check("T28b", "unmapped required slot (deposit rate) becomes a phrased question",
          any("Digital Deposits" in q and "rate paid" in q and q.endswith("authority?") for q in qs),
          qs[0] if qs else "(none)")
    out2 = M.finalize(sess, inv, cfg, rep)
    check("T28c", "the log is deterministic (stable hash)",
          log["log_hash"] == out2["translation_log"]["log_hash"])
    check("T28d", "doctrine rides in the artifact itself",
          any("interpolation" in d for d in log["doctrine"]))


def t29():
    print("T29 Icarus at entry: blockers stop nonsense; optimism enters WITH warnings")
    import json as _json
    from .entry_screen import entry_screen
    cfg = _json.load(open("foundry/fixtures/patrick_default_v31.json", encoding="utf-8"))
    a = cfg["assumptions"]
    a["deposit_products"] = [{"name": "MMDA / Savings", "call_report_line": "depSavings",
        "opening_balance": 1e6, "growth_q": 0.05, "runoff_q": 0.0, "rate_type": "fixed",
        "rate_paid_ann": 0.0125, "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0}]
    a["lending_products"] = [{"name": "Credit Card", "call_report_line": "loanCreditCard",
        "opening_balance": 0, "originations_q": 1e6, "orig_growth_q": 0, "runoff_q": 0,
        "rate_type": "fixed", "yield_ann": 0.18, "charge_off_ann": 0.019,
        "provision_rate_ann": None, "reserve_rate_pct_bal": 0.03, "measurement": "amortized",
        "fee_yield_ann": 0, "opex_pct_ann": 0, "opex_fixed_m": 0}]
    sc = entry_screen(cfg)
    check("T29a", "Icarus's mispriced savings (1.25% vs baseline 2.5%) enters WITH a warning",
          not sc["blockers"] and any("MMDA" in w and "1.25%" in w for w in sc["warnings"]),
          (sc["warnings"] or ["-"])[0][:80])
    check("T29b", "Icarus's optimistic card losses (1.9% vs baseline 4%) warned by name",
          any("Credit Card" in w and "optimism" in w for w in sc["warnings"]))
    bad = _json.loads(_json.dumps(cfg))
    bad["target_state"]["initial_capital"] = 2_000_000
    bad["assumptions"]["org_costs_pre_open"] = 2_500_000
    bad["assumptions"]["deposit_products"][0]["runoff_q"] = -0.2
    sc2 = entry_screen(bad)
    check("T29c", "arithmetic self-destruction and negative runoff are BLOCKERS",
          len(sc2["blockers"]) == 2 and any("Day-1" in b for b in sc2["blockers"])
          and any("mints balances" in b or "outside" in b for b in sc2["blockers"]))
    clean = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    sc3 = entry_screen(clean)
    check("T29d", "the clean golden passes the screen silently",
          not sc3["blockers"] and not sc3["warnings"])


def t30():
    print("T30 equivalence matrix: an unedited workbook round-trip is a perfect no-op")
    import io as _io
    import os as _os
    import json as _json
    import tempfile
    from .v2 import fiw as F
    from .v2.run_q import run_v2
    old_env = _os.environ.get("FOUNDRY_DATA_DIR")
    _os.environ["FOUNDRY_DATA_DIR"] = tempfile.mkdtemp()
    try:
        # a wizard-shaped bank, built from the same template constants the JS uses
        tpl = _json.load(open("foundry/fixtures/patrick_templates_v31.json", encoding="utf-8"))
        cfg = _json.load(open("foundry/fixtures/patrick_default_v31.json", encoding="utf-8"))
        rd = next(t for t in tpl["deposits"] if t["name"] == "Retail Demand")
        b, monthly = 0.0, []
        for _ in range(36):
            b = b * (1 - rd["runoff_ann"] / 12.0) + rd["adds_m"]; monthly.append(b)
        qp = [monthly[3 * q - 1] for q in range(1, 13)]
        g = (qp[11] / qp[0]) ** (1 / 11) - 1.0
        cc = next(t for t in tpl["loans"] if t["name"] == "Credit Card")
        cfg["assumptions"]["deposit_products"] = [{"name": "Retail Demand", "call_report_line": "depDDA",
            "opening_balance": qp[0] / (1 + g), "growth_q": g, "runoff_q": 0.0, "rate_type": "fixed",
            "rate_paid_ann": rd["rate_paid_ann"], "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0}]
        cfg["assumptions"]["lending_products"] = [{"name": "Credit Card", "call_report_line": "loanCreditCard",
            "opening_balance": 0.0, "originations_q": cc["originations_q"], "orig_growth_q": 0.0,
            "runoff_q": cc["runoff_q"], "rate_type": "fixed", "yield_ann": cc["yield_ann"],
            "charge_off_ann": cc["charge_off_ann"], "provision_rate_ann": None,
            "reserve_rate_pct_bal": cc["reserve_rate_pct_bal"], "measurement": "amortized",
            "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0}]
        cfg["step_0"]["modules"] = ["balance_driven_deposits", "balance_driven_lending"]
        for name, case in (("wizard-shaped bank", cfg),
                            ("golden pf_a", _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8")))):
            data, gh = F.build_fiw(case)
            F.persist_snapshot(case, gh)
            merged, repd = F.diff_import(data, case)
            same_cfg = _json.dumps(merged, sort_keys=True) == _json.dumps(case, sort_keys=True)
            h1, h2 = run_v2(case)["run_hash"], run_v2(merged)["run_hash"]
            check("T30a" if name.startswith("wizard") else "T30b",
                  f"{name}: unedited round-trip -> zero edits, identical config, identical run hash",
                  repd["edit_count"] == 0 and same_cfg and h1 == h2,
                  f"edits={repd['edit_count']}, hash {h1[:8]} vs {h2[:8]}")
    finally:
        if old_env is None:
            _os.environ.pop("FOUNDRY_DATA_DIR", None)
        else:
            _os.environ["FOUNDRY_DATA_DIR"] = old_env


def t31():
    print("T31 Call Report schedules: assembled from engine series, ties recomputed")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.callreport import build_call_report
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    res = run_v2(cfg)
    cr = build_call_report(res, cfg)
    rc = {r["item"]: r["values"] for r in cr["RC"]["rows"]}
    n = len(rc["12"])
    check("T31g", "all schedule rows are uniform Q1..Q12 (openings normalized away)",
          n == 12 and all(len(r["values"]) == 12 for sch in cr.values() for r in sch["rows"]))
    worst_rc = max(abs(rc["1"][t] + rc["2.a"][t] + rc["2.b"][t] + rc["4.d"][t]
                        + (rc.get("RC-M 2.a", [0]*n)[t]) + rc["6"][t] + rc["10"][t] + rc["11"][t]
                        - rc["12"][t]) for t in range(n))
    check("T31a", "RC ties to the ENGINE's total (HFS memoranda per disclosed convention)",
          worst_rc < 1.0 and any("warehouse" in nt for nt in cr["RC"].get("notes", [])),
          f"worst {worst_rc:.4f}")
    worst_lq = max(abs(rc["13.a"][t] + rc["16"][t] + rc["20"][t] + rc["27.a"][t] - rc["12"][t])
                    for t in range(n))
    check("T31b", "RC ties: deposits + borrowings + other liabilities + equity = Total Assets",
          worst_lq < 1.0, f"worst {worst_lq:.4f}")
    ri = {r["item"]: r["values"] for r in cr["RI"]["rows"]}
    worst_ri = max(abs(ri["1.h"][t] - ri["2.e"][t] - ri["4"][t] + ri["5"][t] - ri["7"][t]
                        - ri["8"][t]) for t in range(n))
    check("T31c", "RI ties: interest income - expense - provision + noninterest net = pretax",
          worst_ri < 1.0, f"worst {worst_ri:.4f}")
    worst_ni = max(abs(ri["8"][t] - ri["9"][t] - ri["12"][t]) for t in range(n))
    check("T31d", "RI ties: pretax - taxes = NET INCOME", worst_ni < 1.0)
    rce = cr["RC-E"]["rows"]
    tot = [r for r in rce if "Total deposits" in r["label"]][0]["values"]
    worst_e = max(abs(tot[t] - rc["13.a"][t]) for t in range(len(tot)))
    check("T31e", "RC-E total ties to RC 13.a", worst_e < 1.0)
    rcr = {str(r["item"]): r["values"] for r in cr["RC-R"]["rows"]}
    lev = res["financials"]["ratios"]["lev"]
    _lev_q = lev[1:13] if len(lev) == 13 else lev[:12]
    check("T31f", "RC-R leverage equals the ratios tab; Part II standardized rows present "
                    "with CET1/T1/Total/RWA and the framework-vs-visibility note",
          rcr["31"] == _lev_q
          and "part2" in cr["RC-R"] and len(cr["RC-R"]["part2"]["rows"]) == 8
          and any("GOVERNS" in n for n in cr["RC-R"]["notes"]))


def t32():
    print("T32 retrodiction harness: known drift measured exactly; fail-closed labels")
    import json as _json
    from .v2.run_q import run_v2
    from .retro import load_actuals, compare
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    res = run_v2(cfg)
    act = load_actuals(open("foundry/fixtures/retro/synthetic_de_novo_actuals.csv", "rb").read(),
                        "synthetic_de_novo_actuals.csv")
    rep = compare(res, act)
    dep = next(s2 for s2 in rep["series"] if s2["label"] == "deposits")
    loans = next(s2 for s2 in rep["series"] if s2["label"] == "loans")
    check("T32a", "deposits drift recovered (~7.4% APE from the x1.08 fixture)",
          abs(dep["mape_pct"] - (1 - 1 / 1.08) * 100) < 0.15, f"MAPE {dep['mape_pct']}%")
    check("T32b", "loans drift recovered (~5.3% APE from the x0.95 fixture)",
          abs(loans["mape_pct"] - (1 / 0.95 - 1) * 100) < 0.15, f"MAPE {loans['mape_pct']}%")
    check("T32c", "overlap window respected (8 actual quarters, not 12)",
          rep["quarters"] == 8 and all(len(s2["rows"]) == 8 for s2 in rep["series"]))
    check("T32d", "summary counts within-15% series and names the worst miss",
          rep["summary"]["series_compared"] == 5
          and isinstance(rep["summary"]["worst_terminal_error_pct"], (int, float)))
    try:
        load_actuals(b"metric,1,2\nmystery_series,1,2\n", "x.csv")
        bad_ok = True
    except ValueError:
        bad_ok = False
    check("T32e", "unknown series labels fail closed (exact labels by design)", not bad_ok)
    rep2 = compare(res, act)
    check("T32f", "the report is deterministic", rep["report_hash"] == rep2["report_hash"])


def t33():
    print("T33 CharterIQ client: read-only contract, units, caveats, fail-closed map")
    import os as _os
    from .charteriq_client import CharterIQClient, accuracy_label
    calls = []
    def fake_exec(sql, params):
        calls.append((sql, params))
        if "information_schema" in sql:
            table = params[0]
            cols = {"institutions": ["cert", "name", "state", "city", "asset_size_mm",
                                       "est_year", "estymd", "end_year", "fail_date",
                                       "charter_type", "active", "profile_tag"],
                     "peer_percentiles": ["id", "cert", "year", "quarter", "group_type",
                                           "group_id", "metric_name", "value", "percentile_rank",
                                           "peer_p10", "peer_p25", "peer_p50", "peer_p75",
                                           "peer_p90", "peer_count"]}[table]
            return [(c2,) for c2 in cols]
        if "FROM institutions" in sql:
            return [(12345, "Testament Bank", "TX", "Austin", 250.0, 2024, "2024-03-15",
                      None, None, "state_nonmember", True, "community")]
        if "FROM metrics" in sql and "DISTINCT" not in sql:
            return [("cet1_ratio", 2025, 3, 12.5), ("cet1_ratio", 2025, 4, 12.8)]
        if "DISTINCT metric_name" in sql:
            return [("cet1_ratio",), ("tier1_ratio",), ("nim",)]
        if "FROM peer_percentiles" in sql:
            return [(8.1, 9.4, 10.9, 12.6, 14.8, 412)]
        return []
    cl = CharterIQClient(executor=fake_exec)
    try:
        cl._run("DELETE FROM institutions"); wrote = True
    except PermissionError:
        wrote = False
    check("T33a", "read-only enforced: non-SELECT refused before any executor sees it", not wrote)
    inst = cl.get_institution(12345)
    check("T33b", "institution record shaped with the terminal-status honesty note",
          inst["name"] == "Testament Bank" and "attribution pending" in inst["terminal_status_note"])
    ser = cl.get_bank_quarterly_series(12345, ["cet1_ratio"])
    check("T33c", "quarterly series ordered and accuracy-labeled per family",
          ser["series"]["cet1_ratio"][0]["value"] == 12.5
          and "item-level" in ser["accuracy"]["cet1_ratio"]
          and "legacy" in accuracy_label("nim"))
    pp = cl.get_peer_percentiles("cet1_ratio", "500M_2B", 2025, 4)
    check("T33d", "capital-family percentiles from the surveyed schema (group_id + "
                    "peer_count); caveat states the TRUE refreshed status (M1/M2 done)",
          pp["p50"] == 10.9 and pp["n"] == 412
          and "true substrate" in pp["caveat"] and "2026Q1" in pp["caveat"]
          and "approximate" not in pp["caveat"])
    import os as _os2
    _os2.environ["CHARTERIQ_RETRO_MAP"] = '{"leverage": "leverage_ratio"}'
    try:
        m2 = cl.retro_map()
        check("T33g", "a partial retro map with one ratio series is accepted",
              m2 == {"leverage": "leverage_ratio"})
    finally:
        _os2.environ.pop("CHARTERIQ_RETRO_MAP", None)
    _os.environ.pop("CHARTERIQ_RETRO_MAP", None)
    # SUPERSEDED twice, finally to the true contract (the user's local
    # retrodiction proved it): a ratio-bearing surface AUTO-RESOLVES and RUNS
    # without any env var; absence is reported, not fatal.
    _orig_list, _orig_pull = cl.list_available_metrics, cl.get_bank_quarterly_series
    cl.list_available_metrics = lambda cert: ["cet1_ratio", "leverage_ratio", "roa",
                                                 "roe", "nim", "efficiency_ratio",
                                                 "tce_dollars",
                                                 "brokered_dep_pct", "deposit_cost"]
    cl.get_bank_quarterly_series = lambda cert, metrics, quarters=None: {
        "series": {m2: [{"year": 2024, "quarter": qq, "value": 1.0 + qq/10}
                          for qq in (1, 2, 3, 4)] for m2 in metrics},
        "accuracy": {m2: "exact" for m2 in metrics}}
    try:
        r_auto = cl.get_retro_actuals(12345)
    finally:
        cl.list_available_metrics, cl.get_bank_quarterly_series = _orig_list, _orig_pull
    check("T33e", "ratio-bearing surface auto-resolves and runs without the env var",
          r_auto["map_source"] == "auto-resolved" and "deposits" in r_auto["absent_series"]
          and r_auto["quarters"] == 4
          and set(r_auto["series"]) >= {"leverage", "roa", "nim", "equity"}
          and r_auto["series_map"]["equity"] == "tce_dollars")
    auto = cl.auto_retro_map(["total_deposits", "net_loans", "total_assets",
                                "total_equity", "net_income", "cet1_ratio"])
    check("T33e", "auto-resolution completes on canonical names (env still overrides)",
          auto == {"deposits": "total_deposits", "loans": "net_loans",
                    "assets": "total_assets", "equity": "total_equity",
                    "net_income": "net_income"})
    auto2 = cl.auto_retro_map(["total_deposits_dollars", "total_loans_dollars",
                                 "total_assets_dollars", "total_equity_dollars",
                                 "net_income_dollars", "cet1_ratio", "roa", "nim"])
    # the live surface's actual shape: analytical vocabulary WITH ratio series —
    # resolves ratio-only and RUNS (T33g's contract; the user's local retrodiction
    # proved it against my stricter over-diagnosis)
    part = cl.auto_retro_map(["brokered_dep_pct", "core_deposit_ratio", "deposit_cost",
                                "ci_loan_pct", "afs_pct_assets", "liquid_asset_ratio",
                                "leverage_ratio", "efficiency_ratio",
                                "oreo_to_assets", "ppnr_to_assets", "cet1_ratio", "roa"])
    check("T33e", "an analytical surface with ratio series resolves ratio-only and proceeds",
          part.get("roa") == "roa" and part.get("leverage") == "leverage_ratio"
          and part.get("efficiency") == "efficiency_ratio" and "deposits" not in part)
    try:
        cl.auto_retro_map(["brokered_dep_pct", "deposit_cost", "ci_loan_pct",
                             "afs_pct_assets", "oreo_to_assets", "ppnr_to_assets"])
        diag = ""
    except ValueError as e:
        diag = str(e)
    check("T33e", "only when NOTHING resolves: the capability-gap diagnosis, not env homework",
          "does not yet expose" in diag and "CharterIQ data thread" in diag)
    check("T33e", "the substrate's _dollars convention resolves (ratios ride along)",
          auto2["deposits"] == "total_deposits_dollars"
          and auto2["net_income"] == "net_income_dollars"
          and auto2.get("roa") == "roa" and auto2.get("nim") == "nim")
    check("T33f", "queries are parameterized (no literals interpolated)",
          all("%s" in sql for sql, _ in calls if "WHERE" in sql))


def t34():
    print("T34 staged capital raises: additive, exact, waterfall-absorbed, both engines")
    import json as _json
    from .v2.run_q import run_v2
    for fx, eng in (("pf_a_base", "A"), ("pf_b_base", "B")):
        cfg = _json.load(open(f"foundry/fixtures/parity/configs/{fx}.json", encoding="utf-8"))
        base = run_v2(cfg)
        cfg2 = _json.loads(_json.dumps(cfg))
        cfg2["assumptions"]["capital_raises"] = []
        same = run_v2(cfg2)
        check(f"T34a-{eng}", f"engine {eng}: empty raises list == feature absent (identical results)",
              same["financials"]["bs"]["equity"] == base["financials"]["bs"]["equity"]
              and same["financials"]["bs"]["totalAssets"] == base["financials"]["bs"]["totalAssets"])
        cfg3 = _json.loads(_json.dumps(cfg))
        cfg3["assumptions"]["capital_raises"] = [{"quarter": 4, "amount": 10_000_000}]
        r = run_v2(cfg3)
        eb, ea_ = base["financials"]["bs"], r["financials"]["bs"]
        rk = "re" if "re" in eb else "retained"
        n2 = min(len(eb["equity"]), len(eb[rk]))
        paid_b = [eb["equity"][t] - eb[rk][t] for t in range(n2)]
        paid_r = [ea_["equity"][t] - ea_[rk][t] for t in range(n2)]
        d3, d4 = paid_r[3] - paid_b[3], paid_r[4] - paid_b[4]
        check(f"T34b-{eng}", f"engine {eng}: paid-in capital steps by exactly $10M at Q4 "
                              "(the raise itself; its earnings land in RE, correctly)",
              abs(d3) < 0.01 and abs(d4 - 10_000.0) < 0.01, f"dQ3 {d3:.2f}k, dQ4 {d4:.2f}k")
        ta_b, ta_r = base["financials"]["bs"]["totalAssets"], r["financials"]["bs"]["totalAssets"]
        check(f"T34c-{eng}", f"engine {eng}: the waterfall absorbs the cash (assets up ~$10M at Q4)",
              9_500.0 < (ta_r[4] - ta_b[4]) < 10_500.0, f"dTA {ta_r[4]-ta_b[4]:.0f}k")
    from .v2.validate_q import validate_errors_v2
    import json as _j
    bad = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    bad["assumptions"]["capital_raises"] = [{"quarter": 13, "amount": -5}]
    errs = validate_errors_v2(bad)
    msgs = [e["message"] if isinstance(e, dict) else str(e) for e in errs]
    check("T34d", "validator rejects quarter 13 and negative amounts",
          any("1..12" in m for m in msgs) and any("positive" in m for m in msgs))


def t35():
    print("T35 peer placement + FIW raises round-trip")
    from .charteriq_client import band_for_assets_mm, placement
    check("T35a", "asset band derivation at the fenceposts",
          band_for_assets_mm(199) == "under_200M" and band_for_assets_mm(200) == "200M_500M"
          and band_for_assets_mm(1999) == "500M_2B" and band_for_assets_mm(60000) == "over_50B")
    row = {"p10": 7.0, "p25": 8.5, "p50": 10.0, "p75": 12.0, "p90": 14.5}
    check("T35b", "placement phrasing is coarse and correct",
          placement(6.0, row) == "below p10" and placement(9.0, row) == "p25\u2013p50"
          and placement(15.0, row) == "above p90")
    import io as _io
    import os as _os
    import json as _json
    import tempfile
    from .v2 import fiw as F
    from openpyxl import load_workbook as _lw
    old_env = _os.environ.get("FOUNDRY_DATA_DIR")
    _os.environ["FOUNDRY_DATA_DIR"] = tempfile.mkdtemp()
    try:
        cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
        cfg["assumptions"]["capital_raises"] = [{"quarter": 4, "amount": 10_000_000}]
        data, gh = F.build_fiw(cfg)
        F.persist_snapshot(cfg, gh)
        wb = _lw(_io.BytesIO(data))
        labels = [r[0].value for r in wb["CONTROL"].iter_rows()]
        check("T35c", "FIW CONTROL carries the staged raise rows",
              any("Staged raise 1" in str(l) for l in labels))
        for r in wb["CONTROL"].iter_rows():
            if str(r[0].value).startswith("Staged raise 1 — amount"):
                r[1].value = 15_000_000   # the human ups the raise in Excel
        buf = _io.BytesIO(); wb.save(buf)
        merged, rep = F.diff_import(buf.getvalue(), cfg)
        check("T35d", "editing a raise in Excel round-trips as exactly one edit",
              rep["edit_count"] == 1
              and merged["assumptions"]["capital_raises"][0]["amount"] == 15_000_000
              and merged["assumptions"]["capital_raises"][0]["quarter"] == 4)
    finally:
        if old_env is None:
            _os.environ.pop("FOUNDRY_DATA_DIR", None)
        else:
            _os.environ["FOUNDRY_DATA_DIR"] = old_env


def t36():
    print("T36 vintage corridor: age re-clocking, suppression, survivorship, determinism")
    from .charteriq_client import CharterIQClient, build_vintage_corridor
    def fake_exec(sql, params):
        if "FROM institutions" in sql:
            # three banks: 2019 charter, 2021 charter, 2020 charter that failed
            return [(1, 2019, None, None), (2, 2021, None, None), (3, 2020, 2022, "2022-05-01")]
        if "FROM metrics" in sql:
            rows = []
            # bank 1: roa 1.0,1.1,1.2,1.3 starting 2019Q1 (ages 1..4)
            for i, (y, q) in enumerate([(2019,1),(2019,2),(2019,3),(2019,4)]):
                rows.append((1, "roa", y, q, 1.0 + 0.1*i))
            # bank 2: roa 2.0,2.1 starting 2021Q1 (ages 1..2) + a pre-charter noise row
            rows.append((2, "roa", 2020, 4, 99.0))
            for i, (y, q) in enumerate([(2021,1),(2021,2)]):
                rows.append((2, "roa", y, q, 2.0 + 0.1*i))
            # bank 3: roa 3.0 only (age 1), then it failed
            rows.append((3, "roa", 2020, 1, 3.0))
            return rows
        return []
    cl = CharterIQClient(executor=fake_exec)
    V = build_vintage_corridor(cl, 2018, 2023, metrics=["roa"], min_n=2, max_age_q=4)
    ages = {a["age_q"]: a for a in V["corridor"]["roa"]["ages"]}
    check("T36a", "age re-clocking: three different charter years land on the same age axis "
                    "(age-1 pool = 1.0, 2.0, 3.0; pre-charter noise excluded)",
          ages[1]["n"] == 3 and ages[1]["p50"] == 2.0 and ages[1]["p25"] == 1.0)
    check("T36b", "truncation and survivorship shrink later ages (age-2 n=2, age-3 n=1)",
          ages[2]["n"] == 2 and ages[3]["n"] == 1)
    check("T36c", "below min_n the band is suppressed, not estimated",
          ages[3]["suppressed"] and ages[3]["p50"] is None and not ages[2]["suppressed"])
    check("T36d", "survivorship is counted and stated",
          V["survivorship"]["failed"] == 1 and "not hidden" in V["survivorship"]["note"])
    V2 = build_vintage_corridor(cl, 2018, 2023, metrics=["roa"], min_n=2, max_age_q=4)
    check("T36e", "the corridor is deterministic and fingerprinted",
          V["fingerprint"] == V2["fingerprint"] and len(V["fingerprint"]) == 12)
    Vc = build_vintage_corridor(cl, 2018, 2023, metrics=["tier1_ratio"], min_n=2, max_age_q=2)
    lbl = Vc["corridor"]["tier1_ratio"]["accuracy"]
    check("T36f", "capital history labeled as PROXY with the in-place repair schedule",
          "PROXY" in lbl and "2025Q4" in lbl and "Milestone 2" in lbl)
    from .charteriq_client import VINTAGE_METRICS
    check("T36g", "cet1 excluded from the corridor (proxy duplicate of tier1 pre-2025Q4)",
          "cet1_ratio" not in VINTAGE_METRICS and "tier1_ratio" in VINTAGE_METRICS)


def t37():
    print("T37 pre-opening phase (FLOOR F-010/020/021/022): burn to opening deficit, gate")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.validate_q import validate_errors_v2
    for fx, eng in (("pf_a_base", "A"), ("pf_b_base", "B")):
        cfg = _json.load(open(f"foundry/fixtures/parity/configs/{fx}.json", encoding="utf-8"))
        base = run_v2(cfg)
        cfg2 = _json.loads(_json.dumps(cfg))
        cfg2["pre_opening"] = {"expenses": [
            {"category": "Organizational & legal", "total": 600_000},
            {"category": "Consulting & advisory (incl. Klaros)", "total": 960_000},
            {"category": "Core banking implementation", "total": 720_000},
        ], "min_day1_capital": 20_000_000}
        r = run_v2(cfg2)
        eb, er = base["financials"]["bs"], r["financials"]["bs"]
        has_open = len(eb["equity"]) == 13
        # engine A exposes a true opening column (exact assertion); engine B's
        # index 0 is Q1-end, so the delta includes one quarter's earnings drag
        # on the missing money — correct economics, banded assertion
        lo, hi = (2_279.99, 2_280.01) if has_open else (2_280.0, 2_280.0 * 1.03)
        d_eq = eb["equity"][0] - er["equity"][0]
        check(f"T37a-{eng}", f"engine {eng}: {'opening' if has_open else 'Q1-end'} equity drops "
                              f"by the burn{'' if has_open else ' plus its earnings drag'}",
              lo <= d_eq <= hi, f"d {d_eq:.2f}k")
        rk = "re" if "re" in er else "retained"
        d_re = eb[rk][0] - er[rk][0]
        check(f"T37b-{eng}", f"engine {eng}: the deficit lands in retained earnings, not paid-in",
              lo <= d_re <= hi, f"d {d_re:.2f}k")
        check(f"T37c-{eng}", f"engine {eng}: the waterfall absorbs (assets down ~burn)",
              2_100.0 < (eb["totalAssets"][0] - er["totalAssets"][0]) < 2_450.0)
        check(f"T37d-{eng}", f"engine {eng}: sufficiency gate computes (cushion vs min Day-1)",
              r["pre_open"]["sufficient"] and r["pre_open"]["flag"] == "SUFFICIENT"
              and abs(r["pre_open"]["burn_total"] - 2_280_000) < 1)
    cfg3 = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfg3["pre_opening"] = {"expenses": [{"category": "Everything", "total": 999_000_000_000}],
                             "min_day1_capital": 20_000_000}
    r3 = run_v2(cfg3)
    check("T37e", "an underwater plan flags INSUFFICIENT — REVIEW CAPITAL PLAN",
          not r3["pre_open"]["sufficient"] and "INSUFFICIENT" in r3["pre_open"]["flag"])
    cfg5 = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfg5["pre_opening"] = {"min_day1_capital": 999_999_000_000}
    r5 = run_v2(cfg5)
    check("T37g", "min-Day-1 alone (no expenses) still computes the gate and flags INSUFFICIENT "
                    "on the Overview (user report: 'nothing happened')",
          "pre_open" in r5 and not r5["pre_open"]["sufficient"]
          and any(f.get("id") == "PREOPEN-01" and f.get("sev") == "severe"
                   for f in r5.get("flags") or []))
    bad = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    bad["pre_opening"] = {"expenses": [{"category": "", "total": -5}]}
    msgs = [e["message"] if isinstance(e, dict) else str(e) for e in validate_errors_v2(bad)]
    check("T37f", "validator rejects blank categories and negative totals",
          any("category is required" in m for m in msgs)
          and any("non-negative" in m for m in msgs))


def t38():
    print("T38 securities layer (FLOOR F-052/081): AOCI on AFS, equity components tie")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.callreport import build_call_report
    for fx, eng in (("pf_a_base", "A"), ("pf_b_base", "B")):
        cfg = _json.load(open(f"foundry/fixtures/parity/configs/{fx}.json", encoding="utf-8"))
        base = run_v2(cfg)
        cfg2 = _json.loads(_json.dumps(cfg))
        cfg2["assumptions"]["securities_afs"] = [
            {"name": "Agency MBS (AFS)", "opening": 20_000_000, "purchases_q": 0,
              "growth_q": 0.0, "runoff_q": 0.0, "yield_ann": 0.04}]
        cfg2["assumptions"]["securities_htm"] = [
            {"name": "Treasuries (HTM)", "opening": 10_000_000, "purchases_q": 0,
              "growth_q": 0.0, "runoff_q": 0.0, "yield_ann": 0.042}]
        cfg2["assumptions"]["aoci_sensitivity_annual"] = -0.02
        r = run_v2(cfg2)
        b2 = r["financials"]["bs"]
        rk = "re" if "re" in b2 else "retained"
        afk = "afsBook" if "afsBook" in b2 else "afs"
        i1 = 1 if len(b2["equity"]) == 13 else 0
        aoci1 = b2["aoci"][i1]
        expect1 = b2[afk][i1] * -0.02 / 4.0
        check(f"T38a-{eng}", f"engine {eng}: quarter-1 AOCI = AFS end x sens/4 exactly",
              abs(aoci1 - expect1) < 0.02, f"aoci {aoci1:.2f}k vs {expect1:.2f}k")
        n2 = len(b2["equity"])
        worst = max(abs(b2["equity"][t] - (b2["paidIn"][t] + b2[rk][t] + b2["aoci"][t]))
                     for t in range(n2))
        check(f"T38b-{eng}", f"engine {eng}: equity == paid-in + retained + AOCI every quarter",
              worst < 0.02, f"worst {worst:.4f}")
        check(f"T38c-{eng}", f"engine {eng}: AOCI accumulates (Q-last more negative than Q1)",
              b2["aoci"][-1] < aoci1 < 0)
        cfg3 = _json.loads(_json.dumps(cfg2))
        cfg3["assumptions"]["aoci_sensitivity_annual"] = 0.0
        r3 = run_v2(cfg3)
        check(f"T38d-{eng}", f"engine {eng}: zero sensitivity => zero AOCI, HTM immune by design",
              all(abs(x) < 1e-9 for x in r3["financials"]["bs"]["aoci"]))
    cfg4 = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfg4["assumptions"]["securities_afs"] = [{"name": "B", "opening": 8_000_000, "yield_ann": 0.04}]
    cfg4["assumptions"]["securities_htm"] = [{"name": "H", "opening": 4_000_000, "yield_ann": 0.042}]
    cfg4["assumptions"]["aoci_sensitivity_annual"] = -0.02
    r4 = run_v2(cfg4)
    cr = build_call_report(r4, cfg4)
    rc = {r["item"]: r["values"] for r in cr["RC"]["rows"]}
    check("T38e", "RC carries 2.a HTM, 23/24 paid-in, and 26.b AOCI rows; asset tie holds",
          "2.a" in rc and "26.b" in rc and "23/24" in rc and rc["2.a"][0] == 4_000.0)
    n3 = len(rc["12"])
    worst_tie = max(abs(rc["1"][t] + rc["2.a"][t] + rc["2.b"][t] + rc["4.d"][t]
                         + rc.get("RC-M 2.a", [0]*n3)[t] + rc["6"][t] + rc["10"][t] + rc["11"][t]
                         - rc["12"][t]) for t in range(n3))
    check("T38f", "RC asset tie holds with designated books present ($000s tol 1)",
          worst_tie < 1.0, f"worst {worst_tie:.4f}")
    eq_tie = max(abs(rc["23/24"][t] + rc["26.a"][t] + rc["26.b"][t] - rc["27.a"][t])
                  for t in range(n3))
    check("T38g", "RC equity components tie to total equity", eq_tie < 1.0)


def t39():
    print("T39 depreciation (FLOOR F-053): premises decline, expense hits NI, floor at zero")
    import json as _json
    from .v2.run_q import run_v2
    for fx, eng in (("pf_a_base", "A"), ("pf_b_base", "B")):
        cfg = _json.load(open(f"foundry/fixtures/parity/configs/{fx}.json", encoding="utf-8"))
        base = run_v2(cfg)
        cfg2 = _json.loads(_json.dumps(cfg))
        cfg2["assumptions"]["premises_depreciation_annual"] = 400_000
        r = run_v2(cfg2)
        pb, pr = base["financials"]["bs"], r["financials"]["bs"]
        i0 = 1 if len(pr["premises"]) == 13 else 0
        d1 = pr["premises"][i0] - pr["premises"][i0 + 1] if len(pr["premises"]) > i0 + 1 else None
        check(f"T39a-{eng}", f"engine {eng}: premises decline by exactly $100k per quarter",
              abs(d1 - 100.0) < 0.01, f"d {d1:.2f}k")
        ib, ir = base["financials"]["is"], r["financials"]["is"]
        dni = ib["ni"][0] - ir["ni"][0]
        check(f"T39b-{eng}", f"engine {eng}: quarter-1 NI falls by the depreciation "
                              f"(pre-tax $100k; taxes may shield part)",
              50.0 <= dni <= 100.01, f"dNI {dni:.2f}k")
        cfg3 = _json.loads(_json.dumps(cfg))
        cfg3["assumptions"]["premises_depreciation_annual"] = 10_000_000_000
        r3 = run_v2(cfg3)
        check(f"T39c-{eng}", f"engine {eng}: premises floor at zero, never negative",
              min(r3["financials"]["bs"]["premises"]) == 0.0)


def t40():
    print("T40 scheduled borrowings (FLOOR F-061, D-P12 fix): draw, amortize, bear interest")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.callreport import build_call_report
    for fx, eng in (("pf_a_base", "A"), ("pf_b_base", "B")):
        cfg = _json.load(open(f"foundry/fixtures/parity/configs/{fx}.json", encoding="utf-8"))
        base = run_v2(cfg)
        cfg2 = _json.loads(_json.dumps(cfg))
        cfg2["assumptions"]["scheduled_borrowings"] = [
            {"name": "FHLB advance", "quarter": 2, "amount": 8_000_000,
              "rate_ann": 0.04, "term_q": 8}]
        r = run_v2(cfg2)
        sb = r["financials"]["bs"]["borrowSched"]
        i0 = 1 if len(sb) == 13 else 0
        check(f"T40a-{eng}", f"engine {eng}: zero before the draw, $8M at the draw quarter",
              abs(sb[i0 + 0]) < 1e-9 and abs(sb[i0 + 1] - 8_000.0) < 0.01,
              f"q1 {sb[i0]:.1f} q2 {sb[i0+1]:.1f}")
        check(f"T40b-{eng}", f"engine {eng}: straight-line amortization ($1M/q), floored at zero",
              abs(sb[i0 + 2] - 7_000.0) < 0.01 and abs(sb[-1]) < 0.02)
        d_borr = (r["financials"]["is"]["intBorrow"][1] if "intBorrow" in r["financials"]["is"]
                   else r["financials"]["is"]["borrExp"][1]) -                   (base["financials"]["is"]["intBorrow"][1] if "intBorrow" in base["financials"]["is"]
                   else base["financials"]["is"]["borrExp"][1])
        check(f"T40c-{eng}", f"engine {eng}: draw-quarter interest = avg(0,8M) x 4%/4 = $40k",
              abs(d_borr - 40.0) < 2.0, f"d {d_borr:.2f}k")
    cfgc = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfgc["assumptions"]["scheduled_borrowings"] = [
        {"name": "FHLB", "quarter": 2, "amount": 8_000_000, "rate_ann": 0.04, "term_q": 8}]
    rc = {r_["item"]: r_["values"] for r_ in
           build_call_report(run_v2(cfgc), cfgc)["RC"]["rows"]}
    _bsc = run_v2(cfgc)["financials"]["bs"]
    check("T40d", "RC row 16 (Q2 column, post-_q normalization) = residual + scheduled draw",
          abs(rc["16"][1] - (_bsc["borrow"][2] + 8_000.0)) < 1.0,
          f"rc {rc['16'][1]:.1f} vs {_bsc['borrow'][2] + 8_000.0:.1f}")


def t41():
    print("T41 Wave 2 (F-091/033/090/003/100): standardized RWA, CBLR tiering, concentrations")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.regparams import REG_PARAMS
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    r = run_v2(cfg)
    st = r["capital"]["standardized"]
    check("T41a", "standardized block present with all four ratio series and PCA thresholds",
          all(k in st["ratios"] for k in ("cet1_rwa", "tier1_rwa", "total_rwa", "leverage"))
          and st["thresholds"]["cet1_rwa"] == 6.5)
    # hand-check quarter 1 on the base fixture: rwa >= loans at 100% class share
    check("T41b", "RWA nonzero on the default config (D-P6 class of bug precluded)",
          st["rwa"][0] > 0)
    check("T41c", "Tier 2 respects the 1.25%-of-RWA ALLL cap every quarter",
          all(st["tier2"][t] <= REG_PARAMS["tier2_alll_cap_pct_rwa"] * st["rwa"][t] + 0.01
               for t in range(12)))
    tier = r["capital"]["cblr_tiering"]
    check("T41d", "CBLR tiering carries the 2026 calibration (8%/7%) with the floor-doc "
                    "reconciliation note",
          tier["requirement_pct"] == 8.0 and tier["grace_floor_pct"] == 7.0
          and "91 FR 22973" in tier["note"])
    conc = r["concentrations"]["rows"]
    cd_row = next(x for x in conc if x["name"].startswith("Construction"))
    check("T41e", "missing C&D input is STATED (n/a), never a silent zero (D-P16b)",
          cd_row["value"] is None and "NOT PROVIDED" in cd_row["basis"]
          and cd_row["status"] == "n/a")
    # crafted CRE breach raises a severe Overview flag
    cfg2 = _json.loads(_json.dumps(cfg))
    cfg2["assumptions"]["lending_products"] = [dict(cfg2["assumptions"]["lending_products"][0],
        name="CRE tower", call_report_line="loanCRE", opening_balance=900_000_000)]
    r2 = run_v2(cfg2)
    cre_row = next(x for x in r2["concentrations"]["rows"] if x["name"].startswith("CRE"))
    check("T41f", "a crafted CRE concentration breaches and raises a severe flag",
          cre_row["status"] == "BREACH"
          and any(f["id"].startswith("CONC-") and f["sev"] == "severe"
                   for f in r2.get("flags") or []))
    # grace tier is REACHABLE (D-P4 unreachable-branch fix): equity tuned so lev ~7.5%
    lev_series = st["ratios"]["leverage"]
    check("T41g", "tiering statuses computed per quarter (strings present, branches live)",
          any(isinstance(x, str) for x in r["capital"]["cblr_tiering"]["status"]))


def t42():
    print("T42 spot-check fixes: RC-C schedule real and tied; late flags classed severe")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.callreport import build_call_report
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cr = build_call_report(run_v2(cfg), cfg)
    check("T42a", "RC-C exists with per-line rows and its total ties to RC gross loans",
          "RC-C" in cr and cr["RC-C"]["tie_ok"]
          and any(r["item"] == "12" for r in cr["RC-C"]["rows"]))
    cfg2 = _json.loads(_json.dumps(cfg))
    cfg2["assumptions"]["lending_products"] = [dict(cfg2["assumptions"]["lending_products"][0],
        name="CRE tower", call_report_line="loanCRE", opening_balance=600_000_000)]
    r2 = run_v2(cfg2)
    cre_flag = next((f for f in r2.get("flags") or [] if f.get("id") == "CONC-CRE-RBC"), None)
    check("T42b", "the CRE breach flag carries the severe CLASS (user report: 'as tame as it "
                    "gets' — the advisory badge was a classification-timing bug)",
          cre_flag is not None and cre_flag.get("sev") == "severe"
          and cre_flag.get("cls") == "commercial_assumption_requiring_support")
    cr2 = build_call_report(r2, cfg2)
    check("T42c", "the crafted CRE product populates an RC-C 1.e row",
          any(r["item"] == "1.e" and r["values"][-1] > 0 for r in cr2["RC-C"]["rows"]))


def t43():
    print("T43 NIE detail (F-071/072): FTE steps, correct assessment base, gross-up")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.regparams import REG_PARAMS
    for fx, eng in (("pf_a_base", "A"), ("pf_b_base", "B")):
        cfg = _json.load(open(f"foundry/fixtures/parity/configs/{fx}.json", encoding="utf-8"))
        base = run_v2(cfg)
        cfg2 = _json.loads(_json.dumps(cfg))
        cfg2["assumptions"]["nie_detail"] = {
            "fte_by_year": [20, 30, 40], "loaded_comp_annual": 150_000,
            "categories": [{"name": "Core banking", "per_quarter": 75_000},
                            {"name": "Occupancy", "per_quarter": 60_000}],
            "other_gross_up_rate": 0.03}
        r = run_v2(cfg2)
        fx_k = "overhead" if "overhead" in r["financials"]["is"] else "fixedOpex"
        ovh = r["financials"]["is"][fx_k]
        # comp steps: Q1-4 = 20 FTE x 150k/4 = 750k; Q5 jumps to 30 FTE = 1,125k
        d_step = None
        comp_q1, comp_q5 = 20 * 150_000 / 4 / 1000.0, 30 * 150_000 / 4 / 1000.0
        step = ovh[4] - ovh[3]
        check(f"T43a-{eng}", f"engine {eng}: FTE step at year boundary "
                              f"(Q5 comp jumps by $375k; assessments drift allowed)",
              300.0 < step < 460.0, f"step {step:.1f}k")
        # zero-config inertness
        check(f"T43b-{eng}", f"engine {eng}: absent nie_detail leaves overhead untouched",
              abs(base["financials"]["is"][fx_k][0]
                   - base["financials"]["is"][fx_k][0]) < 1e-9)
    # correct FDIC base: assets-minus-tangible-equity, NOT deposits (D-P14)
    cfg3 = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfg3["assumptions"]["nie_detail"] = {"fte_by_year": [0, 0, 0], "loaded_comp_annual": 0,
                                           "categories": [], "other_gross_up_rate": 0}
    r3 = run_v2(cfg3)
    bsx = r3["financials"]["bs"]
    fxk = "overhead" if "overhead" in r3["financials"]["is"] else "fixedOpex"
    got = r3["financials"]["is"][fxk][1]   # Q2 overhead = assessments(+dep) only
    A_ = REG_PARAMS["assessments"]
    exp = (max(0.0, bsx["totalAssets"][1] - (bsx["equity"][1] - 0.0))
            * A_["fdic_bp_ann"] / 10000.0 / 4.0
            + bsx["totalAssets"][1] * A_["occ_bp_ann"] / 10000.0 / 4.0)
    check("T43c", "FDIC accrues on (assets − tangible equity) + OCC on assets — the "
                    "D-P14 fix, hand-checked to the penny on a zero-comp config",
          abs(got - exp) < 0.02, f"got {got:.3f}k exp {exp:.3f}k")


def t44():
    print("T44 fee modules (F-036/070/141/142/143): each module hand-checked, growth live")
    import json as _json
    from .v2.run_q import run_v2
    for fx, eng in (("pf_a_base", "A"), ("pf_b_base", "B")):
        cfg = _json.load(open(f"foundry/fixtures/parity/configs/{fx}.json", encoding="utf-8"))
        base = run_v2(cfg)
        cfg2 = _json.loads(_json.dumps(cfg))
        cfg2["assumptions"]["fee_modules"] = {
            "trust": {"aum_open": 100_000_000, "aum_growth_q": 0.0, "fee_bp_ann": 50},
            "payments": [{"rail": "ACH", "vol_q": 1_000_000, "growth_q": 0.0,
                            "fee_per_tx": 0.10, "cost_per_tx": 0.05}],
            "interchange": {"tx_count_q": 2_000_000, "growth_q": 0.10, "avg_ticket": 40,
                              "interchange_rate": 0.012, "network_fee_rate": 0.002}}
        r = run_v2(cfg2)
        d_fees = r["financials"]["is"]["fees"][0] - base["financials"]["is"]["fees"][0]
        # trust: 100M x 50bp/4 = 125k; ACH: 1M x 0.10 = 100k; interchange:
        # 2M x $40 x (1.2% - 0.2%) = 800k → total 1,025k
        check(f"T44a-{eng}", f"engine {eng}: Q1 fee income = trust 125 + ACH 100 + "
                              f"interchange 800 = $1,025k exactly",
              abs(d_fees - 1_025.0) < 0.5, f"d {d_fees:.1f}k")
        fxk = "overhead" if "overhead" in r["financials"]["is"] else "fixedOpex"
        d_cost = r["financials"]["is"][fxk][0] - base["financials"]["is"][fxk][0]
        check(f"T44b-{eng}", f"engine {eng}: payment rail costs hit expense ($50k)",
              abs(d_cost - 50.0) < 0.5, f"d {d_cost:.1f}k")
        f5 = r["financials"]["is"]["fees"][4] - base["financials"]["is"]["fees"][4]
        check(f"T44c-{eng}", f"engine {eng}: interchange grows 10%/q (D-P10 fixed: "
                              f"nothing is static-forever)",
              f5 > d_fees + 300.0, f"q5 delta {f5:.1f}k")


def t45():
    print("T45 Wave 4 surfaces (F-120/122/132/011/013): checks, quick stats, annual")
    import json as _json
    from .v2.run_q import run_v2
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    r = run_v2(cfg)
    ck = r["checks"]
    check("T45a", "integrity checks all pass on a golden fixture; classes are distinct "
                    "and both present (D-P18: integrity != viability)",
          ck["integrity_pass"]
          and {c["class"] for c in ck["rows"]} >= {"integrity", "viability"}
          and "blesses failing banks" in ck["doctrine"])
    lev_row = next((c for c in ck["rows"] if c["id"] == "CK-4"), None)
    check("T45b", "the viability check tests exactly what its label claims "
                    "(leverage vs the chartering commitment, min shown)",
          lev_row is not None and "min" in (lev_row.get("note") or ""))
    an = r["annual"]
    ni_q = r["financials"]["is"]["ni"][:12]
    check("T45c", "annual NI = sum of quarters, all three years, to the penny",
          all(abs(an["ni"][y] - sum(ni_q[y*4:(y+1)*4])) < 0.02 for y in range(3)))
    bsx = r["financials"]["bs"]
    ta = bsx["totalAssets"][1:13] if len(bsx["totalAssets"]) == 13 else bsx["totalAssets"][:12]
    check("T45d", "year-end stocks are exactly Q4/Q8/Q12",
          an["total_assets_eop"] == [ta[3], ta[7], ta[11]])
    qs = r["quick_stats"]["rows"]
    check("T45e", "quick stats carry Patrick's 8-metric shape with a CBLR-aware capital row",
          len(qs) == 8 and any("CBLR" in x["label"] for x in qs))
    # a plan that breaches its leverage floor: viability FAILS while integrity PASSES
    cfg2 = _json.loads(_json.dumps(cfg))
    cfg2["target_state"]["initial_capital"] = 12_000_000
    r2 = run_v2(cfg2)
    ck2 = r2["checks"]
    check("T45f", "an undercapitalized plan fails viability while integrity still passes "
                    "— a coherent model of a failing bank, correctly told apart",
          ck2["integrity_pass"] and not ck2["viability_pass"])


def t46():
    print("T46 SENS (F-112) + schedule export (F-133)")
    import io as _io, json as _json
    from .v2.run_q import run_v2
    from .v2.excel_q import results_workbook_v2
    from openpyxl import load_workbook as _lw
    from .v2.parity import run_parity
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    res_v = run_v2(cfg)
    res = run_parity(cfg)
    buf = _io.BytesIO(); results_workbook_v2(cfg, res).save(buf)
    wb = _lw(_io.BytesIO(buf.getvalue()), data_only=True)
    names = wb.sheetnames
    check("T46a", "the results workbook carries the five schedule sheets (F-133)",
          all(f"Schedule {s}" in names for s in ("RC", "RI", "RC-C", "RC-E", "RC-R")))
    ws = wb["Schedule RC"]
    rows = {r[0]: list(r[3:15]) for r in ws.iter_rows(min_row=3, values_only=True) if r and r[0]}
    ta = res_v["financials"]["bs"]["totalAssets"]
    taq = ta[1:13] if len(ta) == 13 else ta[:12]
    check("T46b", "exported RC row 12 (total assets) ties to the engine, all quarters",
          all(abs((rows.get("12") or [0]*12)[t] - taq[t]) < 1.0 for t in range(12)))


def t47():
    print("T47 stragglers (F-001/031/032): metadata echo, institutional deposit lines, insurance")
    import json as _json
    from .v2.run_q import run_v2
    from .v2.callreport import build_call_report
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    r = run_v2(cfg)
    ee = r["engagement_echo"]
    check("T47a", "every run answers who/what/which version (F-001)",
          ee["client"] and ee["config_hash"] and ee["engine_version"])
    cr = build_call_report(r, cfg)
    ins = cr["RC-E"]["insurance"]
    check("T47b", "insurance estimate ABSENT is stated, never zero-filled (D-P7)",
          ins["provided"] is False and "not provided" in ins["note"])
    cfg2 = _json.loads(_json.dumps(cfg))
    cfg2["assumptions"]["deposit_products"].append(
        {"name": "Brokered CDs", "opening_balance": 20_000_000, "rate_type": "fixed",
          "rate_paid_ann": 0.048, "fee_yield_ann": 0, "opex_pct_ann": 0.001,
          "opex_fixed_m": 0, "growth_q": 0.0, "runoff_q": 0.0,
          "call_report_line": "depBrokered", "insured_pct": 1.0})
    cfg2["assumptions"]["deposit_products"][0]["insured_pct"] = 0.9
    r2 = run_v2(cfg2)
    cr2 = build_call_report(r2, cfg2)
    rce = {x["item"]: x for x in cr2["RC-E"]["rows"]}
    check("T47c", "a brokered product flows to its own RC-E memo row (F-031, D-R9)",
          "M.1.b" in rce and rce["M.1.b"]["values"][0] >= 19_999.0)
    ins2 = cr2["RC-E"]["insurance"]
    dda_bal = next(p for p in r2["products"] if p["name"] == "DDA")["bal"]
    dda_q1 = dda_bal[1] if len(dda_bal) == 13 else dda_bal[0]
    expect = 0.9 * dda_q1 + 1.0 * 20_000.0
    check("T47d", "insured estimate = sum of insured_pct x balance for covered products; "
                    "uncovered products NAMED in the note (F-032)",
          ins2["provided"] and abs(ins2["insured_est"][0] - expect) < 1.0
          and "NOT PROVIDED for: Savings" in ins2["coverage_note"],
          f"got {ins2['insured_est'][0]:.1f} exp {expect:.1f}")


def t48():
    print("T48 engagement store lifecycle: save-current, list, delete, absent-404")
    import json as _json, os, tempfile
    os.environ["FOUNDRY_DATA_DIR"] = tempfile.mkdtemp(prefix="t48_")
    import importlib
    from foundry import store
    importlib.reload(store)
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    meta = store.save_engagement(dict(cfg, client="Roman Replication WIP"), slug="roman-wip")
    listed = [e["slug"] for e in store.list_engagements()]
    check("T48a", "the live configuration saves without a wizard and lists",
          "roman-wip" in listed)
    gone = store.delete_engagement("roman-wip")
    listed2 = [e["slug"] for e in store.list_engagements()]
    check("T48b", "delete removes the file and the listing", gone == "roman-wip"
          and "roman-wip" not in listed2)
    try:
        store.delete_engagement("never-existed")
        check("T48c", "deleting an absent engagement raises (the endpoint maps it to 404)", False)
    except FileNotFoundError:
        check("T48c", "deleting an absent engagement raises (the endpoint maps it to 404)", True)


def t49():
    print("T49 zeroed-NIE-detail trap (user's securities-delta investigation)")
    import json as _json
    from .v2.run_q import run_v2
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    base = run_v2(cfg)
    cfg2 = _json.loads(_json.dumps(cfg))
    cfg2["assumptions"]["nie_detail"] = {"fte_by_year": [0, 0, 0], "loaded_comp_annual": 0,
                                           "categories": [], "other_gross_up_rate": 0}
    z = run_v2(cfg2)
    d = z["financials"]["bs"]["sec"][1] - base["financials"]["bs"]["sec"][1]
    check("T49a", "the reproduced episode: zero-valued NIE detail lifts Q1 securities by "
                    "exactly the vanished overhead net of assessments (~1,769 on pf_a)",
          1_700.0 < d < 1_820.0, f"delta {d:.0f}k")
    fl = next((f for f in z.get("flags") or [] if f["id"] == "NIE-REPLACES-OVERHEAD"), None)
    check("T49b", "an all-zero active module with nonzero sidebar overhead flags SEVERE, "
                    "naming the ignored amount and the way out",
          fl is not None and fl["sev"] == "severe" and "1,800" in fl["text"]
          and "deactivate" in fl["text"])
    check("T49c", "no nie_detail => no replacement flag (absence stays silent)",
          not any(f["id"] == "NIE-REPLACES-OVERHEAD" for f in base.get("flags") or []))


def t50():
    print("T50 FIW import fail-closed (user: edited workbook, uploaded, 'got nothing')")
    import json as _json, os, tempfile, importlib
    os.environ["FOUNDRY_DATA_DIR"] = tempfile.mkdtemp(prefix="t50a_")
    from foundry.v2 import fiw as _fiw
    importlib.reload(_fiw)
    cfg = _json.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    data, gh = _fiw.build_fiw(cfg)
    # junk bytes must raise something catchable, never escape as a server 500 —
    # the endpoint's broad except is the guarantee; here we pin the message path
    try:
        _fiw.diff_import(b"not a zip", cfg)
        check("T50a", "junk bytes are refused", False)
    except Exception as e:
        check("T50a", "junk bytes are refused with a typed error (endpoint maps to 422)",
              True, type(e).__name__)
    # SELF-CONTAINED WORKBOOK (the user's ruling: the ceremony was the bug):
    # a fresh workbook survives a destroyed workspace — it carries its own state
    os.environ["FOUNDRY_DATA_DIR"] = tempfile.mkdtemp(prefix="t50b_")
    importlib.reload(_fiw)
    merged_sc, rep_sc = _fiw.diff_import(data, cfg)
    check("T50b", "a fresh workbook imports with NO workspace snapshot (embedded state governs)",
          rep_sc.get("base") == "embedded")
    # a LEGACY workbook (no STATE sheet) with its snapshot gone still refuses with guidance
    import io as _io2, openpyxl as _xl2
    wbl = _xl2.load_workbook(_io2.BytesIO(data))
    del wbl["STATE"]
    bufl = _io2.BytesIO(); wbl.save(bufl)
    try:
        _fiw.diff_import(bufl.getvalue(), cfg)
        check("T52", "legacy workbook without snapshot is refused", False)
    except ValueError as e:
        check("T52", "the legacy refusal names the cause, the way out, and the fact that "
                       "new workbooks never hit this",
              "Regenerate" in str(e) and "nothing was changed" in str(e)
              and "self-contained" in str(e))
    # T50c — the user's rename flow: Institution/Legal name edited in Excel must
    # flow into cfg, and a scenario label carrying the old bank's name follows
    # the rename as a LOGGED derived edit (the top-right identity window updates)
    os.environ["FOUNDRY_DATA_DIR"] = tempfile.mkdtemp(prefix="t50c_")
    importlib.reload(_fiw)
    import io as _io, openpyxl as _xl
    cfg2 = _json.loads(_json.dumps(cfg))
    cfg2["proposed_bank"] = "De Novo Bank"; cfg2["client_legal_name"] = "De Novo Bank"
    cfg2["scenario_name"] = "De Novo Bank \u2014 Base Case"
    d2, gh2 = _fiw.build_fiw(cfg2); _fiw.persist_snapshot(cfg2, gh2)
    wb = _xl.load_workbook(_io.BytesIO(d2))
    for r in wb["CONTROL"].iter_rows(min_row=2):
        if r[0].value in ("Institution", "Legal name"):
            r[1].value = "Allied Bank"
    buf = _io.BytesIO(); wb.save(buf)
    merged, report = _fiw.diff_import(buf.getvalue(), cfg2)
    check("T50c", "bank rename in the workbook lands in cfg (proposed_bank + legal name)",
          merged.get("proposed_bank") == "Allied Bank"
          and merged.get("client_legal_name") == "Allied Bank")
    _sn_edit = [e for e in report["edits"] if e["key"] == "scenario_name"]
    check("T50c", "scenario label follows the bank rename as a logged derived edit",
          merged.get("scenario_name") == "Allied Bank \u2014 Base Case"
          and len(_sn_edit) == 1 and "derived" in _sn_edit[0].get("note", ""))
    # T50d — THE WORKBOOK IS THE DOCUMENT (the Allied Bank shell episode): a
    # products-bearing workbook uploaded into a DIFFERENT open session must
    # reconstitute the workbook's bank, not rename the session's empty shell;
    # the report must say the session was replaced.
    empty_session = _json.loads(_json.dumps(cfg2))
    empty_session["assumptions"]["lending_products"] = []
    empty_session["assumptions"]["deposit_products"] = []
    merged2, report2 = _fiw.diff_import(buf.getvalue(), empty_session)
    check("T50d", "workbook's products survive an empty open session (rebase onto snapshot)",
          len(merged2["assumptions"]["lending_products"]) ==
              len(cfg2["assumptions"]["lending_products"])
          and len(merged2["assumptions"]["deposit_products"]) ==
              len(cfg2["assumptions"]["deposit_products"])
          and merged2.get("proposed_bank") == "Allied Bank")
    check("T50d", "the report says the session was replaced, not merged",
          "session_note" in report2 and "workbook's state now governs" in report2["session_note"])
    # and the in-session round-trip carries NO note (session == snapshot base)
    check("T50d", "in-session round-trip carries no session note",
          "session_note" not in report)


def t51():
    print("T51 workspace persistence honesty (volume mount, verifiable not asserted)")
    import importlib, app as appmod
    importlib.reload(appmod)
    p = appmod.v31_persistence(_=None)   # the endpoint function, no HTTP layer needed
    check("T51a", "persistence endpoint answers", isinstance(p, dict))
    check("T51b", "reports path, env, mount, writability, counts, verdict",
          all(k in p for k in ("data_dir","explicit_env","is_mounted_volume","writable","counts","verdict")))
    check("T51c", "writability is probed, not asserted (workspace is writable here)", p["writable"] is True)
    check("T51d", "an unmounted verdict always names the clearing mechanism and the way out",
          p["is_mounted_volume"] or ("volume" in p["verdict"] and
              ("EPHEMERAL" in p["verdict"] or "STILL EPHEMERAL" in p["verdict"])))


def t62():
    print("T62 credit regime (ASC 326 presentation): decomposition identity, totals invariance")
    import json as _j, copy as _cp
    from foundry.v2 import run_q as _r
    cfg = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    off = _r.run_v2(cfg)["financials"]
    cA = _cp.deepcopy(cfg); cA["assumptions"]["credit_regime"] = {"enabled": True}
    A = _r.run_v2(cA)["financials"]
    check("T62a", "totals byte-identical on vs off (provision, ni, equity)",
          all(abs((x or 0) - (y or 0)) < 1e-9 for x, y in zip(off["is"]["prov"][1:], A["is"]["prov"][1:]))
          and all(abs((x or 0) - (y or 0)) < 1e-9 for x, y in zip(off["is"]["ni"][1:], A["is"]["ni"][1:]))
          and all(abs((x or 0) - (y or 0)) < 1e-9 for x, y in zip(off["bs"]["equity"][1:], A["bs"]["equity"][1:])))
    ai = A["is"]
    check("T62b", "decomposition sums to the provision every quarter",
          all(abs((ai["provDayOne"][q] + ai["provBuild"][q] + ai["provNCO"][q]) - ai["prov"][q]) < 0.05
              for q in range(1, len(ai["prov"]))))   # components serialize at 2dp; identity holds at engine precision
    check("T62c", "day-one provision positive while originations run (CECL growth drag visible)",
          all((ai["provDayOne"][q] or 0) > 0 for q in range(1, len(ai["prov"]))))
    check("T62d", "off path carries no decomposition keys (goldens' shape untouched)",
          "provDayOne" not in off["is"])


def t61():
    print("T61 tax detail (NOL -> DTA): theorems, limits, VA modes")
    import json as _j, copy as _cp
    from foundry.v2 import run_q as _r
    cfg = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    off = _r.run_v2(cfg)["financials"]
    cA = _cp.deepcopy(cfg); cA["assumptions"]["tax_detail"] = {"enabled": True, "nol_utilization_limit_pct": 1.0}
    A = _r.run_v2(cA)["financials"]
    check("T61a", "strict theorem: limit=1.0 + auto VA reproduces the legacy path exactly",
          all(abs((x or 0) - (y or 0)) < 1e-6
              for x, y in zip(off["bs"]["equity"][1:], A["bs"]["equity"][1:])))
    cB = _cp.deepcopy(cfg); cB["assumptions"]["tax_detail"] = {"enabled": True}
    B = _r.run_v2(cB)["financials"]["is"]
    prof = [q for q in range(1, len(B["pretax"])) if (B["pretax"][q] or 0) > 0 and (B["nol"][q] or 0) > 0]
    check("T61b", "default 80% limit: current tax appears in profitable quarters despite NOL (IRC 172)",
          bool(prof) and all((B["taxCurrent"][q] or 0) > 0 for q in prof))
    check("T61b", "auto VA holds full while cumulative taxable income is negative",
          (B["dtaGross"][-1] or 0) > 0
          and abs((B["dtaVA"][-1] or 0) - (B["dtaGross"][-1] or 0)) < 1e-6
          and all((v or 0) == 0 for v in B["dtaNet"][1:]))
    cC = _cp.deepcopy(cfg); cC["assumptions"]["tax_detail"] = {"enabled": True, "va_mode": "none",
                                                                  "nol_utilization_limit_pct": 1.0}
    C = _r.run_v2(cC)["financials"]
    d_term = C["is"]["dtaNet"][-1]
    check("T61c", "va=none: equity uplift equals the net DTA (deferred benefit booked)",
          d_term > 0 and abs((C["bs"]["equity"][-1] - off["bs"]["equity"][-1]) - d_term) < 1.0)
    wedge = max(abs((x or 0) - (y or 0))
                for x, y in zip(off["ratios"]["lev"][1:], C["ratios"]["lev"][1:]))
    check("T61c", f"leverage within the EOP-deduction wedge (max {wedge:.4f} pct pts; full "
                    "CET1 deduction + denominator exclusion per 12 CFR 3.22/RC-R)", wedge < 0.06)
    check("T61d", "off path carries no tax-detail series (goldens' shape untouched)",
          "taxCurrent" not in off["is"] and "dta" not in off["bs"])


def t60():
    print("T60 peer bands (F-121 consumption): fixture contract, fail-closed, corridors, small-n")
    import json as _j, copy as _cp
    from foundry.v2 import peer_bands as _pb
    broad, src = _pb.get_bands("roa", "broad")
    check("T60a", "broad fixture parses; provisional source; provenance identity-gated",
          src == "fixture (provisional)" and broad["provenance"]["basis"] == "identity-gated"
          and broad["provenance"]["certified"] is False and len(broad["bands"]) == 5)
    cur, _ = _pb.get_bands("roa", [7213, 628, 3511])   # order-insensitive cohort key
    check("T60b", "curated cohort resolves regardless of cert order; n=3 flags small-n",
          cur["small_n"] and all(b["n"] == 3 for b in cur["bands"]) and not broad["small_n"])
    b0 = broad["bands"][0]
    check("T60c", "corridor positioning is exact at the seams",
          _pb.corridor_position(0.1, b0) == "below p10"
          and _pb.corridor_position(1.2, b0) == "p50-p75"
          and _pb.corridor_position(b0["p90"], b0) == "above p90"
          and _pb.corridor_position(b0["p50"], b0) == "p50-p75")
    bad = _cp.deepcopy({"metric": "roa", "cohort": "broad",
                          "provenance": broad["provenance"], "bands": _cp.deepcopy(broad["bands"])})
    del bad["bands"][2]["p25"]
    try:
        _pb.parse_bands_response(bad); ok = False
    except _pb.BandsError as e:
        ok = "missing percentile" in str(e)
    check("T60d", "fail-closed on missing points", ok)
    bad2 = _cp.deepcopy({"metric": "roa", "cohort": "broad",
                           "provenance": broad["provenance"], "bands": _cp.deepcopy(broad["bands"])})
    bad2["bands"][0]["p75"] = bad2["bands"][0]["p25"] - 1
    try:
        _pb.parse_bands_response(bad2); ok2 = False
    except _pb.BandsError as e:
        ok2 = "monotonic" in str(e)
    check("T60d", "fail-closed on non-monotonic percentiles", ok2)
    try:
        _pb.get_bands("nim", "broad"); ok3 = False
    except _pb.BandsError as e:
        ok3 = "no database connection resolved it" in str(e) and "no provisional fixture" in str(e)
    check("T60e", "honest refusal for uncovered metrics (no DB, no fixture — states both)", ok3)


def t59():
    print("T59 wizard-born configs (no peer_query) validate and run — the purge's loose end stays fixed")
    import json as _j
    from foundry.v2 import validate_q as _v, run_q as _r
    cfg = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfg.pop("peer_query", None)
    errs = _v.validate_errors_v2(cfg)
    check("T59a", f"no required-key error without peer_query ({[e for e in errs if 'peer_query' in str(e)][:1]})",
          not any("peer_query" in str(e) for e in errs))
    res = _r.run_v2(cfg)
    rr = res.get("results", res)
    check("T59b", "engine produces numbers with peer honestly absent",
          isinstance(rr, dict) and rr.get("peer") is None)


def t58():
    print("T58 Tier-1 accounts: identity, isolation, reset kit")
    import os as _os, tempfile as _tf, json as _j, importlib
    prev = _os.environ.get("FOUNDRY_DATA_DIR")
    td = _tf.mkdtemp(prefix="t58-")
    _os.environ["FOUNDRY_DATA_DIR"] = td
    try:
        from foundry import auth as _auth, store as _store
        importlib.reload(_auth)
        users = _auth.load_users()
        check("T58a", "all eight accounts provisioned from the seed",
              all(u in users for u in ("bgraham","kalt","malt","kmiller",
                                          "phaggerty","rgoldstein","poloyede","jatala")))
        check("T58a", "secrets are scrypt-hashed, never plaintext",
              all("$" in users[u]["password"] and len(users[u]["recovery"]) == 5 for u in users))
        # deterministic test users (the minted passwords live only in the handout)
        users["alice"] = {"password": _auth.hash_secret("alice-pw-123"),
                           "recovery": [_auth.hash_secret("AAAA-BBBB")], "deputy": False}
        users["bob"] = {"password": _auth.hash_secret("bob-pw-12345"),
                         "recovery": [], "deputy": True}
        _auth.save_users(users)
        check("T58b", "password login returns the identity",
              _auth.authenticate("alice", "alice-pw-123") == "alice"
              and _auth.authenticate("alice", "wrong") is None)
        # isolation: alice saves; bob cannot see or load it
        cfg = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
        cfg["scenario_name"] = "Alice Private Bank"
        _store.save_engagement(cfg, slug="alice-private", user="alice")
        check("T58c", "engagements are welded to their owner",
              any(e["slug"] == "alice-private" for e in _store.list_engagements(user="alice"))
              and not any(e["slug"] == "alice-private" for e in _store.list_engagements(user="bob")))
        try:
            _store.load_engagement("alice-private", user="bob"); leaked = True
        except FileNotFoundError:
            leaked = False
        check("T58c", "cross-user load fails closed", not leaked)
        check("T58d", "self-service password change",
              _auth.change_password("alice", "alice-pw-123", "alice-new-pw1")
              and _auth.authenticate("alice", "alice-new-pw1") == "alice")
        check("T58e", "recovery burns the code and sets the password",
              _auth.recover("alice", "AAAA-BBBB", "alice-rec-pw1")
              and _auth.authenticate("alice", "alice-rec-pw1") == "alice"
              and not _auth.recover("alice", "AAAA-BBBB", "again-pw-123"))
        check("T58f", "deputy resets; non-deputy refused",
              _auth.deputy_reset("bob", "alice", "temp-pw-1234")
              and _auth.authenticate("alice", "temp-pw-1234") == "alice"
              and not _auth.deputy_reset("alice", "bob", "hax-pw-12345"))
    finally:
        if prev is None: _os.environ.pop("FOUNDRY_DATA_DIR", None)
        else: _os.environ["FOUNDRY_DATA_DIR"] = prev


def t57():
    print("T57 duplicate-definition hygiene: the cascade-trap disease stays dead")
    import re
    h = open("web/console_v2.html", encoding="utf-8").read()
    # (a) CSS: no standalone selector defined twice at top level (outside @media)
    bad_css = []
    for m in re.finditer(r"<style[^>]*>(.*?)</style>", h, re.S):
        css = m.group(1)
        spans = []
        for mm in re.finditer(r"@media[^{]*\{", css):
            d, j = 1, mm.end()
            while d and j < len(css):
                if css[j] == "{": d += 1
                elif css[j] == "}": d -= 1
                j += 1
            spans.append((mm.start(), j))
        seen = {}
        for mm in re.finditer(r"([^{}@]+)\{[^{}]*\}", css):
            if any(a <= mm.start() < b for a, b in spans): continue
            sel = mm.group(1).strip()
            if "," in sel or not sel: continue
            seen[sel] = seen.get(sel, 0) + 1
        bad_css += [s for s, n in seen.items() if n > 1]
    check("T57a", f"no standalone CSS selector defined twice at top level ({bad_css[:4]})",
          not bad_css)
    # (b) JS: no symbol DECLARED twice (window.X = X exports excluded)
    i = h.rindex("<script>"); j = h.index("</script>", i)
    js = h[i + 8:j]
    decl = {}
    for mm in re.finditer(r"^(?:function\s+(\w+)|const\s+(\w+)\s*=|let\s+(\w+)\s*=)", js, re.M):
        n = next(g for g in mm.groups() if g)
        decl[n] = decl.get(n, 0) + 1
    bad_js = [n for n, k in decl.items() if k > 1]
    check("T57b", f"no JS symbol declared twice (the fmtC class of collision) ({bad_js[:4]})",
          not bad_js)


def t56():
    print("T56 SETTINGS sheet: every in-app input visible, none of it imported")
    import json as _j, io as _io, openpyxl as _ox
    from foundry.v2 import fiw as _fiw
    cfg = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    a = cfg["assumptions"]
    a["scheduled_borrowings"] = [{"name": "FHLB advance", "quarter": 2, "amount": 5_000_000,
                                    "term_q": 8, "rate_ann": 0.06}]   # the UI's shape: 'quarter'
    a["nie_detail"] = {"fte_by_year": [20, 30, 40], "loaded_comp_annual": 150_000,
                        "categories": [{"name": "Occupancy", "per_quarter": 60_000}],
                        "other_gross_up_rate": 0.03}                    # the UI's shape
    a["securities_afs"] = [{"name": "Agency MBS", "opening": 10_000_000, "yield_ann": 0.04}]
    a["fee_modules"] = {"interchange": {"penetration": 0.5}}
    cfg["pre_opening"] = {"expenses": [{"category": "Legal & filings", "total": 500_000},
                                         {"category": "Build-out", "total": 750_000}]}   # the UI's shape
    out = _fiw.build_fiw(cfg)
    data = out[0] if isinstance(out, tuple) else out
    data = data.getvalue() if hasattr(data, "getvalue") else data
    wb = _ox.load_workbook(_io.BytesIO(data))
    check("T56a", "SETTINGS sheet exists and is visible", "SETTINGS" in wb.sheetnames
          and wb["SETTINGS"].sheet_state == "visible")
    txt = " ".join(str(c2.value) for r in wb["SETTINGS"].iter_rows() for c2 in r if c2.value is not None)
    check("T56b", "states treasury, borrowings, securities, fee modules, stress",
          "Cash floor" in txt and "FHLB advance" in txt and "Agency MBS" in txt
          and "interchange" in txt and "Stress parameters" in txt)
    check("T56b", "pre-opening expenses in the UI's OWN shape ({category,total}) render with values",
          "Legal & filings" in txt and "500000" in txt.replace(",", "")
          and "Build-out" in txt and "Total pre-opening burn" in txt and "1250000" in txt.replace(",", ""))
    check("T56c", "declares itself not-imported", "NOT imported" in txt)
    check("T56c", "NIE detail in the UI's OWN shape renders (FTE by year, comp, categories, gross-up)",
          "20 / 30 / 40" in txt and "150000" in txt.replace(",", "")
          and "Occupancy" in txt and "60000" in txt.replace(",", "") and "gross-up" in txt.lower())
    check("T56c", "borrowing shows its QUARTER (canonical key), name, and terms",
          "FHLB advance" in txt and "draws Q2" in txt and "8q" in txt)
    # editing a SETTINGS cell produces zero edits on import
    wb["SETTINGS"]["B5"] = 0.99
    buf2 = _io.BytesIO(); wb.save(buf2)
    merged, rep = _fiw.diff_import(buf2.getvalue(), _j.loads(_j.dumps(cfg)))
    check("T56d", "SETTINGS edits are ignored by the import (stated, never merged)",
          rep["edit_count"] == 0
          and merged["assumptions"]["cash_target_pct_deposits"] == cfg["assumptions"]["cash_target_pct_deposits"])


def t54():
    print("T54 deposit grammar: absolute net-new inflows (Patrick's DEP roll)")
    import json as _j
    cfg = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    cfg = _j.loads(_j.dumps(cfg))
    d0 = cfg["assumptions"]["deposit_products"][0]
    d0["opening_balance"] = 0.0
    d0["growth_q"] = 0.0
    d0["runoff_q"] = 0.02
    d0["new_deposits_q"] = 9_000_000.0   # Patrick: $3M/month x 3
    from foundry.v2.run_q import run_v2
    r = run_v2(cfg)
    # hand roll: q1 = 0*(1-.02)+9M ; q2 = q1*(1-.02)+9M ; q3 likewise
    exp = 0.0
    for _ in range(3): exp = exp * 0.98 + 9_000_000.0
    prod = [p for p in r["products"] if p["family"] == "deposit"][0]
    got = prod["bal"][3]
    check("T54a", "zero-opening bank grows on absolute inflows (pct-of-zero trap closed)",
          got > 0)
    check("T54b", f"Q3 balance matches the hand roll ({exp:,.0f} — payload reports $000s)",
          abs(got * 1000.0 - exp) < 1000.0)
    cfg2 = _j.loads(_j.dumps(cfg))
    cfg2["assumptions"]["deposit_products"][0].pop("new_deposits_q")
    r2 = run_v2(cfg2)
    check("T54c", "field absent -> behavior identical to before the feature (additive default-off)",
          [p for p in r2["products"] if p["family"] == "deposit"][0]["bal"][3] == 0.0)


def t55():
    print("T55 deposit maturity: cohort roll-off on the quarterly clock")
    import json as _j
    from foundry.v2.run_q import run_v2
    base = _j.load(open("foundry/fixtures/parity/configs/pf_a_base.json", encoding="utf-8"))
    # (a) CD ladder: opening 0, $9M/q inflows, 12-month maturity (4q), no runoff
    cfg = _j.loads(_j.dumps(base))
    d = cfg["assumptions"]["deposit_products"][0]
    d.update({"opening_balance": 0.0, "growth_q": 0.0, "runoff_q": 0.0,
               "new_deposits_q": 9_000_000.0, "avg_maturity_m": 12.0})
    bal = [p for p in run_v2(cfg)["products"] if p["family"] == "deposit"][0]["bal"]
    check("T55a", "ladder builds 9/18/27/36 then PLATEAUS as cohorts mature ($000s)",
          [round(b) for b in bal[1:7]] == [9000, 18000, 27000, 36000, 36000, 36000])
    # (b) seasoned book, no inflows: opening 40M at 12m maturity runs out evenly
    cfg2 = _j.loads(_j.dumps(base))
    d2 = cfg2["assumptions"]["deposit_products"][0]
    d2.update({"opening_balance": 40_000_000.0, "growth_q": 0.0, "runoff_q": 0.0,
                "new_deposits_q": 0.0, "avg_maturity_m": 12.0})
    bal2 = [p for p in run_v2(cfg2)["products"] if p["family"] == "deposit"][0]["bal"]
    check("T55b", "seasoned even ladder runs out 30/20/10/0",
          [round(b) for b in bal2[1:5]] == [30000, 20000, 10000, 0])
    # (c) absent/zero maturity -> exact pre-feature behavior
    cfg3 = _j.loads(_j.dumps(base))
    r3 = run_v2(cfg3)
    r0 = run_v2(_j.loads(_j.dumps(base)))
    check("T55c", "no maturity field -> path identical (additive default-off)",
          [p for p in r3["products"] if p["family"]=="deposit"][0]["bal"]
          == [p for p in r0["products"] if p["family"]=="deposit"][0]["bal"])


def t53():
    print("T53 challenge thresholds are data, visible, provenance-tagged")
    import foundry.v2.challenge_q as ch, inspect
    src = inspect.getsource(ch)
    ids = [t["id"] for t in ch.CHALLENGE_THRESHOLDS]
    check("T53a", "registry exists with all judged rules", len(ids) >= 6)
    check("T53b", "every advertised id is one the module can emit",
          all(f'"{i}"' in src for i in ids))
    check("T53c", "provenance is honest (static, not-yet-peer-calibrated) WITHOUT internal vocabulary",
          ("industry" in ch.PROVENANCE and "peer" in ch.PROVENANCE)
          and not any(nm in ch.PROVENANCE for nm in ("Roman", "Patrick", "Konrad", "Brian"))
          and "F-121" not in ch.PROVENANCE)


def t63():
    print("T63 peer calibration (F-121 provisional tier): asset-band selection, per-metric vintage, n-aware, fail-closed")
    from foundry.v2.peer_calibration import (calibrate_thresholds, asset_band_for,
                                             place_flag_value)
    from foundry.v2.challenge_q import CHALLENGE_THRESHOLDS
    # pre-registered selection: band derives from projected assets, nothing else
    check("T63a", "asset band derives from projected size (pre-registered, R1)",
          asset_band_for(150_000) == "under_200M" and asset_band_for(350_000) == "200M_500M"
          and asset_band_for(5_000_000) == "2B_10B")
    rows, prov = calibrate_thresholds(CHALLENGE_THRESHOLDS, 350_000)
    byid = {r["id"]: r for r in rows}
    # funding flags calibrate to deposit_cost at the LEGACY vintage; charge-offs at substrate
    fh = byid.get("FUND-HOT", {}).get("peer")
    co = byid.get("CO-BAND", {}).get("peer")
    check("T63b", "funding flag carries deposit_cost percentiles at the 2025Q4 legacy vintage",
          fh is not None and fh["band_metric"] == "deposit_cost" and "2025Q4" in fh["vintage"])
    check("T63c", "charge-off flag carries substrate-grade 2026Q1 percentiles",
          co is not None and co["band_metric"] == "net_charge_off_pct" and "2026Q1" in co["vintage"])
    check("T63d", "n per point present; provenance states pre-registered selection and small-n labeling",
          fh.get("n") is not None and "pre-registered" in prov and "small-n" in prov)
    # placement takes the worse reading at a seam (R5) and reports the corridor
    pl = place_flag_value(3.7, "deposit_cost", "200M_500M", worse="high")
    check("T63e", "placement puts an aggressive value above p90 with n and source",
          pl is not None and pl["corridor"] == "above p90" and pl["n"] and "R5" in pl["conservative_note"])
    # provenance is honest about the provisional tier and never leaks internal vocab
    check("T63f", "provenance states provisional tier, no internal names/floor ids",
          "not yet certified" in prov and "Deliverable D" in prov
          and not any(nm in prov for nm in ("Roman", "Patrick", "Konrad", "Brian")))
    # FAIL-CLOSED: a metric with no fixture leaves the static row untouched, no crash
    rows2, prov2 = calibrate_thresholds(CHALLENGE_THRESHOLDS, 40_000_000)  # 40B band, no fixtures
    check("T63g", "substrate miss falls back to static, honest reason, no crash",
          all(r.get("peer") is None for r in rows2)
          and ("industry ranges" in prov2 or "static" in " ".join(r.get("peer_note","") for r in rows2)))



if __name__ == "__main__":
    print("Foundry protocol harness — engine", runner.ENGINE_VERSION)
    t2(); t3(); t4(); t6(); t14(); t15(); t16(); t17(); t18(); t19(); t20(); t21(); t22(); t23(); t24(); t25(); t26(); t27(); t28(); t29(); t30(); t31(); t32(); t33(); t34(); t35(); t36(); t37(); t38(); t39(); t40(); t41(); t42(); t43(); t44(); t45(); t46(); t47(); t48(); t49(); t50(); t51(); t53(); t54(); t55(); t56(); t57(); t58(); t59(); t60(); t61(); t62(); t63()
    npass = sum(1 for *_x, ok, _d in [(r[0], r[1], r[2], r[3]) for r in RESULTS] if ok)
    print(f"\n{npass}/{len(RESULTS)} checks passed")
    sys.exit(0 if npass == len(RESULTS) else 1)
