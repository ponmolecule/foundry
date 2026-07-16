"""Mode T, stage T-1: ingest + recon (TRANSLATION_PIPELINE.md stages 0-1).

Ingest normalizes any arriving workbook or CSV into a cell/table inventory —
every sheet including hidden ones (the Patrick lesson: his PEER tab was
hidden), dimensions, populated density, and detected time axes. Recon turns
the inventory into the human-readable "what does this file contain" report:
candidate series matched by a label lexicon, cadence detected from axis
patterns, magnitudes summarized — and units always flagged UNVERIFIED, never
guessed from magnitude (the CET1 lesson). No stage here interprets meaning;
that is the mapping session's job (T-2), with a human confirming every
assignment.
"""
import csv
import io
import json
import hashlib
import re

LEXICON = {
    "deposits": ["deposit", "dda", "mmda", "savings", "checking", "cd ", "cds",
                  "brokered", "money market"],
    "loans": ["loan", "lending", "origination", "mortgage", "credit card",
               "c&i", "cre", "consumer", "installment", "receivable"],
    "pricing": ["yield", "rate", "apr", "coupon", "cost of funds", "spread",
                 "fed funds", "prime", "sofr", "interest"],
    "credit": ["charge-off", "charge off", "nco", "provision", "allowance",
                "alll", "reserve", "delinquen", "loss"],
    "capital": ["capital", "equity", "tier 1", "cet1", "leverage", "raise"],
    "treasury": ["securities", "afs", "htm", "cash", "fed funds sold",
                  "borrowing", "fhlb", "liquidity"],
    "fees": ["fee", "interchange", "service charge", "trust", "baas",
              "gain on sale", "servicing"],
    "opex": ["expense", "salaries", "compensation", "fte", "headcount",
              "occupancy", "marketing", "vendor", "assessment"],
    "volume": ["accounts", "customers", "members", "units", "transactions",
                "balance"],
}

_MONTHS = re.compile(r"^(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)", re.I)
_Q = re.compile(r"^q[1-9]\d*$", re.I)
_YR = re.compile(r"^(19|20)\d{2}$")
_MNUM = re.compile(r"^m(onth)?\s*\d{1,3}$", re.I)
_YLBL = re.compile(r"^y(ear)?\s*\d{1,2}$", re.I)


def _axis_cadence(labels):
    """Detect the cadence of a header run. Returns (cadence, span) or None."""
    labs = [str(x).strip() for x in labels if x not in (None, "")]
    if len(labs) < 3:
        return None
    def frac(pat):
        return sum(1 for x in labs if pat.match(x)) / len(labs)
    if frac(_MONTHS) > 0.6 or frac(_MNUM) > 0.6:
        return ("monthly", len(labs))
    if frac(_Q) > 0.6:
        return ("quarterly", len(labs))
    if frac(_YR) > 0.6 or frac(_YLBL) > 0.6:
        return ("annual", len(labs))
    # consecutive integer runs (1..N): span is evidence, cadence is NOT —
    # twelve could be months or quarters; the mapping session decides.
    try:
        nums = [int(float(x)) for x in labs]
        if all(b - a == 1 for a, b in zip(nums, nums[1:])):
            return ("periodic", len(nums))
    except (ValueError, TypeError):
        pass
    return None


def ingest_xlsx(data):
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)
    inv = {"kind": "xlsx", "sheets": []}
    for ws in wb.worksheets:
        rows = list(ws.iter_rows(values_only=True))
        populated = sum(1 for row in rows for v in row if v is not None)
        axes = []
        for ri, row in enumerate(rows[:30]):
            cad = _axis_cadence(list(row)[1:])
            if cad:
                axes.append({"row": ri + 1, "cadence": cad[0], "span": cad[1]})
        inv["sheets"].append({
            "name": ws.title,
            "state": ws.sheet_state,          # visible | hidden | veryHidden
            "dims": [ws.max_row, ws.max_column],
            "populated_cells": populated,
            "time_axes": axes[:4],
            "_rows": rows,                     # carried for recon; dropped in report
        })
    return inv


def ingest_csv(data):
    text = data.decode("utf-8", errors="replace")
    rows = list(csv.reader(io.StringIO(text)))
    axes = []
    for ri, row in enumerate(rows[:10]):
        cad = _axis_cadence(row[1:])
        if cad:
            axes.append({"row": ri + 1, "cadence": cad[0], "span": cad[1]})
    return {"kind": "csv", "sheets": [{
        "name": "(csv)", "state": "visible",
        "dims": [len(rows), max((len(r) for r in rows), default=0)],
        "populated_cells": sum(1 for r in rows for v in r if v not in ("", None)),
        "time_axes": axes[:4], "_rows": rows,
    }]}


def ingest(data, filename=""):
    if filename.lower().endswith(".csv") or (data[:4] != b"PK\x03\x04" and b"," in data[:200]):
        return ingest_csv(data)
    return ingest_xlsx(data)


def _classify(label):
    low = str(label).lower()
    return [theme for theme, words in LEXICON.items() if any(w in low for w in words)]


def recon(inv):
    """Inventory -> report. Candidates matched by lexicon; units UNVERIFIED."""
    report = {"kind": inv["kind"], "sheets": [], "candidates": [],
              "hidden_sheets": [], "notes": [
                  "Units are UNVERIFIED throughout — declared at mapping, never "
                  "inferred from magnitude.",
                  "Candidate matching is lexical; the mapping session (human-"
                  "confirmed) decides meaning. Nothing below is an assignment."]}
    for sh in inv["sheets"]:
        rows = sh["_rows"]
        report["sheets"].append({k: v for k, v in sh.items() if k != "_rows"})
        if sh["state"] != "visible":
            report["hidden_sheets"].append(sh["name"])
        sheet_cadence = sh["time_axes"][0]["cadence"] if sh["time_axes"] else None
        for ri, row in enumerate(rows):
            label = row[0] if row else None
            if label is None or not str(label).strip():
                continue
            themes = _classify(label)
            if not themes:
                continue
            vals = [v for v in list(row)[1:] if isinstance(v, (int, float))]
            report["candidates"].append({
                "sheet": sh["name"], "row": ri + 1,
                "label": str(label).strip()[:80], "themes": themes,
                "numeric_cells": len(vals),
                "series_like": len(vals) >= 3,
                "cadence_guess": sheet_cadence if len(vals) >= 3 else None,
                "magnitude": [min(vals), max(vals)] if vals else None,
                "units": "UNVERIFIED",
            })
    report["candidate_count"] = len(report["candidates"])
    report["report_hash"] = hashlib.sha256(
        json.dumps(report, sort_keys=True, default=str).encode()).hexdigest()[:12]
    return report


def recon_file(data, filename=""):
    return recon(ingest(data, filename))
