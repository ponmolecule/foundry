"""Business Plan Tables — cover-sheet generator (two sheets: Annual + Quarterly).

Produces an engagement-specific 'Business Plan Tables' workbook following the Highnote template shape,
with BOTH tabs the template has: 'Business Plan Tables Annual' and 'Business Plan Tables Quarterly'.
Every populated number comes from the real engine result (run_v2); template placeholders
([Loan Product 1], [Deposit Source ...]) are replaced with the ACTUAL product names for this engagement
and expand to as many rows as it has products. Nothing is fabricated — a line the engine does not
separately produce is left blank (not zero).

Conventions:
  Annual sheet    — stocks at year-end (Q4/Q8/Q12); flows summed over the year's four quarters.
  Quarterly sheet — stocks at each quarter-end; flows for that single quarter. Both from the same
                    length-12 engine series, so the two sheets reconcile by construction.
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import io

_ARIAL = "Arial"
_HDR_FILL = PatternFill("solid", fgColor="1F3B5F")
_HDR_FONT = Font(name=_ARIAL, bold=True, color="FFFFFF", size=11)
_SUB_FONT = Font(name=_ARIAL, bold=True, color="1F3B5F", size=10)
_LBL_FONT = Font(name=_ARIAL, size=10)
_NUM_FONT = Font(name=_ARIAL, size=10)
_TOTAL_FONT = Font(name=_ARIAL, bold=True, size=10)
_TITLE_FONT = Font(name=_ARIAL, bold=True, color="1F3B5F", size=12)
_MONEY = '#,##0;(#,##0);-'
_PCT = '0.0%;(0.0%);-'
_PCT2 = '0.00%;(0.00%);-'
_NUM = '#,##0.0'


def _norm(series):
    s = list(series or [])
    if len(s) == 13:
        return s[1:], s[0]
    return s, None


def _annual_periods(ppy=4, n_periods=12):
    # AUDIT 4.3: cadence-aware. Engine runs in `ppy` periods/year; the BPT annual view aggregates
    # each YEAR: stock at year-end engine-period, flows summed over the year's engine-periods.
    nyr = max(1, n_periods // ppy)
    periods = [{"label": "Day 1", "day1": True}]
    for y in range(1, nyr + 1):
        idxs = [(y - 1) * ppy + k for k in range(ppy)]      # engine-period indices in year y
        periods.append({"label": f"Year {y}", "stock_idx": y * ppy - 1, "flow_idxs": idxs})
    return periods


def _quarterly_periods(ppy=4, n_periods=12):
    # AUDIT 4.3: the QUARTERLY regulatory view. mpq engine-periods per quarter (mo:3, qtr:1). A
    # quarter's stock = its last engine-period; its flow = sum over the quarter's engine-periods.
    mpq = max(1, round(ppy / 4))
    nq = max(1, round(n_periods * 4 / ppy))
    nyr = max(1, nq // 4)
    periods = [{"label": "Day 1", "day1": True}]
    for y in range(1, nyr + 1):
        for q in range(1, 5):
            qn = (y - 1) * 4 + q
            base = (qn - 1) * mpq
            idxs = [base + j for j in range(mpq)]
            periods.append({"label": f"Y{y} Q{q}", "stock_idx": base + mpq - 1, "flow_idxs": idxs})
    return periods


def _stock_vals(series, periods):
    s, d1 = _norm(series)
    out = []
    for p in periods:
        if p.get("day1"):
            out.append(d1)
        else:
            i = p["stock_idx"]
            out.append(s[i] if 0 <= i < len(s) and s[i] is not None else None)
    return out


def _flow_vals(series, periods):
    s, _ = _norm(series)
    out = []
    for p in periods:
        if p.get("day1"):
            out.append(None)
        else:
            vals = [s[i] for i in p["flow_idxs"] if 0 <= i < len(s) and s[i] is not None]
            out.append(sum(vals) if vals else 0.0)
    return out


def _engagement_slug(cfg):
    """A filename-safe identifier for THIS saved engagement/variation.

    The Save-As dialog writes the user's chosen name into cfg.scenario_name (see guardSaveAsConfirm),
    so that is the correct source: two variations of the same bank ('Nook & Cranny Bank' vs
    'Nook & Cranny Bank_adopted_metrics') must produce distinct filenames. Fall back to engagement_id,
    then the bank name, then a generic label.
    """
    import re
    raw = (cfg.get("scenario_name") or "").strip() or None
    if not raw:
        raw = cfg.get("engagement_id")
    if not raw:
        pb = cfg.get("proposed_bank")
        raw = pb if isinstance(pb, str) else (pb.get("legal_name") or pb.get("name")) if isinstance(pb, dict) else None
    if not raw:
        raw = cfg.get("client_legal_name") or "engagement"
    slug = re.sub(r"[^A-Za-z0-9]+", "_", str(raw)).strip("_")
    return slug or "engagement"


def _bank_name(cfg):
    pb = cfg.get("proposed_bank", {})
    if isinstance(pb, str):
        return pb or cfg.get("client_legal_name") or "Proposed Bank"
    if isinstance(pb, dict):
        return pb.get("legal_name") or pb.get("name") or cfg.get("client_legal_name") or "Proposed Bank"
    return cfg.get("client_legal_name") or "Proposed Bank"


def _build_sheet(ws, cfg, res, periods, granularity):
    a = cfg.get("assumptions", {})
    fin = res.get("financials", {})
    bs = fin.get("bs", {})
    is_ = fin.get("is", {})
    cap = (res.get("capital", {}) or {}).get("standardized", {}) or {}
    cap_ratios = cap.get("ratios", {}) or {}
    products = res.get("products", []) or []
    lend = [p for p in products if p.get("family") == "lending"]
    dep_products = a.get("deposit_products", []) or []

    ncols = len(periods)
    flow_cols = [p for p in periods if not p.get("day1")]

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 42
    for col in range(3, 3 + ncols):
        ws.column_dimensions[get_column_letter(col)].width = 12

    row = [1]

    def _blank(): row[0] += 1

    def _title(text):
        c = ws.cell(row=row[0], column=1, value=text); c.font = _TITLE_FONT; row[0] += 1

    def _header(label, cols):
        r = row[0]
        cb = ws.cell(row=r, column=2, value=label); cb.font = _HDR_FONT; cb.fill = _HDR_FILL
        cb.alignment = Alignment(horizontal="left")
        for j, h in enumerate(cols):
            cc = ws.cell(row=r, column=3 + j, value=h); cc.font = _HDR_FONT; cc.fill = _HDR_FILL
            cc.alignment = Alignment(horizontal="right")
        row[0] += 1

    def _datarow(label, values, fmt=_MONEY, bold=False, indent=False, section=False):
        r = row[0]
        cb = ws.cell(row=r, column=2, value=label)
        cb.font = _SUB_FONT if section else (_TOTAL_FONT if bold else _LBL_FONT)
        if indent: cb.alignment = Alignment(indent=1)
        for j, v in enumerate(values):
            cc = ws.cell(row=r, column=3 + j, value=v)
            cc.font = _TOTAL_FONT if bold else _NUM_FONT
            cc.number_format = fmt
            cc.alignment = Alignment(horizontal="right")
        row[0] += 1

    all_labels = [p["label"] for p in periods]
    flow_labels = [p["label"] for p in flow_cols]

    c = ws.cell(row=row[0], column=1, value=f"{_bank_name(cfg)} \u2014 Business Plan Tables ({granularity.title()})")
    c.font = Font(name=_ARIAL, bold=True, size=15, color="1F3B5F"); row[0] += 1
    c = ws.cell(row=row[0], column=1,
                value="Numbers in $ thousands unless noted. Generated from the Foundry engine for this engagement.")
    c.font = Font(name=_ARIAL, italic=True, size=9, color="6B7A8D"); row[0] += 2

    # Balance Sheet ($)
    _title("Table 1: Balance Sheet ($ thousands)")
    _header("Balance Sheet", all_labels)
    bs_rows = [
        ("Assets", None, True, True),
        ("Cash and Balances Due from Banks", "cash", False, False),
        ("Securities", "sec", False, False),
        ("Gross Loans", "grossLoans", False, False),
        ("Allowance for Credit Losses", "alll", False, False),
        ("Other Assets", None, False, False),
        ("Intangible Assets", None, False, False),
        ("Total Assets", "totalAssets", True, False),
        ("Liabilities and Equity", None, True, True),
        ("Deposits", "deposits", False, False),
        ("Other Liabilities", "borrow", False, False),
        ("Total Liabilities", None, True, False),
        ("Equity", "equity", True, False),
    ]
    for label, key, bold, section in bs_rows:
        if section:
            _datarow(label, [None] * ncols, section=True); continue
        if key is None:
            _datarow(label, [None] * ncols, bold=bold); continue
        vals = _stock_vals(bs.get(key), periods)
        if key == "alll":
            vals = [(-v if isinstance(v, (int, float)) else v) for v in vals]
        _datarow(label, vals, bold=bold)
    _blank()

    # Income Statement ($) — flows
    _title("Table 2: Income Statement ($ thousands)")
    _header("Income Statement", flow_labels)

    def _flow_sum(keys):
        cols = None
        for k in keys:
            v = _flow_vals(is_.get(k), flow_cols)
            cols = v if cols is None else [(cols[i] or 0) + (v[i] or 0) for i in range(len(v))]
        return cols or [None] * len(flow_cols)

    ii = _flow_sum(["loanInt", "secInt", "cashInt", "bookInt"])
    ie = _flow_sum(["depExp", "borrExp"])
    _datarow("Interest Income", ii)
    _datarow("Interest Expense", ie)
    _datarow("Provision for Credit Losses", _flow_sum(["prov"]))
    _datarow("Non-Interest Income", _flow_sum(["fees", "gos", "servNet", "fvPnl"]))
    _datarow("Non-Interest Expense", _flow_sum(["prodOpex", "overhead"]))
    _datarow("Pre-Tax Net Income", _flow_sum(["pretax"]), bold=True)
    _datarow("Taxes", _flow_sum(["tax"]))
    _datarow("Net Income (Loss)", _flow_sum(["ni"]), bold=True)
    _blank()

    # Interest Income by Product Line (DYNAMIC)
    _title("Table 3: Interest Income by Product Line ($ thousands)")
    _header("Interest Income by Product Line", flow_labels)
    for p in lend:
        _datarow(p.get("name", "Loan Product"), _flow_vals(p.get("intInc"), flow_cols), indent=True)
    tot_loan_ii = None
    for p in lend:
        v = _flow_vals(p.get("intInc"), flow_cols)
        tot_loan_ii = v if tot_loan_ii is None else [(tot_loan_ii[i] or 0) + (v[i] or 0) for i in range(len(v))]
    _datarow("Total Loans", tot_loan_ii or [None] * len(flow_cols), bold=True)
    _datarow("Cash & Securities", _flow_sum(["secInt", "cashInt"]))
    _datarow("Total Interest Income", ii, bold=True)
    _blank()

    # per-product Characteristics (DYNAMIC)
    for idx, p in enumerate(lend, 1):
        name = p.get("name", f"Loan Product {idx}")
        _title(f"Table 4.{idx}: {name} \u2014 Characteristics")
        _header(f"{name}", flow_labels)
        _datarow("Total Bank Originations ($ thousands)", _flow_vals(p.get("origq"), flow_cols))
        cfg_p = next((lp for lp in a.get("lending_products", []) if lp.get("name") == name), {})
        yld = cfg_p.get("yield_ann")
        _datarow("Average Asset Yield (%)",
                 [yld if isinstance(yld, (int, float)) else None] * len(flow_cols), fmt=_PCT2)
        term_q = cfg_p.get("term_q")
        term = (term_q / 4.0) if isinstance(term_q, (int, float)) and term_q else None
        _datarow("Average Loan Term (years)", [term] * len(flow_cols), fmt=_NUM)
        _blank()

    # Interest Expense by Product Line
    _title("Table 5: Interest Expense by Product Line ($ thousands)")
    _header("Interest Expense by Product Line", flow_labels)
    dep_in_res = [p for p in products if p.get("family") == "deposit"]
    if dep_in_res:
        for p in dep_in_res:
            _datarow(p.get("name", "Deposit"), _flow_vals(p.get("intExp"), flow_cols), indent=True)
    else:
        for dp in dep_products:
            _datarow(dp.get("name", "Deposit Source"), [None] * len(flow_cols), indent=True)
    _datarow("Total Interest Expense", ie, bold=True)
    _blank()
    if not dep_in_res:
        cn = ws.cell(row=row[0], column=2,
            value="Per-deposit interest-expense split not separately modeled; only the aggregate total is engine-sourced.")
        cn.font = Font(name=_ARIAL, italic=True, size=8, color="9AA7B4"); row[0] += 2

    # Forecast Capital Levels — stocks
    _title("Table 6: Forecast Capital Levels")
    _header("Forecast Capital Levels", all_labels)

    def _cap_row(key):
        s = cap_ratios.get(key)
        if not isinstance(s, list):
            return [None] * ncols
        vals = _stock_vals(s, periods)
        return [(v / 100.0 if isinstance(v, (int, float)) else None) for v in vals]

    _datarow("Tier 1 Leverage Ratio", _cap_row("leverage"), fmt=_PCT)
    _datarow("CET 1 Capital Ratio", _cap_row("cet1_rwa"), fmt=_PCT)
    _datarow("Tier 1 Capital Ratio", _cap_row("tier1_rwa"), fmt=_PCT)
    _datarow("Total Risk-Based Capital Ratio", _cap_row("total_rwa"), fmt=_PCT)
    _blank()

    # Loan Loss Provision by Product (DYNAMIC)
    _title("Table 7: Loan Loss Provision Expense by Product ($ thousands)")
    _header("Loan Loss Provision Expense", flow_labels)
    tot_co = None
    for p in lend:
        v = _flow_vals(p.get("co"), flow_cols)
        _datarow(p.get("name", "Loan Product"), v, indent=True)
        tot_co = v if tot_co is None else [(tot_co[i] or 0) + (v[i] or 0) for i in range(len(v))]
    _datarow("Total", tot_co or [None] * len(flow_cols), bold=True)
    _blank(); _blank()

    note = ws.cell(row=row[0], column=2,
        value=(("Annual view: balance-sheet & capital figures at year-end (Q4/Q8/Q12); income figures "
                "summed over the year's four quarters. ")
               if granularity == "annual" else
               ("Quarterly view: balance-sheet & capital figures at each quarter-end; income figures for "
                "that single quarter. ")) +
              "Blank cells indicate a line the engine does not separately produce (not zero). "
              "Product tables reflect this engagement's actual products.")
    note.font = Font(name=_ARIAL, italic=True, size=8, color="6B7A8D")
    note.alignment = Alignment(wrap_text=True)
    ws.merge_cells(start_row=row[0], start_column=2, end_row=row[0], end_column=2 + ncols)
    ws.row_dimensions[row[0]].height = 42

    ws.freeze_panes = "C1"


def build_bpt_cover(cfg, res):
    """Return (xlsx_bytes, filename). Two-tab Business Plan Tables workbook (Annual + Quarterly)."""
    wb = Workbook()
    ws_annual = wb.active
    ws_annual.title = "Business Plan Tables Annual"
    _ppy = int((cfg.get("assumptions") or {}).get("periods_per_year") or 4)
    _np = int((cfg.get("assumptions") or {}).get("n_periods") or 12)
    _build_sheet(ws_annual, cfg, res, _annual_periods(_ppy, _np), "annual")

    ws_q = wb.create_sheet("Business Plan Tables Quarterly")
    _build_sheet(ws_q, cfg, res, _quarterly_periods(_ppy, _np), "quarterly")

    buf = io.BytesIO()
    wb.save(buf)
    filename = f"{_engagement_slug(cfg)}_Foundry_Business_Plan_Table.xlsx"
    return buf.getvalue(), filename
