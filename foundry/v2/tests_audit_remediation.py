"""Independent remediation gate for audit e7220e6 (2026-09-04).

This suite is deliberately orthogonal to the historical golden/parity harness. It tests
calendar semantics, regulatory identities, reporting aggregation, UI/engine contracts,
and security invariants that can remain wrong while old fixtures stay green.

Run: python -m foundry.v2.tests_audit_remediation
"""
from __future__ import annotations
import copy, io, json, os, re, subprocess, tempfile
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
PF_A = ROOT / "foundry/fixtures/parity/configs/pf_a_base.json"
PF_B = ROOT / "foundry/fixtures/parity/configs/pf_b_base.json"


def _load(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _monthly_a():
    c = _load(PF_A)
    c["assumptions"]["periods_per_year"] = 12
    c["assumptions"]["n_periods"] = 36
    return c


def _row(schedule, item):
    return next(r for r in schedule["rows"] if r.get("item") == item)


def _browser_calendar_probe(html: str, ppy: int, opening: str):
    """Execute the ACTUAL browser calendar/rate-curve helpers under Node.

    This is intentionally behavioral rather than a source-string check: Test 6 is
    specifically about the browser path preserving a literal monthly opening month.
    """
    lo = html.index("function _parseISO")
    hi = html.index("function _interpDated", lo)
    funcs = html[lo:hi]
    script = f"""
let cfg={{charter_profile:{{target_opening:{json.dumps(opening)}}}}};
function PPY(){{ return {int(ppy)}; }}
function MO_PER_Q(){{ return Math.max(1, Math.round(PPY()/4)); }}
{funcs}
const out={{
  p1:_modelPeriodEnd(1).toISOString().slice(0,10),
  block1:_modelQuarterEnd(1).toISOString().slice(0,10)
}};
console.log(JSON.stringify(out));
"""
    got = subprocess.run(["node", "-e", script], cwd=ROOT, capture_output=True,
                         text=True, check=True).stdout.strip().splitlines()[-1]
    return json.loads(got)


def _browser_label_probe(html: str, ppy: int, n_periods: int):
    """Execute canonical browser cadence-label helpers under the requested cadence."""
    lo = html.index("function NP()")
    hi = html.index("// Regulator-facing submission horizon", lo)
    funcs = html[lo:hi]
    script = f"""
let cfg={{assumptions:{{periods_per_year:{int(ppy)},n_periods:{int(n_periods)}}}}};
const METHODOLOGY_HTML='';
{funcs}
console.log(JSON.stringify({{
  period:PLAB('full'), p1:PERLAB(1), horizon:HORIZONLAB(),
  event:EVENTUNIT(), scheduled:EVENTSCHEDLABEL()
}}));
"""
    got = subprocess.run(["node", "-e", script], cwd=ROOT, capture_output=True,
                         text=True, check=True).stdout.strip().splitlines()[-1]
    return json.loads(got)


def _browser_nie_entry_probe(html: str):
    """Execute the actual NIE activation/manual-add handlers under Node."""
    add_src = html[html.index("function _newNieDetail()"):
                   html.index("window.nieCatPaste = function(")]
    on_src = html[html.index("window.nieOn = function()"):
                  html.index("// Static mirror of foundry/v2/fee_catalog.py")]
    script = f"""
global.window={{}};
let cfg={{assumptions:{{}}}};
function renderContent(){{}}
function refresh(){{}}
{add_src}
{on_src}
window.nieOn();
const afterOn=JSON.parse(JSON.stringify(cfg.assumptions.nie_detail.categories));
window.nieCatAdd();
const afterAdd=JSON.parse(JSON.stringify(cfg.assumptions.nie_detail.categories));
console.log(JSON.stringify({{afterOn,afterAdd}}));
"""
    got = subprocess.run(["node", "-e", script], cwd=ROOT, capture_output=True,
                         text=True, check=True).stdout.strip().splitlines()[-1]
    return json.loads(got)


def _browser_fin_colspan_probe(html: str, n_periods: int, v21: bool = False):
    """Execute the table-span helper with a native-cadence horizon."""
    lo = html.index("function finColspan")
    hi = html.index("function headerRow", lo)
    funcs = html[lo:hi]
    script = f"""
const V21={str(bool(v21)).lower()};
function NP(){{ return {int(n_periods)}; }}
{funcs}
console.log(JSON.stringify({{
  noOpen:finColspan(false,false),
  withOpen:finColspan(true,false),
  withOpenTotal:finColspan(true,true)
}}));
"""
    got = subprocess.run(["node", "-e", script], cwd=ROOT, capture_output=True,
                         text=True, check=True).stdout.strip().splitlines()[-1]
    return json.loads(got)


def _browser_copy_probe(html: str, ppy: int, n_periods: int):
    """Execute the ACTUAL cadence-sensitive financial/help copy under Node."""
    hlo = html.index("function NP()")
    hhi = html.index("// Regulator-facing submission horizon", hlo)
    helpers = html[hlo:hhi]
    clo = html.index("const CALC_TEXT =")
    chi = html.index("let _noteTerms", clo)
    calc = html[clo:chi]
    script = f"""
let cfg={{assumptions:{{periods_per_year:{int(ppy)},n_periods:{int(n_periods)}}}}};
const METHODOLOGY_HTML='';
{helpers}
{calc}
console.log(JSON.stringify({{
  paidin:calcText('paidin'),
  leverage:calcText('leverage'),
  schedBorrow:calcText('schedBorrow'),
  stressTitle:'Tier 1 Leverage Ratio by '+PLAB('full').replace(/^./,c=>c.toUpperCase())
}}));
"""
    got = subprocess.run(["node", "-e", script], cwd=ROOT, capture_output=True,
                         text=True, check=True).stdout.strip().splitlines()[-1]
    return json.loads(got)


def main():
    from foundry.v2.run_q import run_v2, _cblr_state_machine
    from foundry.v2.parity import run_parity
    from foundry.v2.callreport import build_call_report
    from foundry.v2 import bpt_cover
    from foundry.v2.excel_q import results_workbook_v2
    from foundry.v2.timebase import (model_period_end_date, submission_end_period,
                                     submission_period_label, quarterly_value_to_period)
    from foundry.v2.income_modules import nie_detail_series, managed_notional_series
    from foundry.v2.engine_q_a import dated_rate_fn
    from foundry.v2 import rate_fetch
    from foundry.v2.regparams import REG_PARAMS

    passed = failed = 0
    def ck(name, cond, detail=""):
        nonlocal passed, failed
        if cond:
            passed += 1; print(f"PASS {name}" + (f" — {detail}" if detail else ""))
        else:
            failed += 1; print(f"FAIL {name}" + (f" — {detail}" if detail else ""))

    # A. Calendar/cadence semantics -------------------------------------------------
    c = _monthly_a(); a = c["assumptions"]
    for p in a["deposit_products"]:
        p["growth_q"] = 0; p["runoff_q"] = 0
    for p in a["lending_products"]:
        p["opening_balance"] = 0; p["originations_q"] = 0; p["orig_growth_q"] = 0
        p["runoff_q"] = 0; p["charge_off_ann"] = 0
    dep = a["deposit_products"][0]; dep["opening_balance"] = 24_000_000; dep["avg_maturity_m"] = 24
    r = run_v2(copy.deepcopy(c)); xp = next(x for x in r["products"] if x["name"] == dep["name"])
    ck("A1 24-month CD survives month 8", xp["bal"][8] > 0, f"M8={xp['bal'][8]:.2f}")
    ck("A2 24-month CD matures at month 24", abs(xp["bal"][24]) < .01, f"M24={xp['bal'][24]:.2f}")

    c2 = copy.deepcopy(c); lp = c2["assumptions"]["lending_products"][0]
    lp["opening_balance"] = 24_000_000; lp["structure"] = "term"; lp["term_q"] = 8
    r2 = run_v2(c2); xl = next(x for x in r2["products"] if x["name"] == lp["name"])
    ck("A3 eight-quarter loan survives month 8", xl["bal"][8] > 0, f"M8={xl['bal'][8]:.2f}")
    ck("A4 eight-quarter loan matures at month 24", abs(xl["bal"][24]) < .01, f"M24={xl['bal'][24]:.2f}")

    c3 = copy.deepcopy(c); c3["assumptions"]["capital_raises"] = [{"quarter": 4, "amount": 12_000_000}]
    rr0, rr1 = run_v2(c), run_v2(c3)
    eq0, eq1 = rr0["financials"]["bs"]["equity"], rr1["financials"]["bs"]["equity"]
    ck("A5 Q4 capital raise does not land in month 4", abs(eq1[4]-eq0[4]) < .01)
    ck("A6 Q4 capital raise lands at month 10 (Q4 start)", eq1[10]-eq0[10] > 11_900)

    # legacy quarter rates/flows preserve calendar-quarter economics under monthly cadence
    gq = .12; gm = quarterly_value_to_period("growth", gq, 12)
    ck("A7 quarterly growth converts by compounding", abs((1+gm)**3 - (1+gq)) < 1e-12)
    ck("A8 quarterly dollar flow converts evenly", quarterly_value_to_period("originations", 3_000_000, 12) == 1_000_000)

    # ISO opening month must be preserved, not coerced to quarter start.
    cd = _monthly_a(); cd["charter_profile"] = {"target_opening": "2027-05-15"}
    ck("A9 monthly ISO opening preserves May", str(model_period_end_date(cd, 1, 12)) == "2027-05-31")
    ck("A10 M12 lands April 2028", str(model_period_end_date(cd, 12, 12)) == "2028-04-30")
    html = (ROOT/"web/console_v2.html").read_text(encoding="utf-8")
    bp_m = _browser_calendar_probe(html, 12, "2027-05-15")
    bp_q = _browser_calendar_probe(html, 4, "2027-05-15")
    ck("A10b browser monthly M1 ends May 31", bp_m["p1"] == "2027-05-31", str(bp_m))
    ck("A10c browser monthly first three-month curve block ends July 31",
       bp_m["block1"] == "2027-07-31", str(bp_m))
    ck("A10d browser quarterly opening keeps containing-quarter end",
       bp_q["p1"] == "2027-06-30" and bp_q["block1"] == "2027-06-30", str(bp_q))

    # Explicit cadence-aware event periods allow monthly placement without overloading
    # the backward-compatible `quarter` field. A Month-24 raise must land in Month 24.
    ep0 = _monthly_a(); ep0["assumptions"]["capital_raises"] = []
    ep1 = copy.deepcopy(ep0); ep1["assumptions"]["capital_raises"] = [{"period":24,"amount":12_000_000}]
    er0, er1 = run_v2(ep0), run_v2(ep1)
    ee0, ee1 = er0["financials"]["bs"]["equity"], er1["financials"]["bs"]["equity"]
    ck("A11 explicit Month-24 raise does not land in Month 23", abs(ee1[23]-ee0[23]) < .01)
    ck("A12 explicit Month-24 raise lands in Month 24", ee1[24]-ee0[24] > 11_900)

    # A scheduled FHLB advance placed at M4 with a four-quarter term is a 12-month
    # bullet: M4-M15 outstanding, M16 matured; interest = 12MM*6%/12 = 60k/month.
    fb = _monthly_a(); fb["assumptions"]["capital_raises"] = []
    fb["assumptions"]["scheduled_borrowings"] = [{
        "name":"FHLB test", "period":4, "amount":12_000_000, "rate_ann":.06, "term_q":4
    }]
    fr = run_v2(fb); fbs=fr["financials"]["bs"]; fis=fr["financials"]["is"]
    ck("A13 FHLB M4 draw starts in Month 4", max(abs(x) for x in fbs["borrowSched"][:4]) < .01 and abs(fbs["borrowSched"][4]-12_000)<.01)
    ck("A14 FHLB four-quarter bullet remains through Month 15 and matures Month 16",
       all(abs(fbs["borrowSched"][m]-12_000)<.01 for m in range(4,16)) and abs(fbs["borrowSched"][16])<.01)
    ck("A15 FHLB monthly interest is exactly 60k while outstanding",
       all(abs(fis["borrExp"][m-1]-60)<.01 for m in range(4,16)) and abs(fis["borrExp"][15])<.01)
    ck("A16 wholesale peak includes scheduled FHLB when residual plug is zero",
       abs(fr["scenarios"]["base"]["peak_borrowings"]-12_000)<.01 and max(fbs["borrow"])<.01)

    # B. Regulatory capital --------------------------------------------------------
    base = _load(PF_A); rb = run_v2(copy.deepcopy(base))
    big = copy.deepcopy(base); big["assumptions"]["other_assets"] += 100_000_000
    rg = run_v2(big)
    drwa = rg["capital"]["standardized"]["rwa"][0] - rb["capital"]["standardized"]["rwa"][0]
    ck("B1 general other assets carry positive standardized RWA", drwa > 80_000, f"ΔRWA={drwa:.2f}k")

    dta_cfg = copy.deepcopy(base); dta_cfg["assumptions"]["tax_detail"] = {"enabled": True, "va_mode": "none"}
    rd = run_v2(dta_cfg)
    dta = rd["financials"]["bs"].get("dta") or []
    maxdiff = max(abs(x-y) for x,y in zip(rd["capital"]["standardized"]["cet1"], rd["capital"]["rows"]["tier1"]))
    ck("B2 test actually produces NOL DTA", max(dta) > 0, f"max DTA={max(dta):.2f}k")
    ck("B3 standardized CET1 uses same applicable deductions as Tier 1", maxdiff < .01, f"max diff={maxdiff:.4f}k")
    ck("B4 capital-shortfall reports submission endpoint", rd["capital_shortfall"].get("submission_endpoint") == "Q12")

    # C. NOL roll-forward ----------------------------------------------------------
    ri = rb["financials"]["is"]
    # pf_a has losses through period 11 then small positive P12 with an NOL balance.
    used = ri["nol"][10] - ri["nol"][11]
    expected = .8 * max(0.0, ri["pretax"][11])
    ck("C1 default NOL balance burns only the deduction actually used", abs(used-expected) < .02,
       f"used={used:.2f}, expected={expected:.2f}")

    # D. Regulator-facing Q12 vs computational horizon + Day-1 normalization -------
    qper = bpt_cover._quarterly_periods(12, 36, 12)
    vals = bpt_cover._stock_vals(list(range(37)), qper, 36)
    ck("D1 monthly BPT Q1 stock = M3", vals[1] == 3, f"got={vals[1]}")
    ck("D2 monthly BPT Q12 stock = M36", vals[-1] == 36, f"got={vals[-1]}")
    aper = bpt_cover._annual_periods(12, 36, 12)
    avals = bpt_cover._stock_vals(list(range(37)), aper, 36)
    ck("D3 monthly BPT Year1 stock = M12", avals[1] == 12, f"got={avals[1]}")
    ck("D4 monthly BPT Year3 stock = M36", avals[-1] == 36, f"got={avals[-1]}")
    q20 = bpt_cover._quarterly_periods(4, 20, 12)
    q20vals = bpt_cover._stock_vals(list(range(21)), q20, 20)
    ck("D5 20Q computational model retains regulator-facing Q12 BPT", len(q20)-1 == 12 and q20vals[-1] == 12)
    longc = copy.deepcopy(base); longc["assumptions"]["n_periods"] = 20
    ck("D6 explicit submission endpoint stays Q12 on 20Q model", submission_end_period(longc,4,20)==12 and submission_period_label(longc,4,20)=="Q12")

    # Pro Forma Exhibit core sheets are native cadence; Call Report schedules remain quarterly.
    mc = _monthly_a(); pr = run_parity(copy.deepcopy(mc)); wb = results_workbook_v2(mc, pr)
    ws = wb["Balance Sheet"]
    bs_hdr = [c.value for c in ws[1]]
    ta_row = next(row for row in ws.iter_rows(values_only=True) if row and row[0] == "TOTAL ASSETS")
    ck("D7 monthly Pro Forma Balance Sheet exposes Open + M1-M36",
       bs_hdr[5:] == ["Open"] + [f"M{i}" for i in range(1,37)], str(bs_hdr[5:9])+" ... "+str(bs_hdr[-2:]))
    ck("D8 monthly Pro Forma BS M1/M36 are native period-end stocks",
       abs(ta_row[6] - pr["bs"]["totalAssets"][1]) < .01 and
       abs(ta_row[-1] - pr["bs"]["totalAssets"][36]) < .01)
    isws = wb["Income Statement"]; is_hdr=[c.value for c in isws[1]]
    ni_row = next(row for row in isws.iter_rows(values_only=True) if row and row[0] == "NET INCOME (LOSS)")
    ck("D9 monthly Pro Forma Income Statement exposes M1-M36 native flows",
       is_hdr[5:] == [f"M{i}" for i in range(1,37)] and
       abs(ni_row[5]-pr["is"]["ni"][0]) < .01 and abs(ni_row[-1]-pr["is"]["ni"][35]) < .01)
    rtws=wb["Ratios"]; rt_hdr=[c.value for c in rtws[1]]
    lev_row = next(row for row in rtws.iter_rows(values_only=True) if row and row[1] == "lev")
    ck("D10 monthly Pro Forma Ratios exposes M1-M36 native ratios",
       rt_hdr[2:] == [f"M{i}" for i in range(1,37)] and len(lev_row[2:]) == 36 and
       abs(lev_row[2]-pr["ratios"]["lev"][0]) < .01 and abs(lev_row[-1]-pr["ratios"]["lev"][35]) < .01)
    ri_ws=wb["Schedule RI"]; ri_hdr=[c.value for c in ri_ws[2]]
    ck("D11 monthly Pro Forma Call Report schedules remain Q1-Q12",
       ri_hdr[3:] == [f"Q{i}" for i in range(1,13)], str(ri_hdr[3:]))

    qpr = run_parity(copy.deepcopy(base)); qwb = results_workbook_v2(base, qpr)
    ck("D12 quarterly Pro Forma core sheets remain Q1-Q12",
       [c.value for c in qwb["Balance Sheet"][1]][5:] == ["Open"]+[f"Q{i}" for i in range(1,13)] and
       [c.value for c in qwb["Income Statement"][1]][5:] == [f"Q{i}" for i in range(1,13)] and
       [c.value for c in qwb["Ratios"][1]][2:] == [f"Q{i}" for i in range(1,13)])

    bpt_bytes, _ = bpt_cover.build_bpt_cover(mc, run_v2(copy.deepcopy(mc)))
    bpt_wb = load_workbook(io.BytesIO(bpt_bytes), read_only=True, data_only=True)
    ck("D13 monthly model Business Plan Tables remain Annual + Quarterly only",
       bpt_wb.sheetnames == ["Business Plan Tables Annual", "Business Plan Tables Quarterly"],
       str(bpt_wb.sheetnames))

    # Existing Q12-named scenario / concentration fields are submission metrics, not
    # accidental terminal metrics when computation extends beyond three years.
    lr = run_v2(copy.deepcopy(longc)); sb = lr["scenarios"]["base"]
    li = lr["financials"]["is"]["ni"]
    lta = lr["financials"]["bs"]["totalAssets"]
    ck("D14 long model scenario cumulative NI stops at submission Q12",
       abs(sb["cum_ni"]-sum(li[:12])) < .02 and abs(sb["cum_ni_full"]-sum(li[:20])) < .02)
    ck("D15 long model Q12 assets are not computational terminal assets",
       abs(sb["q12_total_assets"]-lta[12]) < .02 and abs(sb["terminal_total_assets"]-lta[20]) < .02)
    ck("D16 concentration exhibit is explicitly Q12 on 20Q model", lr["concentrations"]["as_of"] == "Q12")

    from foundry.v2 import exec_view_gen
    md = exec_view_gen.build(mc, run_v2(copy.deepcopy(mc)))
    qd = exec_view_gen.build(base, run_v2(copy.deepcopy(base)))
    ld = exec_view_gen.build(longc, lr)
    def _cum_metric(d): return next(x for x in d["METRICS"] if x.get("id") == "cumni")
    ck("D17 Executive Summary generic horizon labels follow cadence/horizon",
       md["MODEL"]["horizonLabel"] == "36 months" and _cum_metric(md)["sub"] == "36 months total" and
       qd["MODEL"]["horizonLabel"] == "12 quarters" and _cum_metric(qd)["sub"] == "12 quarters total" and
       ld["MODEL"]["horizonLabel"] == "20 quarters" and _cum_metric(ld)["sub"] == "20 quarters total")
    ck("D18 Executive Summary cumulative NI uses full 20Q computational horizon",
       _cum_metric(ld)["value"] == f"${sb['cum_ni_full']/1000:.1f}M" and
       abs(sb["cum_ni_full"]-sb["cum_ni"]) > .01,
       f"full={sb['cum_ni_full']:.2f}, submission={sb['cum_ni']:.2f}, shown={_cum_metric(ld)['value']}")

    # E. Call Report canonical quarterly API ---------------------------------------
    mr = run_v2(copy.deepcopy(mc)); cr = build_call_report(mr, mc)
    ri12 = _row(cr["RI"], "12")["values"]
    ck("E1 monthly Call Report exposes 12 regulatory quarters", len(ri12) == 12)
    ck("E2 Schedule RI Q1 net income sums M1-M3", abs(ri12[0] - sum(mr["financials"]["is"]["ni"][:3])) < .01)
    rc = cr["RC"]
    hfs = _row(rc,"4.a")["values"]; hfi = _row(rc,"4.b")["values"]; acl = _row(rc,"4.c")["values"]; net = _row(rc,"4.d")["values"]
    ck("E3 RC 4.d = HFI less ACL, excluding HFS", max(abs(net[i] - (hfi[i]+acl[i])) for i in range(len(net))) < .02)
    rcc_total = _row(cr["RC-C"], "12")["values"]
    ck("E4 RC-C total = RC 4.a + 4.b", max(abs(rcc_total[i]-(hfs[i]+hfi[i])) for i in range(len(rcc_total))) < .02)
    rcr_t1 = _row(cr["RC-R"], "26")["values"]
    ck("E5 RC-R Tier1 uses canonical capital row", max(abs(x-y) for x,y in zip(rcr_t1, mr["capital"]["rows"]["tier1"][2::3])) < .02)

    # F. CBLR quarter-state machine ------------------------------------------------
    P = REG_PARAMS["cblr"]
    ck("F1 exactly 8% does not satisfy strict >8% requirement",
       _cblr_state_machine([P["requirement"]],[True],P)[0] != "ok")
    st = _cblr_state_machine([.079]*5, [True]*5, P)
    ck("F2 four-quarter grace expires in fifth regulatory quarter", st[:4] == ["grace"]*4 and st[4] == "EXHAUSTED", str(st))
    # Eight historic grace quarters outside the rolling 20 must age out.
    seq = [.079]*4 + [.081]*17 + [.079]
    st2 = _cblr_state_machine(seq, [True]*len(seq), P)
    ck("F3 old grace usage rolls out of previous-20 window", st2[-1] == "grace", f"last={st2[-1]}")

    # G. Profile B public contract -------------------------------------------------
    pb = _load(PF_B); prb = run_v2(copy.deepcopy(pb))
    vals = [round(prb["scenarios"][k]["cum_ni"],2) for k in ("base","credit","rate","combined","dfast_severe")]
    ck("G1 Profile B scenarios are economically distinct", len(set(vals)) >= 4, str(vals))
    fl = copy.deepcopy(pb); fa=fl["assumptions"]; fa["rate_path_q"]=[.04]*12; fa["rate_path_longer_run"]=.04
    fp=fa["lending_products"][0]; fp["rate_type"]="float"; fp["index"]="sofr"; fp["index_spread"]=.03; fp.pop("yield_ann",None)
    fr=run_v2(fl); fx=next(x for x in fr["products"] if x["name"]==fp["name"])
    ck("G2 Profile B SOFR+300bp prices at 7% on 4% curve", abs(fx["rateQ"][0]-7.0)<1e-9, str(fx["rateQ"][:1]))
    cap = copy.deepcopy(pb); cap["assumptions"]["capital_raises"]=[{"quarter":1,"amount":1_000_000}]
    rc0,rc1=run_v2(pb),run_v2(cap)
    ck("G3 Profile B Q1 capital raise hits Q1", abs((rc1["financials"]["bs"]["paidIn"][0]-rc0["financials"]["bs"]["paidIn"][0])-1000)<.01)

    # H. Dated rate curves + FRED provenance ---------------------------------------
    dc = _monthly_a(); dc["charter_profile"]={"target_opening":"2027-05-15"}
    rf = dated_rate_fn({"2027-05-31":.04,"2027-08-31":.03}, .03, dc, 12)
    ck("H1 dated curve anchors to actual monthly calendar", abs(rf(1)-.04)<1e-12 and abs(rf(4)-.03)<1e-12)

    def fake_get(url):
        sid = re.search(r"series_id=([^&]+)", url).group(1)
        obs = {
            "DFEDTARU":[{"date":"2026-09-03","value":"3.75"}],
            "DFEDTARL":[{"date":"2026-09-03","value":"3.50"}],
            "FEDTARMD":[{"date":"2028-01-01","value":"3.4"},{"date":"2027-01-01","value":"3.6"},{"date":"2026-01-01","value":"3.8"}],
            "FEDTARMDLR":[{"date":"2026-01-01","value":"3.1"}],
            "EFFR":[{"date":"2026-09-03","value":"3.64"}],
        }[sid]
        return json.dumps({"observations":obs})
    snap = rate_fetch.fetch_policy("x", fake_get)
    _an=snap["fomc"]["anchors"]
    ck("H2 FRED SEP target years map to same calendar years",
       list(_an)==["2026-12-31","2027-12-31","2028-12-31"] and abs(_an["2026-12-31"]-.038)<1e-12 and abs(_an["2027-12-31"]-.036)<1e-12 and abs(_an["2028-12-31"]-.034)<1e-12, str(_an))
    ck("H3 FRED does not mislabel observation/target date as SEP vintage", snap["fomc"]["source_vintage"] is None and snap["vintage"]["sep"] is None)
    ck("H4 current policy date is labeled observation, not asserted statement", snap["current_policy"]["observation_date"]=="2026-09-03" and snap["current_policy"]["statement_date"] is None)

    # I. NIE / fee cadence ---------------------------------------------------------
    nie = {"categories":[{"name":"x","per_period":100.0,"trajectory":"linear","growth_per_period":.10}]}
    ns = nie_detail_series({"n_periods":4,"nie_detail":nie}, ppy=4)["categories"]
    ck("I1 canonical NIE linear-growth field is consumed", [round(x,1) for x in ns] == [100.0,110.0,121.0,133.1], str(ns))
    legacy_nie = {"categories":[{"name":"x","per_quarter":300.0,"trajectory":"flat"}]}
    nsm = nie_detail_series({"n_periods":3,"nie_detail":legacy_nie}, ppy=12)["categories"]
    ck("I2 legacy $300/qtr NIE becomes $100/month", all(abs(x-100)<1e-9 for x in nsm), str(nsm))
    mn = {"day1":1000.0,"trajectory":"proportional","growth_q":.12}
    _avg,mns = managed_notional_series(mn, 3, ppy=12)
    ck("I3 managed-notional legacy quarterly growth preserves quarter economics", abs(mns[-1]/1000.0-(1.12)) < 1e-10, str(mns))

    # J. Challenge / generic reporting / UI contracts ------------------------------
    html = (ROOT/"web/console_v2.html").read_text(encoding="utf-8")
    exec_html = (ROOT/"foundry/v2/assets/exec_view_template.html").read_text(encoding="utf-8")
    ck("J1 preset insertion converts historical quarterly preset economics", "materializeQuarterlyPreset(pr.p)" in html)
    ck("J2 NIE UI writes canonical growth_spec for new Growth trajectories",
       'cat.growth_spec={rate:g,period:gp,method:gm,anchor:ga,anchor_month:gam}' in html)
    ck("J3 Executive chart uses cadence-aware endpoint labels", "MODEL.periodEndLabel" in exec_html)
    ck("J4 fee/catalog growth labels use active cadence", "`%/${PLAB()}`" in html)
    ck("J5 Call Report UI does not double-collapse canonical quarterly values",
       'REG_FLOW_Q(r.values)' not in html and 'REG_STOCK_Q(r.values)' not in html)
    ck("J6 regulator submission horizon is distinct from computational horizon in UI",
       'function SUBQ()' in html and 'function SUBEND()' in html and 'AT_SUB(' in html)
    ck("J7 stress leverage chart consumes full native-cadence series",
       's.lev_by_period || s.lev_by_q || []' in html and 's.lev_submission_q || []' not in html
       and len(mr["scenarios"]["base"]["lev_by_period"]) == 36)
    ck("J8 methodology period wording resolves from engagement cadence",
       "Loan balances roll forward each {{period}}" in html and "function methodologyHTML()" in html and "h += methodologyHTML();" in html)
    ck("J9 generic financial descriptions use cadence-aware period text",
       "function calcText(key)" in html and "Pro Forma Income Statement (${PPY()===12?'monthly':'quarterly'})" in html)
    ck("J10 raise/borrowing event selectors are cadence-aware rather than hard-coded Qtr",
       "function EVENTUNIT()" in html and "function SETEVENT(ev,v)" in html and "${EVENTSCHEDLABEL()}<select" in html)
    ck("J11 monthly event UI stores explicit period instead of overloading quarter",
       "ev.period=v; delete ev.quarter" in html)
    ck("J12 FHLB presentation consistently says bullet, not amortizing",
       "Scheduled Borrowings (FHLB/Term, bullet)" in html and "Scheduled Borrowings (FHLB/Term, amortizing)" not in html)
    ck("J13 browser calendar has explicit native-period end helper",
       "function _openingModelStart()" in html and "PPY()===12 ? d.getUTCMonth()" in html
       and "function _modelPeriodEnd(p)" in html and "return _modelPeriodEnd(q*MO_PER_Q());" in html
       and "function _openingQuarterStart()" not in html)
    ck("J14 generic Executive Summary cumulative NI is full-horizon while regulatory Q12 remains explicit",
       "base.cum_ni_full ?? base.cum_ni" in html and "${HORIZONLAB()} total" in html
       and "Q1–Q${SUBQ()} submission totals" in html and "Product (Q1–Q${SUBQ()} submission totals" in html)
    ck("J15 generic monthly/quarterly wording has no known fixed-quarter remnants",
       "Engine-computed output for this quarter" not in html
       and '"first profitable quarter"' not in html
       and "End-of-quarter balance" not in html
       and "from its stated quarter" not in html
       and "from its draw quarter" not in html
       and ">Net Income by Quarter<" not in html
       and ">Tier 1 Leverage Ratio by Quarter<" not in html)
    mlab = _browser_label_probe(html, 12, 36); qlab = _browser_label_probe(html, 4, 20)
    ck("J16 cadence helpers render month/M1/36 months vs quarter/Q1/20 quarters",
       mlab == {"period":"month","p1":"M1","horizon":"36 months","event":"Mth","scheduled":"Mth"}
       and qlab == {"period":"quarter","p1":"Q1","horizon":"20 quarters","event":"Qtr","scheduled":"Qtr"},
       f"monthly={mlab}, quarterly={qlab}")
    mcopy = _browser_copy_probe(html, 12, 36); qcopy = _browser_copy_probe(html, 4, 20)
    ck("J16b rendered financial/stress copy follows monthly vs quarterly cadence",
       "scheduled month" in mcopy["paidin"] and "month-end" in mcopy["leverage"]
       and "scheduled draw month" in mcopy["schedBorrow"] and "term in quarters" in mcopy["schedBorrow"]
       and mcopy["stressTitle"] == "Tier 1 Leverage Ratio by Month"
       and "scheduled quarter" in qcopy["paidin"] and "quarter-end" in qcopy["leverage"]
       and "scheduled draw quarter" in qcopy["schedBorrow"] and "term in quarters" in qcopy["schedBorrow"]
       and qcopy["stressTitle"] == "Tier 1 Leverage Ratio by Quarter",
       f"monthly={mcopy}, quarterly={qcopy}")
    ck("J17 event UI uses requested Mth/Qtr cadence and full computational horizon",
       'function EVENTUNIT(){ return PPY()===12?"Mth":PPY()===1?"Yr":"Qtr"; }' in html
       and html.count('Array.from({length:NP()}') >= 2
       and 'cfg.assumptions.n_periods = (wiz.years || 3) * (wiz.ppy || 4)' in html
       and '[[12,"Monthly"],[4,"Quarterly"]]' in html
       and '[[3,"3 years"],[5,"5 years"],[7,"7 years"]]' in html)
    ck("J18 scheduled-event wording distinguishes model period from contractual FHLB term",
       html.count('title="Scheduled model period"') >= 2
       and 'title="contractual term (quarters)"' in html
       and 'Those are modeled as bullet advances: the full draw is held flat from its scheduled draw {{period}} through maturity (the stated term in quarters)' in html)
    nie_entry = _browser_nie_entry_probe(html)
    ck("J19 Operating Expense detail starts empty and supports one-at-a-time category entry",
       nie_entry["afterOn"] == []
       and nie_entry["afterAdd"] == [{"name":"", "per_period":0}]
       and "Paste categories" in html and "+ Add one manually" in html
       and "Core banking & tech" not in html
       and 'categories:[{name:"Core banking & tech"' not in html,
       str(nie_entry))
    ck("J20 scheduled raise/borrowing row label is only the compact cadence token",
       mlab["scheduled"] == "Mth" and qlab["scheduled"] == "Qtr"
       and 'function EVENTSCHEDLABEL(){ return EVENTUNIT(); }' in html
       and 'return "Scheduled "+EVENTUNIT()' not in html)
    span36 = _browser_fin_colspan_probe(html, 36, False)
    span20ref = _browser_fin_colspan_probe(html, 20, True)
    ck("J21 Balance Sheet section bands span the complete native-cadence table width",
       span36 == {"noOpen":37,"withOpen":38,"withOpenTotal":39}
       and span20ref == {"noOpen":22,"withOpen":23,"withOpenTotal":24}
       and html.count('colspan="${finColspan(bsHasOpen,false)}"') == 3
       and 'colspan="${V21?15:14}"' not in html,
       f"36m={span36}, 20q+ref={span20ref}")

    # K. Security hardening ---------------------------------------------------------
    app_src = (ROOT/"app.py").read_text(encoding="utf-8")
    ck("K1 no hard-coded legacy fallback password", "solstice-2026" not in app_src and 'FOUNDRY_PASS", "' not in app_src)
    ck("K2 legacy credential disabled unless explicit env vars exist", "if not USER or not PASS" in app_src)
    ck("K3 auth cookie uses Secure and HttpOnly", "secure=_cookie_secure()" in app_src and "httponly=True" in app_src)
    ck("K4 v2 frozen engagements are owner-scoped on disk", 'os.path.join("clients_v2", _user_key)' in app_src)
    ck("K5 legacy v1 uploaded engagements are owner-scoped", "legacy_engagements" in app_src and "_legacy_engagements_for(user)" in app_src)

    # L. Methodology/docs must not claim generic quarter-only mechanics -------------
    eng_src = (ROOT/"foundry/v2/engine_q_a.py").read_text(encoding="utf-8")
    ck("L1 engine exposes canonical timebase rather than hardcoding duration Q", "quarters_to_periods" in eng_src and "months_to_periods" in eng_src)

    print(f"\n{passed} passed, {failed} failed")
    return 0 if failed == 0 else 1

if __name__ == "__main__":
    raise SystemExit(main())
