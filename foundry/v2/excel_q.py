"""Foundry v2 — Excel intake and exhibit layer (A.14, A.15).

A.14 config workbook: banker-native intake for the v2 schema. Product instances
are rows (one sheet per product family), per-period overrides live on a Vectors
sheet, scalar assumptions and engagement metadata are key/value, and a generated
Data Dictionary documents every field. workbook -> parse -> identical config:
the T15 discipline extended to v2 (many doors, one hallway).

A.15 results workbook: the filing exhibit — balance sheet, income statement,
ratios — with the canonical config hash on the cover. Every number ties to the
engine output exactly.
"""
import io
import json
import hashlib
from openpyxl import Workbook, load_workbook

DEP_COLS = ["name", "opening_balance", "growth_q", "runoff_q", "rate_type",
            "rate_paid_ann", "index_spread", "fee_yield_ann", "opex_pct_ann",
            "opex_fixed_m", "call_report_line"]
LEND_COLS = ["name", "opening_balance", "volume_mode", "originations_q", "orig_growth_q",
             "growth_q", "runoff_q", "rate_type", "yield_ann", "index_spread",
             "charge_off_ann", "provision_rate_ann", "reserve_rate_pct_bal",
             "measurement", "discount_spread_ann", "fee_yield_ann", "opex_pct_ann",
             "opex_fixed_m", "call_report_line",
             "mb_sale_pct_of_orig", "mb_gain_on_sale_margin", "mb_warehouse_hold_q",
             "mb_servicing_retained_pct", "mb_servicing_fee_bp_ann",
             "mb_msr_cap_rate_pct_upb", "mb_msr_decay_q"]
OBS_COLS = ["name", "notional", "growth_q", "fee_yield_ann", "opex_fixed_m"]
SEC_COLS = ["name", "opening", "purchases_q", "growth_q", "runoff_q", "yield_ann",
            "fee_yield_ann", "opex_fixed_m"]
JSON_TOP = ["step_minus_1", "step_0", "step_0a", "step_1", "assumption_tags",
            "constraints", "target_state", "peer_query", "scenario_overlays",
            "parity_expectation"]
SCALAR_TOP = ["engagement_id", "schema_version", "client_legal_name", "proposed_bank",
              "hq", "prepared_by", "config_version", "config_frozen", "parity_profile"]
OV_FIELDS = ["growth_q", "runoff_q", "rate_paid_ann", "yield_ann", "index_spread",
             "charge_off_ann", "originations_q"]

DICT_ROWS = [
    ("Deposits", "opening_balance", "number", "$", "required", "> 0 typical", "Q1 opening balance."),
    ("Deposits", "growth_q", "number", "rate/qtr", "required", "[-0.5, 1.0]", "Quarterly balance growth."),
    ("Deposits", "runoff_q", "number", "rate/qtr", "optional", "[0, 1]", "Quarterly runoff; negative mints balances and is rejected."),
    ("Deposits", "rate_type", "text", "—", "required", "fixed | float", "float prices at index + index_spread off the rate path."),
    ("Deposits", "rate_paid_ann", "number", "annual rate", "fixed only", "[0, 0.30]", "Rate paid on balances."),
    ("Lending", "originations_q", "number", "$/qtr", "originations mode", ">= 0", "New volume per quarter, grown by orig_growth_q."),
    ("Lending", "charge_off_ann", "number", "annual rate", "required", "[0, 0.40]", "Net charge-offs on average balances."),
    ("Lending", "provision_rate_ann", "number", "annual rate", "optional", "[0, 0.40]", "Expected-loss provisioning; blank = charge-off rate; entity ALLL floor still applies."),
    ("Lending", "measurement", "text", "—", "required", "amortized | fair_value", "fair_value carries no ALLL; requires discount_spread_ann."),
    ("Lending", "mb_sale_pct_of_orig", "number", "share", "OTS only", "[0, 1]", "Share of originations designated held-for-sale."),
    ("Lending", "mb_msr_cap_rate_pct_upb", "number", "share of UPB", "servicing only", "[0, 0.03]", "MSR capitalized at settlement, part of the gain on sale."),
    ("Vectors", "(any driver)", "number", "per field", "optional", "field's range", "Per-quarter override; blank quarters use the product's baseline."),
    ("Assumptions", "rate_path_q", "12 numbers", "annual rates", "required", "[0, 0.30]", "Forward index path; glides to rate_path_longer_run beyond Q12."),
    ("Assumptions", "tax_semantics", "text", "—", "required", "profile-defined", "Tax sequencing election (parity modes documented in ENGINE_SPEC)."),
]


def _cfg_hash(cfg):
    return hashlib.sha256(json.dumps(cfg, sort_keys=True, default=str).encode()).hexdigest()[:12]


def _write_products(ws, cols, rows):
    ws.append(cols)
    for p in rows:
        mb = p.get("mortgage_banking") or {}
        out = []
        for c in cols:
            if c.startswith("mb_"):
                out.append(mb.get(c[3:]))
            else:
                v = p.get(c)
                out.append(v)
        ws.append(out)


def _read_products(ws, cols, has_mb=False):
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    header = list(rows[0])
    out = []
    for r in rows[1:]:
        if r[0] is None:
            continue
        p, mb = {}, {}
        for k, v in zip(header, r):
            if k is None:
                continue
            if k.startswith("mb_"):
                if v is not None:
                    mb[k[3:]] = v
            elif v is not None or k in ("provision_rate_ann", "reserve_rate_pct_bal"):
                p[k] = v
        if has_mb and mb:
            p["mortgage_banking"] = mb
        out.append(p)
    return out


def workbook_from_config_v2(cfg):
    wb = Workbook()
    eng = wb.active
    eng.title = "Engagement"
    eng.append(["key", "value"])
    for k in SCALAR_TOP:
        if k in cfg:
            eng.append([k, cfg[k]])
    for k in JSON_TOP:
        if k in cfg and cfg[k] is not None:
            eng.append([k, json.dumps(cfg[k], sort_keys=True)])

    a = cfg["assumptions"]
    asum = wb.create_sheet("Assumptions")
    asum.append(["key", "value"])
    for k, v in a.items():
        if k in ("deposit_products", "lending_products", "obs_exposures",
                 "securities_afs", "securities_htm"):
            continue
        if isinstance(v, (list, dict)):
            asum.append([k, json.dumps(v)])
        else:
            asum.append([k, v])

    _write_products(wb.create_sheet("Deposits"), DEP_COLS, a.get("deposit_products") or [])
    _write_products(wb.create_sheet("Lending"), LEND_COLS, a.get("lending_products") or [])
    _write_products(wb.create_sheet("OBS"), OBS_COLS, a.get("obs_exposures") or [])
    _write_products(wb.create_sheet("SecuritiesAFS"), SEC_COLS, a.get("securities_afs") or [])
    _write_products(wb.create_sheet("SecuritiesHTM"), SEC_COLS, a.get("securities_htm") or [])

    vec = wb.create_sheet("Vectors")
    vec.append(["product", "field"] + [f"Q{q}" for q in range(1, 13)])
    for fam in ("deposit_products", "lending_products", "obs_exposures"):
        for p in a.get(fam) or []:
            for f, m in (p.get("overrides") or {}).items():
                row = [p["name"], f] + [m.get(str(q)) for q in range(1, 13)]
                vec.append(row)

    dd = wb.create_sheet("DataDictionary")
    dd.append(["sheet", "field", "type", "units", "required for", "range", "description"])
    for r in DICT_ROWS:
        dd.append(list(r))
    return wb


def parse_workbook_v2(data):
    wb = load_workbook(io.BytesIO(data) if isinstance(data, (bytes, bytearray)) else data,
                       data_only=True)
    cfg = {}
    for k, v in ((r[0], r[1]) for r in wb["Engagement"].iter_rows(min_row=2, values_only=True)):
        if k is None:
            continue
        cfg[k] = json.loads(v) if k in JSON_TOP else v
    a = {}
    for k, v in ((r[0], r[1]) for r in wb["Assumptions"].iter_rows(min_row=2, values_only=True)):
        if k is None:
            continue
        if isinstance(v, str) and v[:1] in "[{":
            a[k] = json.loads(v)
        else:
            a[k] = v
    a["deposit_products"] = _read_products(wb["Deposits"], DEP_COLS)
    a["lending_products"] = _read_products(wb["Lending"], LEND_COLS, has_mb=True)
    a["obs_exposures"] = _read_products(wb["OBS"], OBS_COLS)
    afs = _read_products(wb["SecuritiesAFS"], SEC_COLS)
    htm = _read_products(wb["SecuritiesHTM"], SEC_COLS)
    a["securities_afs"] = afs or None
    a["securities_htm"] = htm or None
    by_name = {p["name"]: p for fam in ("deposit_products", "lending_products", "obs_exposures")
               for p in a[fam]}
    for r in wb["Vectors"].iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        p = by_name.get(r[0])
        if p is None:
            continue
        ov = p.setdefault("overrides", {})
        m = {str(q): r[1 + q] for q in range(1, 13) if r[1 + q] is not None}
        if m:
            ov[r[1]] = m
    cfg["assumptions"] = a
    if "scenario_overlays" not in cfg:
        cfg["scenario_overlays"] = None
    return cfg


def results_workbook_v2(cfg, res):
    """A.15 — the exhibit. Every number ties to the parity-shaped results exactly."""
    wb = Workbook()
    cover = wb.active
    cover.title = "Cover"
    cover.append([cfg.get("proposed_bank", "Pro Forma"), ""])
    cover.append(["Config hash (canonical)", _cfg_hash(cfg)])
    cover.append(["Schema version", cfg.get("schema_version")])
    cover.append(["Config frozen", cfg.get("config_frozen")])
    cover.append(["Units", "$ thousands"])
    cover.append(["Note", "Identical configuration reproduces identical figures and this hash, forever."])

    from .callreport import code_for_result, code_for_line
    from . import present
    derived = present.derived_lines(res, cfg)
    prods = res.get("products") or []

    def sheet_from_layout(ws, layout, fin, n, has_open):
        cols = (["Open"] if has_open else []) + [f"Q{q}" for q in range(1, (n - 1 if has_open else n) + 1)]
        ws.append(["Line item ($000s)", "key", "Schedule", "Item", "Code"] + cols)
        for row in layout:
            t = row["t"]
            if t == "spacer":
                ws.append([]); continue
            if t == "section":
                ws.append([row["label"]]); continue
            if t == "detail":
                fam_p = [p for p in prods if p["family"] == row["family"] and p.get("bal")]
                if not fam_p:
                    continue
                ws.append([row["label"]])
                for p in fam_p:
                    c = code_for_line(p.get("line")) or ("", "", "", "")
                    arr = p["bal"]
                    if not has_open and len(arr) == n + 1:
                        arr = arr[1:]
                    ws.append(["    " + p["name"], "", c[0], c[1], c[2]] + list(arr))
                continue
            if t == "identity":
                ws.append([row["label"], "identity", "", "", ""] +
                          ["OK" if abs(x) <= 0.02 else x for x in derived["identity"]])
                continue
            key = row["key"]
            arr = fin.get(key, derived.get(key))
            if arr is None:
                continue
            c = code_for_result(key) or ("", "", "", "")
            label = ("    " if row.get("indent") else "") + row["label"]
            vals = [None if x is None else (-x if row.get("negate") else x) for x in arr]
            ws.append([label, key, c[0], c[1], c[2]] + vals)

    n_bs = len(res["bs"]["totalAssets"])
    sheet_from_layout(wb.create_sheet("Balance Sheet"), present.BS_LAYOUT, res["bs"], n_bs, n_bs == 13)
    n_is = len(res["is"]["ni"])
    sheet_from_layout(wb.create_sheet("Income Statement"), present.IS_LAYOUT, res["is"], n_is, False)

    if res.get("ratios"):
        rt = wb.create_sheet("Ratios")
        rt.append(["Ratio (%)", "key"] + [f"Q{q}" for q in range(1, n_is + 1)])
        for k, arr in res["ratios"].items():
            if arr and any(x is not None for x in arr):
                rt.append([present.RATIO_LABELS.get(k, k), k] + list(arr))
    # FLOOR F-001: engagement cover sheet — every artifact answers who/what/which version
    try:
        ee = res.get("engagement_echo") if isinstance(res, dict) else None
        if ee:
            cv = wb.create_sheet("Engagement", 0)
            for k, lbl in (("client", "Client"), ("engagement_id", "Engagement"),
                            ("prepared_by", "Prepared by"), ("config_version", "Config version"),
                            ("config_hash", "Config hash"), ("engine_version", "Engine")):
                cv.append([lbl, ee.get(k)])
    except Exception:
        pass
    # FLOOR F-133: Call-Report-named schedule sheets, per-line references
    try:
        from .callreport import build_call_report
        _shim = res if "financials" in res else {
            "financials": {"bs": res["bs"], "is": res["is"],
                             "ratios": res.get("ratios") or {}},
            "products": res.get("products") or [], "capital": res.get("capital")}
        cr = build_call_report(_shim, cfg)
        for sched in ("RC", "RI", "RC-C", "RC-E", "RC-R"):
            if sched not in cr:
                continue
            ws = wb.create_sheet(f"Schedule {sched}")
            ws.append([cr[sched].get("title", sched)])
            ws.append(["Item", "Code", "Line"] + [f"Q{q}" for q in range(1, 13)])
            for row in cr[sched]["rows"]:
                ws.append([row["item"], row["code"], row["label"]] + list(row["values"]))
            if sched == "RC-R" and cr[sched].get("part2"):
                ws.append([])
                ws.append([cr[sched]["part2"]["title"]])
                for row in cr[sched]["part2"]["rows"]:
                    ws.append([row["item"], row["code"], row["label"]] + list(row["values"]))
            for note in (cr[sched].get("notes") or []):
                ws.append([])
                ws.append([f"note: {note}"])
    except Exception:
        pass   # schedules are additive; the workbook core never fails on them

    return wb
