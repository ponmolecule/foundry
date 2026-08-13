"""Foundry Input Workbook (FIW) — the per-engagement Excel (INPUT_SPEC §7).

Unlike the full config workbook (excel_q.py, every field always), the FIW is
generated from one engagement's state: CONTROL from identity and capital, one
ASSM sheet per product family actually present, LIMITS, and a README carrying
the generation hash that makes the diff-import (build step 3) possible.
Progressive disclosure extends into the paper: a deposits-only bank gets no
loan sheet; a product without originate-to-sell shows no mb_ rows.

Sheet grammar (stable for the diff-import): column A holds a machine key
(family.index.field) — hidden; columns B..D are Product | Field | Value with
units in E. Fact rows (Call Report line) are marked read-only in E.
"""
import io
import json
import hashlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation
from foundry.v2.cadence import workbook_units
from foundry.v2 import wbunits

GOLD = PatternFill("solid", fgColor="FFF3D9")
HDR = Font(bold=True)

# LOAN_FIELDS was replaced by the branch-aware LEND_FIELDS below (v41): full field set
# restored, single contextual rate cell (__RATE__) routed by rate_type at import.
MB_FIELDS = [
    ("mortgage_banking.sale_pct_of_orig", "Originations sold", "share [0,1]"),
    ("mortgage_banking.gain_on_sale_margin", "Gain-on-sale margin", "share"),
    ("mortgage_banking.warehouse_hold_q", "Warehouse period", "quarters held for sale"),
    ("mortgage_banking.servicing_retained_pct", "Servicing retained", "share of sold"),
    ("mortgage_banking.servicing_fee_bp_ann", "Servicing fee", "bp/yr"),
    ("mortgage_banking.msr_cap_rate_pct_upb", "MSR cap rate", "% of UPB"),
    ("mortgage_banking.msr_decay_q", "Servicing runoff", "rate/qtr"),
]
DEP_FIELDS = [
    ("call_report_line", "Call Report line", "fact — resolved, do not edit"),
    ("opening_balance", "Opening balance", "$"),
    ("growth_q", "Balance growth", "rate/qtr"),
    ("new_deposits_q", "New deposits", workbook_units("new_deposits_q")),
    ("avg_maturity_m", "Average maturity", workbook_units("avg_maturity_m")),
    ("runoff_q", "Runoff", "rate/qtr (annual / 4)"),
    ("rate_type", "Rate type", "fixed | float  (editable)"),
    # __RATE__ is a synthetic column: one contextual cell whose meaning is set by rate_type.
    # The generator writes it; diff_import routes it into rate_paid_ann OR index_spread.
    ("__RATE__", "Rate paid / Spread over SOFR (% p.a.)",
     "fixed: annual rate (4.90% = 0.049).  float: spread over SOFR, may be negative"),
    ("__RATE_BASELINE__", "Interpreted as (live)", "fact — informational, follows the Rate type dropdown, not imported"),
    ("fee_yield_ann", "Fee yield", "annual % of avg balance"),
    ("opex_pct_ann", "Op cost (% of avg bal)", "annual rate"),
    ("opex_fixed_q", "Op cost fixed", workbook_units("opex_fixed_q")),
    ("insured_pct", "Insured share", "share [0,1]"),
]
LEND_FIELDS = [
    ("call_report_line", "Call Report line", "fact — resolved, do not edit"),
    ("opening_balance", "Opening balance (Day 1)", "$"),
    ("originations_q", "New originations", workbook_units("originations_q")),
    ("orig_growth_q", "Origination growth", "rate/qtr"),
    ("runoff_q", "Prepayment / paydown", "rate/qtr (annual /4)"),
    ("rate_type", "Rate type", "fixed | float  (editable)"),
    ("__RATE__", "Yield / Spread over SOFR (% p.a.)",
     "fixed: annual yield (6.50% = 0.065).  float: spread over SOFR, may be negative"),
    ("__RATE_BASELINE__", "Interpreted as (live)", "fact — informational, follows the Rate type dropdown, not imported"),
    ("measurement", "Measurement", "amortized | fair_value  (editable)"),
    ("discount_spread_ann", "DCF discount spread (fair_value only)", "annual rate"),
    ("charge_off_ann", "Net charge-offs", "annual rate"),
    ("provision_rate_ann", "Provision rate (blank = NCO)", "annual rate"),
    ("reserve_rate_pct_bal", "ALLL reserve", "% of balance"),
    ("fee_yield_ann", "Fee yield", "annual % of avg balance"),
    ("opex_pct_ann", "Op cost (% of avg bal)", "annual rate"),
    ("opex_fixed_q", "Op cost fixed", workbook_units("opex_fixed_q")),
]
LINE_LABELS = {
    "loanCommercial": "Loans: Commercial & Industrial", "loanConsumer": "Loans: Consumer",
    "loanCreditCard": "Loans: Credit Card", "loanMortgage": "Loans: 1-4 Family Residential",
    "loanOther": "Loans: Other", "loanCRE": "Loans: Commercial Real Estate",
    "depDDA": "Deposits: Transaction (DDA)", "depSavings": "Deposits: Savings & MMDA",
    "depTime": "Deposits: Time", "depBrokered": "Deposits: Brokered",
    "depSweep": "Deposits: Sweep (reciprocal/one-way)",
    "depInstitutional": "Deposits: Institutional & listing-service",
}


def _version_stamp():
    """Lineage stamp for a generated workbook. Two independent markers, each honest:
    the human-assigned BUILD_STAMP (always present, from the committed file) and the
    deployment's git commit SHA (machine-truth, only when the platform exposes it via
    env var). Never fabricates a SHA — absent means 'not available', not a stale hash."""
    import os
    build = ""
    try:
        # BUILD_STAMP lives at the repo/app root (one level above foundry/).
        root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        p = os.path.join(root, "BUILD_STAMP")
        if os.path.exists(p):
            build = open(p, encoding="utf-8").read().strip()
    except Exception:
        build = ""
    if not build:
        build = "unknown"
    # The true deployed commit, if the platform injects it. Check an ordered list of
    # candidate env vars so the exact name isn't a single point of failure.
    sha = ""
    for var in ("RAILWAY_GIT_COMMIT_SHA", "GIT_COMMIT_SHA", "GIT_SHA",
                "SOURCE_VERSION", "COMMIT_SHA"):
        v = os.environ.get(var)
        if v:
            sha = v.strip()[:12]
            break
    return build, (sha or "not available")


# Editable per-category field lists for the flat assumption arrays. Same grammar as
# LOAN_FIELDS/DEP_FIELDS: (config_key, human_label, units_note). "fact" in the note
# marks a row read-only (never gold, never read back by diff_import).
SEC_FIELDS = [
    ("name", "Name", "label (editable)"),
    ("opening", "Opening balance", "$"),
    ("purchases_q", "Purchases / additions", "$/quarter"),
    ("growth_q", "Balance growth", "rate/qtr"),
    ("runoff_q", "Runoff / maturities", "rate/qtr"),
    ("yield_ann", "Average yield", "annual rate"),
]
OBS_FIELDS = [
    ("name", "Name", "label (editable)"),
    ("notional", "Notional", "$"),
    ("growth_q", "Notional growth", "rate/qtr"),
    ("fee_yield_ann", "Fee yield", "annual % of notional"),
    ("opex_pct_ann", "Op cost", "annual % of notional"),
    ("opex_fixed_q", "Op cost fixed", workbook_units("opex_fixed_q")),
]
SCHED_FIELDS = [
    ("name", "Name", "label (editable)"),
    ("amount", "Draw amount", "$"),
    ("quarter", "Draw quarter", "quarter 1-12"),
    ("term_q", "Term to maturity", "quarters (bullet)"),
    ("rate_ann", "Rate", "annual rate"),
]
RAISE_FIELDS = [
    ("quarter", "Raise quarter", "quarter 1-12"),
    ("amount", "Raise amount", "$"),
]
PREOPEN_FIELDS = [
    ("category", "Category", "text"),
    ("total", "Total", "$"),
]


def cfg_hash(cfg):
    return hashlib.sha256(json.dumps(cfg, sort_keys=True, default=str).encode()).hexdigest()[:12]


def _kv_sheet(ws, rows):
    ws.append(["Item", "Value", "Note"])
    for c in ws[1]:
        c.font = HDR
    for label, value, note, editable in rows:
        ws.append([label, value, note])
        if editable:
            ws.cell(ws.max_row, 2).fill = GOLD
        if isinstance(value, (int, float)) and abs(value) >= 1000:
            ws.cell(ws.max_row, 2).number_format = "#,##0"
    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 46


def _rate_of(p):
    """The single contextual rate value for a product: rate_paid_ann/yield_ann when fixed,
    index_spread when float. Returns None if the applicable field is absent."""
    if p.get("rate_type") == "float":
        return p.get("index_spread")
    return p.get("rate_paid_ann") if "rate_paid_ann" in p else p.get("yield_ann")


def _family_sheet(ws, fam, products, fields):
    ws.append(["key", "Product", "Field", "Value", "Units / note"])
    for c in ws[1]:
        c.font = HDR
    # data-validation dropdowns: rate_type (fixed|float) and measurement (amortized|fair_value).
    dv_rt = DataValidation(type="list", formula1='"fixed,float"', allow_blank=False)
    dv_ms = DataValidation(type="list", formula1='"amortized,fair_value"', allow_blank=False)
    ws.add_data_validation(dv_rt)
    ws.add_data_validation(dv_ms)
    for i, p in enumerate(products):
        rt_row = None  # remember the rate_type cell's row for this product (for the live helper)
        for fkey, flabel, funits in fields:
            if fkey.startswith("mortgage_banking.") and not p.get("mortgage_banking"):
                continue
            if fkey == "discount_spread_ann" and p.get("measurement") != "fair_value":
                continue
            if fkey == "__RATE__":
                val = _rate_of(p)
            elif fkey == "__RATE_BASELINE__":
                val = None  # set below as a live formula, once rt_row is known
            elif "." in fkey:
                a, b = fkey.split(".", 1)
                val = (p.get(a) or {}).get(b)
            else:
                val = p.get(fkey)
            if fkey == "call_report_line":
                val = LINE_LABELS.get(val, val)
            ws.append([f"{fam}.{i}.{fkey}", p.get("name", ""), flabel,
                       wbunits.to_workbook(val, funits) if val is not None else "",
                       wbunits.export_label(funits)])
            row = ws.max_row
            # attach the dropdown + remember the row for the discriminator cells
            if fkey == "rate_type":
                dv_rt.add(ws.cell(row, 4))
                rt_row = row
            elif fkey == "measurement":
                dv_ms.add(ws.cell(row, 4))
            elif fkey == "__RATE_BASELINE__" and rt_row is not None:
                # LIVE helper: the choice in the rate_type dropdown immediately sets what the
                # value cell means. Formula, informational only, never read on import.
                ws.cell(row, 4).value = (
                    f'=IF($D${rt_row}="float",'
                    '"Spread over SOFR (annual %; may be negative)",'
                    '"Rate paid / yield (annual %)")')
            if "fact" not in funits:
                ws.cell(row, 4).fill = GOLD
            if isinstance(val, (int, float)) and abs(val) >= 1000:
                ws.cell(row, 4).number_format = "#,##0"
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 24
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 40


def _array_sheet(ws, arrkey, items, fields):
    """Editable sheet for a flat assumption array (securities, obs, scheduled
    borrowings, capital raises). Same grammar as _family_sheet: hidden machine key
    in column A ('arrkey.index.field'), Value in D, units in E, gold unless 'fact'.
    diff_import reads these back by the column-A key."""
    ws.append(["key", "Item", "Field", "Value", "Units / note"])
    for c in ws[1]:
        c.font = HDR
    for i, it in enumerate(items):
        label = it.get("name", it.get("category", f"{arrkey} {i + 1}"))
        for fkey, flabel, funits in fields:
            val = it.get(fkey)
            ws.append([f"{arrkey}.{i}.{fkey}", label, flabel,
                       wbunits.to_workbook(val, funits) if val is not None else "",
                       wbunits.export_label(funits)])
            if "fact" not in funits:
                ws.cell(ws.max_row, 4).fill = GOLD
            if isinstance(val, (int, float)) and abs(val) >= 1000:
                ws.cell(ws.max_row, 4).number_format = "#,##0"
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 30


def _nie_sheet(ws, nd):
    """Editable sheet for nie_detail — a heterogeneous structure (scalars + a fixed FTE ladder +
    a categories array), so each leaf gets a dotted machine key that diff_import reads back via
    the generic ARRAY_SHEETS reader. Same grammar as _fee_sheet. NIE detail drives total
    non-interest expense (it REPLACES the corporate-overhead lump when present), so it earns a
    first-class editable sheet like every other driving input — previously it was read-only on
    SETTINGS with only FTE+comp shown."""
    ws.append(["key", "Section", "Field", "Value", "Units / note"])
    for c in ws[1]:
        c.font = HDR

    def _row(key, section, field, val, units):
        ws.append([key, section, field,
                   wbunits.to_workbook(val, units) if val is not None else "",
                   wbunits.export_label(units)])
        if "fact" not in units:
            ws.cell(ws.max_row, 4).fill = GOLD
        if isinstance(val, (int, float)) and abs(val) >= 1000:
            ws.cell(ws.max_row, 4).number_format = "#,##0"

    # FTE ladder (one row per modeled year) — the headcount that drives loaded comp.
    fte = nd.get("fte_by_year") or []
    for i, n in enumerate(fte):
        _row(f"nie_detail.fte_by_year.{i}", "Staffing", f"FTE — year {i + 1}", n, "headcount")
    # comp + gross-up scalars.
    if nd.get("loaded_comp_annual") is not None:
        _row("nie_detail.loaded_comp_annual", "Staffing", "Loaded comp per FTE",
             nd.get("loaded_comp_annual"), "$/year (fully loaded)")
    if nd.get("other_gross_up_rate") is not None:
        _row("nie_detail.other_gross_up_rate", "Overhead", "Other gross-up rate",
             nd.get("other_gross_up_rate"), "share of subtotal")
    # regulatory assessment rates — engagement assumptions (12 CFR 327 FDIC / 12 CFR 8 OCC).
    # Only emitted when overridden; a blank cell on re-import falls back to the REG_PARAMS default.
    if nd.get("fdic_bp_ann") is not None:
        _row("nie_detail.fdic_bp_ann", "Assessments", "FDIC assessment rate",
             nd.get("fdic_bp_ann"), "bp/year (blank = 5.0 default)")
    if nd.get("occ_bp_ann") is not None:
        _row("nie_detail.occ_bp_ann", "Assessments", "OCC assessment rate",
             nd.get("occ_bp_ann"), "bp/year (blank = 1.5 default)")
    # category lines (array of {name, per_quarter}).
    for i, cat in enumerate(nd.get("categories") or []):
        _row(f"nie_detail.categories.{i}.name", f"Category {i + 1}", "Name",
             cat.get("name"), "label (editable)")
        _row(f"nie_detail.categories.{i}.per_quarter", f"Category {i + 1}", "Amount",
             cat.get("per_quarter"), "$/quarter")
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 28


def _fee_sheet(ws, fm):
    """Editable sheet for fee_modules — a nested/heterogeneous structure, so each
    scalar leaf gets a dotted machine key. Payments is a list of rails; each rail's
    fields key as 'fee_modules.payments.<i>.<field>'. diff_import reads these back."""
    ws.append(["key", "Module", "Field", "Value", "Units / note"])
    for c in ws[1]:
        c.font = HDR

    def _row(key, module, field, val, units):
        ws.append([key, module, field,
                   wbunits.to_workbook(val, units) if val is not None else "",
                   wbunits.export_label(units)])
        if "fact" not in units:
            ws.cell(ws.max_row, 4).fill = GOLD
        if isinstance(val, (int, float)) and abs(val) >= 1000:
            ws.cell(ws.max_row, 4).number_format = "#,##0"

    ic = fm.get("interchange") or {}
    for k, lab, u in (("tx_count_q", "Transactions/qtr", "count"),
                       ("growth_q", "Growth", "rate/qtr"),
                       ("avg_ticket", "Avg ticket", "$/unit"),
                       ("interchange_rate", "Interchange rate", "share"),
                       ("network_fee_rate", "Network fee rate", "share")):
        if k in ic:
            _row(f"fee_modules.interchange.{k}", "interchange", lab, ic.get(k), u)
    for i, rail in enumerate(fm.get("payments") or []):
        _row(f"fee_modules.payments.{i}.rail", f"payments[{i}]", "Rail", rail.get("rail"), "label (editable: ACH, wires, RTP, FedNow, card)")
        for k, lab, u in (("vol_q", "Volume/qtr", "count"), ("growth_q", "Growth", "rate/qtr"),
                           ("fee_per_tx", "Fee per tx", "$/tx"), ("cost_per_tx", "Cost per tx", "$/tx")):
            if k in rail:
                _row(f"fee_modules.payments.{i}.{k}", f"payments[{i}]", lab, rail.get(k), u)
    sc = fm.get("service_charges") or {}
    for k, lab, u in (("accounts", "Accounts", "count"), ("growth_q", "Growth", "rate/qtr"),
                       ("fee_m", "Fee per acct", "$/account/month")):
        if k in sc:
            _row(f"fee_modules.service_charges.{k}", "service_charges", lab, sc.get(k), u)
    tr = fm.get("trust") or {}
    for k, lab, u in (("aum_open", "AUM opening", "$"), ("aum_growth_q", "AUM growth", "rate/qtr"),
                       ("fee_bp_ann", "Fee", "bp/yr")):
        if k in tr:
            _row(f"fee_modules.trust.{k}", "trust", lab, tr.get(k), u)
    ba = fm.get("baas") or {}
    for k, lab, u in (("programs", "Programs", "programs"), ("accts_per_program", "Accts/program", "count"),
                       ("growth_q", "Growth", "rate/qtr"), ("rev_per_acct_m", "Rev per acct", workbook_units("rev_per_acct_m"))):
        if k in ba:
            _row(f"fee_modules.baas.{k}", "baas", lab, ba.get(k), u)
    ws.column_dimensions["A"].hidden = True
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 26


def build_fiw(cfg):
    a = cfg["assumptions"]
    ch = cfg.get("charter_profile") or {}
    gh = cfg_hash(cfg)
    wb = Workbook()

    rd = wb.active
    rd.title = "README"
    rd.append(["FOUNDRY INPUT WORKBOOK"]); rd["A1"].font = Font(bold=True, size=14)
    rd.append([f"Engagement: {cfg.get('proposed_bank', '')}"])
    rd.append([f"Scenario: {cfg.get('scenario_name', '')}"])
    rd.append(["Generation hash", gh])
    _bstamp, _sha = _version_stamp()
    rd.append(["Model build stamp", _bstamp])
    rd.append(["Model git commit", _sha])
    rd.append([])
    rd.append(["How to use this workbook"])
    rd.append(["Gold cells are yours to edit. Plain cells are facts or context — the app"])
    rd.append(["resolves them and the import will not read edits to them."])
    rd.append(["This workbook contains only the sheets your products require. Upload it"])
    rd.append(["back through the same door it came from; the import validates everything"])
    rd.append(["and lists open questions rather than guessing."])
    rd.append([])
    rd.append(["Units — the same as the app"])
    rd.append(["Every cell asks for exactly the number you would type in the app. Dollar"])
    rd.append(["amounts are in THOUSANDS: type 12,000 to mean $12,000,000. The unit is stated"])
    rd.append(["in each row's 'Units / note' column ($000s, $000s/quarter, $000s/month)."])
    rd.append(["Per-unit prices (a $42 ticket, a 30-cent per-transaction fee) are plain dollars,"])
    rd.append(["labelled $/unit or $/tx. Counts (accounts, transactions) are also in thousands,"])
    rd.append(["labelled 000s. Rates are decimals (0.049 = 4.9%)."])
    rd.append([f"Units convention: v{wbunits.WB_UNITS_VERSION} ($000s). Older workbooks in whole dollars still import correctly."])
    rd.column_dimensions["A"].width = 74

    ct = wb.create_sheet("CONTROL")
    _kv_sheet(ct, [
        ("Institution", cfg.get("proposed_bank", ""), "", True),
        ("Legal name", cfg.get("client_legal_name", ""), "", True),
        ("Home state / market", cfg.get("hq", ""), "", True),
        ("Charter type", ch.get("charter_type", ""), "national | national_trust | state_member | state_nonmember", True),
        ("Primary federal regulator", ch.get("regulator", ""), "fact — resolved from charter type", False),
        ("Target opening", ch.get("target_opening", ""), "", True),
        ("Pre-opening period (months)", ch.get("pre_open_months", ""), "", True),
        ("CBLR election", "yes" if ch.get("cblr_election") else "no", "community bank leverage framework", True),
        ("Initial capital ($000s)", wbunits.to_workbook((cfg.get("target_state") or {}).get("initial_capital", "") or None, "$") if (cfg.get("target_state") or {}).get("initial_capital", "") not in ("", None) else "", "", True),
        ("Scenario name", cfg.get("scenario_name", ""), "", True),
    ] + [(f"Staged raise {i+1} — quarter", r.get("quarter", ""), "1..12", True)
          for i, r in enumerate(a.get("capital_raises") or [])]
      + [(f"Staged raise {i+1} — amount ($000s)", wbunits.to_workbook(r.get("amount", "") or None, "$") if r.get("amount", "") not in ("", None) else "", "", True)
          for i, r in enumerate(a.get("capital_raises") or [])])

    if a.get("lending_products"):
        _family_sheet(wb.create_sheet("ASSM_LOANS"), "lending", a["lending_products"], LEND_FIELDS + MB_FIELDS)
    if a.get("deposit_products"):
        _family_sheet(wb.create_sheet("ASSM_DEPOSITS"), "deposit", a["deposit_products"], DEP_FIELDS)

    lm = wb.create_sheet("LIMITS")
    rows = []
    for con in (cfg.get("constraints") or []):
        rows.append((con.get("name", con.get("metric", "constraint")), con.get("value", ""),
                      con.get("source", ""), True))
    rows.append(("Regulatory thresholds", "resolved from REG_PARAMS", "versioned, cited — never typed", False))
    _kv_sheet(lm, rows)

    # Editable sheets for the flat assumption arrays and optional structures —
    # progressive disclosure: a sheet appears only when its category is populated.
    # These round-trip through diff_import exactly like the loan/deposit sheets.
    if a.get("securities_afs"):
        _array_sheet(wb.create_sheet("ASSM_SEC_AFS"), "securities_afs", a["securities_afs"], SEC_FIELDS)
    if a.get("securities_htm"):
        _array_sheet(wb.create_sheet("ASSM_SEC_HTM"), "securities_htm", a["securities_htm"], SEC_FIELDS)
    if a.get("obs_exposures"):
        _array_sheet(wb.create_sheet("ASSM_OBS"), "obs_exposures", a["obs_exposures"], OBS_FIELDS)
    if a.get("scheduled_borrowings"):
        _array_sheet(wb.create_sheet("ASSM_BORROWINGS"), "scheduled_borrowings",
                     a["scheduled_borrowings"], SCHED_FIELDS)
    if a.get("capital_raises"):
        _array_sheet(wb.create_sheet("ASSM_RAISES"), "capital_raises", a["capital_raises"], RAISE_FIELDS)
    if a.get("fee_modules"):
        _fee_sheet(wb.create_sheet("ASSM_FEES"), a["fee_modules"])
    if a.get("nie_detail"):
        _nie_sheet(wb.create_sheet("ASSM_NIE"), a["nie_detail"])
    po = (cfg.get("pre_opening") or {}).get("expenses")
    if po:
        _array_sheet(wb.create_sheet("ASSM_PREOPEN"), "pre_opening.expenses", po, PREOPEN_FIELDS)

    buf = io.BytesIO()
    _settings_sheet(wb, cfg)   # human-readable review of every in-app-configured input
    _embed_state(wb, cfg)      # the workbook carries its own generation state
    wb.save(buf)
    return buf.getvalue(), gh


# ---------------------------------------------------------------- snapshots
import os


def _snap_dir():
    base = os.environ.get("FOUNDRY_DATA_DIR", os.path.join(os.getcwd(), "data"))
    d = os.path.join(base, "fiw")
    os.makedirs(d, exist_ok=True)
    return d


STATE_SHEET = "STATE"
_CHUNK = 30000

def _settings_sheet(wb, cfg):
    """SETTINGS: read-only statement of everything configured in-app — the
    inputs that have no editable cells elsewhere in this workbook. A reviewer
    sees every entered value; the import deliberately ignores this sheet."""
    a = cfg.get("assumptions") or {}
    ws = wb.create_sheet("SETTINGS")
    ws.append(["SETTINGS — configured in-app (read-only)"])
    ws["A1"].font = Font(bold=True, size=12)
    ws.append(["Edits on this sheet are NOT imported. These inputs are configured in the app; "
               "they are stated here so the workbook is a complete record of the engagement."])
    ws.append([])
    def sec(title): ws.append([title]); ws[f"A{ws.max_row}"].font = Font(bold=True)
    def row(label, val, unit=""):
        ws.append([f"  {label}", "" if val is None else val, unit])
        if isinstance(val, (int, float)) and abs(val) >= 1000:
            ws.cell(row=ws.max_row, column=2).number_format = "#,##0"
    sec("Treasury (funding waterfall)")
    row("Cash floor (% of deposits)", a.get("cash_target_pct_deposits"), "rate")
    row("Yield on cash", a.get("cash_yield"), "annual rate")
    row("Yield on securities", a.get("securities_yield"), "annual rate")
    row("Borrowing rate", a.get("borrow_rate_ann"), "annual rate")
    sec("Overhead & other balance sheet")
    row("Corporate overhead", a.get("overhead_q"), "$/quarter")
    row("Overhead growth", a.get("overhead_growth_q"), "rate/qtr")
    row("Premises & equipment", a.get("premises_equipment"), "$")
    row("Premises depreciation", a.get("premises_depreciation_annual"), "$/year")
    row("Intangibles", a.get("intangibles"), "$")
    row("Other assets", a.get("other_assets"), "$")
    row("Other liabilities", a.get("other_liabilities"), "$")
    sec("Scheduled borrowings")
    for i, b in enumerate(a.get("scheduled_borrowings") or [], 1):
        row(b.get("name") or f"Borrowing {i}",
            f"draws Q{b.get('quarter', b.get('start_q'))} \u00b7 {float(b.get('amount') or 0):,.0f} \u00b7 "
            f"{b.get('term_q')}q amortizing at {b.get('rate_ann')}", "")
    if not (a.get("scheduled_borrowings") or []): row("(none)", "")
    sec("Pre-opening expenses")
    _po_ex = (cfg.get("pre_opening") or {}).get("expenses") or []
    for e in _po_ex:
        row(e.get("category") or e.get("label") or "expense",
            e.get("total") if e.get("total") is not None else e.get("amount"), "$")
    if _po_ex:
        row("Total pre-opening burn", sum(float(e.get("total", e.get("amount", 0)) or 0) for e in _po_ex), "$ (reduces Day-1 capital)")
    else:
        row("(none)", "")
    sec("Securities books")
    for p in (a.get("securities_afs") or []): row(f"AFS — {p.get('name')}", p.get("opening"), f"yield {p.get('yield_ann')}")
    for p in (a.get("securities_htm") or []): row(f"HTM — {p.get('name')}", p.get("opening"), f"yield {p.get('yield_ann')}")
    if not ((a.get("securities_afs") or []) + (a.get("securities_htm") or [])): row("(none)", "")
    row("AOCI sensitivity", a.get("aoci_sensitivity_annual"), "% of AFS/yr")
    sec("NIE detail")
    nd = a.get("nie_detail") or {}
    if nd:
        if nd.get("fte_by_year") is not None:
            row("FTE by year (Y1/Y2/Y3)", " / ".join(str(x) for x in nd.get("fte_by_year") or []), "headcount")
        if nd.get("loaded_comp_annual") is not None:
            row("Loaded comp per FTE", nd.get("loaded_comp_annual"), "$/year")
        for cat in (nd.get("categories") or []):
            row(cat.get("name") or "category",
                cat.get("per_quarter") if cat.get("per_quarter") is not None else cat.get("amount_q"), "$/quarter")
        if nd.get("other_gross_up_rate") is not None:
            row("Other gross-up rate", nd.get("other_gross_up_rate"), "rate")
        if nd.get("fdic_bp_ann") is not None:
            row("FDIC assessment rate", nd.get("fdic_bp_ann"), "bp/yr (default 5.0)")
        if nd.get("occ_bp_ann") is not None:
            row("OCC assessment rate", nd.get("occ_bp_ann"), "bp/yr (default 1.5)")
    else:
        row("(not active)", "")
    sec("Credit regime (ASC 326)")
    _crx = a.get("credit_regime")
    if _crx:
        row("Module", "active", "ACL vocabulary + day-one provision decomposition")
        _lp = a.get("lending_products") or []
        row("Amortized cost + CECL ACL", sum(1 for p in _lp if p.get("measurement") != "fair_value"), "per-product election")
        row("Fair value option", sum(1 for p in _lp if p.get("measurement") == "fair_value"), "ASC 825, irrevocable at origination")
        row("Incurred loss", "not offered", "CECL mandatory for HFI amortized-cost loans")
    else:
        row("(not active)", "", "ALLL vocabulary; single provision line")
    sec("Tax detail (NOL \u2192 DTA)")
    _tdx = a.get("tax_detail")
    if _tdx:
        row("Module", "active", "ASC 740 current/deferred split")
        row("NOL utilization limit", _tdx.get("nol_utilization_limit_pct", 0.8), "of taxable income (IRC 172)")
        row("Valuation allowance mode", _tdx.get("va_mode", "auto"), "")
        row("CET1 treatment", "full deduction of net NOL-DTA", "12 CFR 3.22(a)")
    else:
        row("(not active)", "", "100% NOL shield, no DTA booked")
    sec("Fee modules")
    fm = a.get("fee_modules") or {}
    for k, v in fm.items():
        if v: row(k, str(v)[:120], "")
    if not any(fm.values()) if fm else True: row("(none active)", "")
    sec("SOFR rate path (Q1..Q12)")
    row("Path", ", ".join(str(x) for x in (a.get("rate_path_q") or [])), "annual")
    row("Longer run", a.get("rate_path_longer_run"), "annual")
    sec("Stress parameters")
    for k, v in (cfg.get("stress_params") or {}).items(): row(k, v, "")
    ws.column_dimensions["A"].width = 44
    ws.column_dimensions["B"].width = 40


def _embed_state(wb, cfg):
    ws = wb.create_sheet(STATE_SHEET)
    ws.sheet_state = "veryHidden"
    blob = json.dumps(cfg, sort_keys=True)
    for i in range(0, len(blob), _CHUNK):
        ws.cell(row=i // _CHUNK + 1, column=1, value=blob[i:i + _CHUNK])

def _read_embedded_state(wb):
    if STATE_SHEET not in wb.sheetnames:
        return None
    parts = []
    for r in wb[STATE_SHEET].iter_rows(min_col=1, max_col=1):
        if r[0].value is None:
            break
        parts.append(str(r[0].value))
    try:
        return json.loads("".join(parts))
    except Exception:
        return None

def persist_snapshot(cfg, gh):
    with open(os.path.join(_snap_dir(), gh + ".json"), "w", encoding="utf-8") as fh:
        json.dump(cfg, fh)


def load_snapshot(gh):
    p = os.path.join(_snap_dir(), gh + ".json")
    if not os.path.exists(p):
        return None
    return json.load(open(p, encoding="utf-8"))


# ---------------------------------------------------------------- diff import
CONTROL_PATHS = {
    "Institution": ("proposed_bank",),
    "Legal name": ("client_legal_name",),
    "Home state / market": ("hq",),
    "Charter type": ("charter_profile", "charter_type"),
    "Target opening": ("charter_profile", "target_opening"),
    "Pre-opening period (months)": ("charter_profile", "pre_open_months"),
    "CBLR election": ("charter_profile", "cblr_election"),
    "Initial capital ($000s)": ("target_state", "initial_capital"),
    "Initial capital ($)": ("target_state", "initial_capital"),
    "Scenario name": ("scenario_name",),
}
import re as _re
_RAISE_ROW = _re.compile(r"^Staged raise (\d+) \u2014 (quarter|amount)", 0)


def _raise_row_match(label):
    m = _re.match(r"^Staged raise (\d+) — (quarter|amount)", str(label))
    return (int(m.group(1)) - 1, m.group(2)) if m else None
FAM_ARR = {"lending": "lending_products", "deposit": "deposit_products"}


def _get(cfg, path):
    o = cfg
    for k in path:
        if o is None:
            return None
        o = o.get(k) if isinstance(o, dict) else None
    return o


def _set(cfg, path, val):
    o = cfg
    for k in path[:-1]:
        o = o.setdefault(k, {})
    o[path[-1]] = val


def _path_root(cfg, parts):
    """Resolve where a dotted array-sheet key is rooted. 'pre_opening.expenses...'
    is top-level; everything else (securities_afs, obs_exposures, scheduled_borrowings,
    capital_raises, fee_modules) lives under cfg['assumptions']. Returns (container,
    remaining_parts) where remaining_parts still leads with the array/dict name."""
    if parts and parts[0] == "pre_opening":
        return cfg, parts
    return cfg.setdefault("assumptions", {}), parts


def _walk(container, parts, create=False):
    """Walk a dotted path (with integer indices for list positions) to the parent of
    the final leaf. Returns (parent, leaf_key) or (None, None) if absent and not creating."""
    o = container
    for k in parts[:-1]:
        if k.isdigit():
            idx = int(k)
            if not isinstance(o, list):
                return None, None
            if idx >= len(o):
                if not create:
                    return None, None
                while len(o) <= idx:
                    o.append({})
            o = o[idx]
        else:
            if not isinstance(o, dict):
                return None, None
            if k not in o or o[k] is None:
                if not create:
                    return None, None
                o[k] = {}
            o = o[k]
    return o, parts[-1]


def _resolve_path(cfg, parts):
    container, p = _path_root(cfg, parts)
    parent, leaf = _walk(container, p, create=False)
    if parent is None:
        return None
    if leaf.isdigit():
        i = int(leaf)
        return parent[i] if isinstance(parent, list) and i < len(parent) else None
    return parent.get(leaf) if isinstance(parent, dict) else None


def _apply_path(cfg, parts, val):
    container, p = _path_root(cfg, parts)
    parent, leaf = _walk(container, p, create=True)
    if parent is None:
        return
    if leaf.isdigit():
        i = int(leaf)
        if isinstance(parent, list):
            while len(parent) <= i:
                parent.append({})
            parent[i] = val
    elif isinstance(parent, dict):
        parent[leaf] = val


def _num_eq(a, b):
    """Equal within Excel's round-trip precision. openpyxl/Excel carry ~15 sig digits, so a
    value like 33333.333333333336 reads back as 33333.33333333334 — an absolute 1e-12 threshold
    would call that a spurious edit. Relative tolerance absorbs the representation loss."""
    if isinstance(a, (int, float)) and isinstance(b, (int, float)):
        return abs(a - b) <= 1e-9 + 1e-9 * max(abs(a), abs(b))
    return (a or None) == (b or None)


def diff_import(data, current_cfg):
    """Filled FIW -> (merged cfg, edits report). Fail-closed on unknown
    generation state; only cells that DIFFER from the snapshot are applied,
    so untouched defaults keep their provenance and in-app edits made since
    generation survive unless the workbook explicitly changed that cell."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(data), data_only=True)
    if "README" not in wb.sheetnames:
        raise ValueError("not a Foundry Input Workbook (no README sheet)")
    gh = None
    for r in wb["README"].iter_rows():
        if r[0].value == "Generation hash":
            gh = r[1].value
    if not gh:
        raise ValueError("no generation hash on README — cannot diff-import")
    snap = _read_embedded_state(wb)
    base_src = "embedded"
    if snap is None:                      # legacy workbook (pre self-contained format)
        snap = load_snapshot(str(gh))
        base_src = "workspace snapshot"
    if snap is None:
        raise ValueError(f"generation state {gh} not found — this workbook predates the "
                          "self-contained format and its workspace snapshot is gone (older run, "
                          "another instance, or a redeploy). Regenerate the input workbook from "
                          "the current configuration and reapply your edits; nothing was changed. "
                          "Workbooks generated from now on carry their own state and never hit this.")
    # THE WORKBOOK IS THE DOCUMENT: import rebases onto the workbook's own
    # generation state (the bank it describes), with the human's edits on top.
    # The open session never silently supplies the base — uploading a workbook
    # from another engagement reconstitutes THAT bank, and the report says so.
    merged = json.loads(json.dumps(snap))
    edits = []
    warnings = []
    session_differed = json.dumps(current_cfg, sort_keys=True) != json.dumps(snap, sort_keys=True)

    if "CONTROL" in wb.sheetnames:
        for r in wb["CONTROL"].iter_rows(min_row=2):
            label, val = r[0].value, r[1].value
            rr = _raise_row_match(label)
            if rr is not None:
                idx, field = rr
                if field == "amount" and "$000s" in str(label):
                    val = wbunits.from_workbook(val, "$000s")
                arr = (snap["assumptions"].get("capital_raises") or [])
                old_v = arr[idx].get(field) if idx < len(arr) else None
                new_v = int(val) if field == "quarter" and val not in ("", None) else (
                          float(val) if val not in ("", None) else None)
                if new_v != old_v and new_v is not None:
                    tgt = merged["assumptions"].setdefault("capital_raises", [])
                    while len(tgt) <= idx:
                        tgt.append({"quarter": 1, "amount": 0})
                    tgt[idx][field] = new_v
                    edits.append({"key": f"capital_raises.{idx}.{field}",
                                   "from": old_v, "to": new_v})
                continue
            path = CONTROL_PATHS.get(label)
            if not path:
                continue
            if label in ("Initial capital ($000s)",) and "$000s" in str(label):
                val = wbunits.from_workbook(val, "$000s")
            old = _get(snap, path)
            if label == "CBLR election":
                val = str(val).strip().lower() in ("yes", "y", "true", "1")
                if bool(val) == bool(old):
                    continue
            if val != old and not (val in ("", None) and old in ("", None)):
                edits.append({"key": ".".join(str(x) for x in path), "from": old, "to": val})
                _set(merged, path, val)

    # identity follow-through: renaming the bank renames a scenario label that
    # carried the old bank's name — derived, and logged like any other edit.
    _old_pb = str(snap.get("proposed_bank") or "")
    _new_pb = str(merged.get("proposed_bank") or "")
    _sn = str(merged.get("scenario_name") or "")
    if _new_pb and _new_pb != _old_pb and _old_pb and _sn.startswith(_old_pb):
        _renamed = _new_pb + _sn[len(_old_pb):]
        edits.append({"key": "scenario_name", "from": _sn, "to": _renamed,
                       "note": "derived — scenario label followed the bank rename"})
        merged["scenario_name"] = _renamed

    def _norm_num(v):
        """Normalize a cell value for comparison: numbers compare as floats (so 0.049,
        4.9%, and float-epsilon representations don't create spurious 'changed')."""
        if v in ("", None):
            return None
        if isinstance(v, (int, float)):
            return float(v)
        try:
            return float(v)
        except (TypeError, ValueError):
            return v

    # Deposit/loan sheets carry a synthetic contextual rate cell (__RATE__) whose meaning is
    # set by rate_type, plus an informational baseline (__RATE_BASELINE__, a 'fact' row).
    # We collect each product's rows first (so rate_type is known before routing __RATE__),
    # then apply — routing the rate into rate_paid_ann/yield_ann OR index_spread and CLEARING
    # the inactive field, per the discriminated-union contract.
    for sheet, fam in (("ASSM_LOANS", "lending"), ("ASSM_DEPOSITS", "deposit")):
        if sheet not in wb.sheetnames:
            continue
        arr = snap["assumptions"].get(FAM_ARR[fam], [])
        # gather rows by product index
        by_idx = {}
        for r in wb[sheet].iter_rows(min_row=2):
            key, val, units = r[0].value, r[3].value, (r[4].value or "")
            if not key or "fact" in str(units):
                continue
            _, idx, field = key.split(".", 2)
            by_idx.setdefault(int(idx), {})[field] = wbunits.from_workbook(val, units)

        for idx, cells in by_idx.items():
            if idx >= len(arr):
                continue
            base = arr[idx]
            # --- identity heuristic (NOT proof; see PROTOCOL_GAPS): does this row still
            # correspond to the exported product? Require the call-report line AND opening
            # balance at this index to match the snapshot. If not, we cannot trust
            # value-unchanged grading and fall back to a broad notice.
            _line_lbl = LINE_LABELS.get(base.get("call_report_line"), base.get("call_report_line"))
            _id_trusted = (cells.get("call_report_line") in (None, _line_lbl)
                           and (("opening_balance" not in cells)
                                or _norm_num(cells.get("opening_balance")) == _norm_num(base.get("opening_balance"))))

            # rate_type after edits (falls back to snapshot)
            new_rt = cells.get("rate_type") or base.get("rate_type")
            old_rt = base.get("rate_type")

            tgt = merged["assumptions"].setdefault(FAM_ARR[fam], [])
            if idx >= len(tgt):
                continue

            for field, val in cells.items():
                if field in ("__RATE_BASELINE__",):
                    continue
                if field == "__RATE__":
                    # route the one contextual value by the (possibly edited) rate_type,
                    # and CLEAR the inactive field — but ONLY when something actually changed
                    # (a clean re-import must be a no-op, or the hash drifts). "Changed" means
                    # the rate value differs from the active field OR the rate_type flipped.
                    rate_val = _norm_num(val)
                    if new_rt == "float":
                        old = base.get("index_spread")
                        fld, inactive = "index_spread", ("rate_paid_ann", "yield_ann")
                    else:
                        fld = "rate_paid_ann" if fam == "deposit" else "yield_ann"
                        old = base.get(fld)
                        inactive = ("index_spread",)
                    _rate_changed = not _num_eq(_norm_num(old), rate_val)
                    _type_flipped = new_rt != old_rt
                    if _rate_changed or _type_flipped:
                        tgt[idx][fld] = rate_val
                        for _ifld in inactive:
                            tgt[idx].pop(_ifld, None)
                        if _rate_changed:
                            edits.append({"key": f"{fam}.{idx}.{fld}",
                                           "from": old, "to": rate_val})
                    continue
                # ordinary field
                if "." in field:
                    a, b = field.split(".", 1)
                    old = (base.get(a) or {}).get(b)
                else:
                    old = base.get(field)
                newv = _norm_num(val) if field not in ("rate_type", "measurement",
                                                         "call_report_line", "name") else (
                        None if val in ("", None) else val)
                changed = not _num_eq(old, newv)
                if changed:
                    if "." in field:
                        a, b = field.split(".", 1)
                        tgt[idx].setdefault(a, {})[b] = newv
                    else:
                        tgt[idx][field] = newv
                    edits.append({"key": f"{fam}.{idx}.{field}", "from": old, "to": newv})

            # --- rate_type-change warning (graded, identity-aware) ---
            if new_rt != old_rt:
                rate_cell = _norm_num(cells.get("__RATE__"))
                old_rate = (base.get("index_spread") if old_rt == "float"
                            else (base.get("rate_paid_ann") if fam == "deposit" else base.get("yield_ann")))
                pname = base.get("name", f"{fam}[{idx}]")
                if not _id_trusted:
                    warnings.append({"severity": "review",
                        "message": f"row identity uncertain for {sheet} index {idx}; product "
                                    "ordering may have changed — review ALL contextual rate values"})
                elif rate_cell is not None and _norm_num(old_rate) == rate_cell:
                    warnings.append({"severity": "high",
                        "message": f"'{pname}': rate type changed {old_rt}->{new_rt} but the value "
                                    f"{rate_cell} was left unchanged; it will now be read as "
                                    f"{'a spread over SOFR' if new_rt=='float' else 'an absolute rate'}, "
                                    "not the prior meaning — confirm this is intended"})
                else:
                    warnings.append({"severity": "notice",
                        "message": f"'{pname}': rate type changed {old_rt}->{new_rt} (value also changed)"})

    # Flat assumption-array sheets (securities, OBS, scheduled borrowings, raises,
    # pre-opening expenses) and the fee-modules sheet. Same diff mechanic: read the
    # column-A machine key, compare the Value cell to the snapshot, apply only real
    # changes, skip 'fact' rows. Keys are dotted paths into the config.
    def _coerce(old, val):
        if val in ("", None):
            return None
        if isinstance(old, (int, float)) and isinstance(val, str):
            try:
                return float(val)
            except ValueError:
                return val
        return val

    ARRAY_SHEETS = {
        "ASSM_SEC_AFS": ("securities_afs", 3),
        "ASSM_SEC_HTM": ("securities_htm", 3),
        "ASSM_OBS": ("obs_exposures", 3),
        "ASSM_BORROWINGS": ("scheduled_borrowings", 3),
        "ASSM_RAISES": ("capital_raises", 3),
        "ASSM_PREOPEN": ("pre_opening.expenses", 3),
        "ASSM_FEES": ("fee_modules", 3),
        "ASSM_NIE": ("nie_detail", 3),
    }
    _dropped_rows = []          # hand-added rows with content but no valid machine key
    for sheet, (_root, valcol) in ARRAY_SHEETS.items():
        if sheet not in wb.sheetnames:
            continue
        for r in wb[sheet].iter_rows(min_row=2):
            key = r[0].value
            val = r[valcol].value
            units = (r[4].value or "")
            if not key or "fact" in str(units):
                # A row the user filled in (Item/Field/Value present) but with no machine key in
                # column A is a NEW row added in Excel. diff_import keys off column A, so such a row
                # is otherwise dropped SILENTLY. Detect it and warn loudly instead. (Stopgap: adding
                # new instruments via the workbook is not yet supported — do it in the app's + Add
                # controls, which create the machine key. Remove this branch when add-rows lands.)
                _item = r[1].value; _field = r[2].value
                if not key and (_item not in (None, "") or val not in (None, "")):
                    _dropped_rows.append({"sheet": sheet, "item": _item, "field": _field, "value": val})
                continue
            # key is a dotted path, e.g. securities_afs.0.opening,
            # fee_modules.payments.1.fee_per_tx, pre_opening.expenses.2.total
            parts = str(key).split(".")
            val = wbunits.from_workbook(val, units)
            old = _resolve_path(snap, parts)
            newv = _coerce(old, val)
            changed = not _num_eq(old, newv)
            if changed and newv is not None:
                _apply_path(merged, parts, newv)
                edits.append({"key": str(key), "from": old, "to": newv})

    if _dropped_rows:
        # one high-severity warning, listing the affected sheets + row count, so the user is never
        # surprised by a silently-lost instrument again.
        _by_sheet = {}
        for d in _dropped_rows:
            _by_sheet.setdefault(d["sheet"], set()).add(str(d.get("item") or "(unnamed)"))
        _detail = "; ".join(f"{s}: {', '.join(sorted(names))}" for s, names in _by_sheet.items())
        warnings.append({
            "severity": "high",
            "code": "added_rows_ignored",
            "message": ("Rows added directly in the workbook were NOT imported and had no effect: "
                        + _detail + ". The input workbook edits existing line items; it cannot add new "
                        "ones (they carry no internal key). Add new instruments in the app using the "
                        "'+ Add' controls, then re-download the workbook if you want them in it."),
            "rows": _dropped_rows,
        })


    rep = {"generation_hash": str(gh), "edits": edits, "edit_count": len(edits),
            "base": base_src, "warnings": warnings,
            "warning_count": len(warnings),
            "high_severity_warnings": [w for w in warnings if w.get("severity") == "high"]}
    if session_differed:
        rep["session_note"] = ("the open session differed from this workbook's engagement — "
                                "the workbook's state now governs (the session's unsaved "
                                "differences were replaced, not merged)")
    return merged, rep
