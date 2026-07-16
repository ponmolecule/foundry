"""Build the v3.1 default configuration from Patrick's ASSM_BS values.

Run where the source workbook is available; the emitted JSON is committed so
deployments never need the workbook. Direction of truth: his cells are the
assumption values; the app's extra fields (fees, OTS block, per-product opex,
measurement, rate structure) remain present at neutral values — the superset
survives, nothing is invented.

Conversions (each recorded in the emitted config's conversion_notes):
- units: workbook $000s -> config raw dollars (x1000)
- loans: monthly originations x3 -> originations_q; annual prepay /4 -> runoff_q
- deposits: his flat $-adds cannot be a %-growth scalar from a zero base, so the
  generator computes his monthly path (B_m = B_{m-1} * (1 - runoff/12) + add),
  samples quarterly EOP, sets opening = Q1 EOP and pins growth via per-quarter
  overrides (the bridge pattern) — his path exactly, in our machinery
- CD/brokered maturities: unsupported (M11); carried as a caveat note
- treasury/premises/borrowing prices -> globals; other-assets-% -> caveat
"""
import json
import sys

from openpyxl import load_workbook

SRC = sys.argv[1] if len(sys.argv) > 1 else \
    "/mnt/project/Klaros_Bank_Charter_Financial_Model_v1_0_Patrick.xlsx"
TEMPLATE = "foundry/fixtures/parity/configs/pf_a_base.json"
OUT = "foundry/fixtures/patrick_default_v31.json"
TARGETS = "foundry/fixtures/patrick_default_targets.json"

K = 1000.0
LOANS = [  # (name, row0, call_report_line key)
    ("Consumer Installment", 7, "loanConsumer"),
    ("Credit Card", 13, "loanCreditCard"),
    ("Small Business / C&I", 19, "loanCommercial"),
    ("Residential Mortgage", 25, "loanMortgage"),
    ("Commercial Real Estate", 31, "loanOther"),
]
DEPOSITS = [  # (name, row0, line key)
    ("Retail Demand", 38, "depDDA"),
    ("MMDA / Savings", 43, "depSavings"),
    ("Time Deposits / CDs", 48, "depTime"),
    ("Brokered", 53, "depTime"),
    ("Sweep / Program", 58, "depDDA"),
    ("Institutional", 63, "depDDA"),
]


def monthly_deposit_path(add_k, runoff_ann, months=36):
    b, out = 0.0, []
    for _ in range(months):
        b = b * (1 - runoff_ann / 12.0) + add_k * K
        out.append(b)
    return out


def q_eop(path):
    return [path[3 * q - 1] for q in range(1, 13)]


def pin_growth(qpath):
    ov = {"1": 0.0}
    for t in range(2, 13):
        prev = qpath[t - 2]
        ov[str(t)] = (qpath[t - 1] / prev - 1.0) if prev else 0.0
    return ov


def main():
    ws = load_workbook(SRC, data_only=True)["ASSM_BS"]
    v = lambda r: ws.cell(r, 3).value or 0.0

    cfg = json.load(open(TEMPLATE, encoding="utf-8"))
    a = cfg["assumptions"]
    notes = []

    lending = []
    for name, r0, line in LOANS:
        lending.append({
            "name": name, "call_report_line": line,
            "opening_balance": v(r0) * K,
            "originations_q": v(r0 + 1) * K * 3.0,
            "orig_growth_q": 0.0,
            "runoff_q": v(r0 + 3) / 4.0,
            "yield_ann": v(r0 + 2), "rate_type": "fixed",
            "charge_off_ann": v(r0 + 4), "provision_rate_ann": None,
            "reserve_rate_pct_bal": v(r0 + 5),
            "measurement": "amortized",
            "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0,
        })
    notes.append("loans: monthly originations x3 -> quarterly; annual prepay /4 -> runoff_q; "
                  "opening balances are Day-1 zero per source")

    deposits, dep_targets = [], {}
    for name, r0, line in DEPOSITS:
        add, cost, runoff = v(r0 + 1), v(r0 + 2), v(r0 + 4)
        path = monthly_deposit_path(add, runoff)
        qp = q_eop(path)
        dep_targets[name] = qp
        deposits.append({
            "name": name, "call_report_line": line,
            "opening_balance": qp[0], "growth_q": 0.0,
            "runoff_q": 0.0,  # runoff is inside the pinned path
            "rate_paid_ann": cost, "rate_type": "fixed",
            "fee_yield_ann": 0.0, "opex_pct_ann": 0.0, "opex_fixed_m": 0.0,
            "overrides": {"growth_q": pin_growth(qp)},
        })
        mat = v(r0 + 3)
        if mat:
            notes.append(f"{name}: average maturity {mat:.0f} months is not modeled "
                          "(maturity-ladder liability is a phase-2 mechanic); cost and "
                          "path carried, term structure disclosed as unsupported")
    notes.append("deposits: flat $-adds converted by computing the source's monthly path "
                  "(B*(1-runoff/12)+add), sampling quarterly EOP, opening = Q1 EOP, growth "
                  "pinned per quarter via overrides — path exact, not approximated")

    a["lending_products"] = lending
    a["deposit_products"] = deposits
    a["securities_yield"] = v(70)       # AFS average yield
    a["cash_yield"] = v(76)             # IB deposits at other banks rate
    a["borrow_rate_ann"] = v(84)        # FHLB advance rate
    a["premises_equipment"] = v(81) * K
    notes.append("securities yield = AFS cell; HTM yield differs (4.2%) — single-book "
                  "engine, AFS value used, HTM noted; other-assets-% (2% of assets) "
                  "unsupported as a formula, template constant kept")

    cfg["proposed_bank"] = "De Novo Bank"
    cfg["client_legal_name"] = "De Novo Bank (in organization)"
    cfg["scenario_name"] = "De Novo Bank — baseline assumptions (source: engagement workbook)"
    cfg["engagement_id"] = "ENG-NEW"
    cfg["parity_expectation"] = None
    cfg["conversion_notes"] = notes

    json.dump(cfg, open(OUT, "w", encoding="utf-8"), indent=1)
    json.dump(dep_targets, open(TARGETS, "w", encoding="utf-8"), indent=1)
    print(f"wrote {OUT} ({len(lending)} loans, {len(deposits)} deposits) + targets")


if __name__ == "__main__":
    main()
